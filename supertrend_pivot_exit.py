"""
Supertrend Intraday Backtesting Strategy — v3.1 (ChoCh Exit Added)
─────────────────────────────────────────────────────────────────────────────
DATA SOURCE : CSV files  OR  yfinance  (toggle DATA_SOURCE in CONFIG)

  CSV MODE    (DATA_SOURCE = "csv")
    BASE_CSV  — your 1m (or any TF) OHLC file
    HTF_CSV   — your HTF OHLC file  (set None to disable HTF filter)

  YFINANCE MODE  (DATA_SOURCE = "yfinance")
    YF_TICKER        — Yahoo Finance ticker  e.g. "^NSEI", "RELIANCE.NS"
    YF_BASE_INTERVAL — base TF  e.g. "1m", "2m", "5m"
    YF_HTF_INTERVAL  — HTF      e.g. "15m", "30m", "1h"  (None = disabled)
    YF_PERIOD        — lookback  e.g. "5d", "7d", "30d", "60d"
                       Note: 1m data is limited to the last 7 days by Yahoo
    YF_START / YF_END — use a fixed date range instead of YF_PERIOD
                        set YF_START = None to use YF_PERIOD

CSV FORMAT  : Standard columns (case-insensitive, auto-detected)
  datetime, open, high, low, close, volume
  DateTime column accepts any standard format (auto-parsed)

FILTERS:
  1. HTF Filter        — only trade in direction of HTF Supertrend
                         (auto-disabled if HTF_CSV = None  or  YF_HTF_INTERVAL = None)
  2. Trade Window      — only enter during profitable time slots
  3. Confirm Candles   — wait N candles after ST flip before entering
  4. Min Gap           — skip if price too close to ST line (choppy)
  5. Volume            — require volume > N × 20-bar average
  6. Max Trades/Day    — hard cap on daily entries

ENTRIES:
  Long  : ST below price  →  enter at close
  Short : ST above price  →  enter at close

EXITS (whichever fires first):
  Hard SL      →  fixed % loss from entry
  ST SL        →  price crosses ST line (base TF or HTF ST depending on config)
  ChoCh Exit   →  Change of Character:
                    Long  → price creates a new Higher High after entry,
                            then breaks BELOW that recorded HH → exit long
                    Short → price creates a new Lower Low after entry,
                            then breaks ABOVE that recorded LL → exit short
  EOD          →  force exit at 15:15 IST

EXIT IMPROVEMENTS:
  EXIT_USE_HTF_ST   — use the wider HTF ST line as exit instead of tight 1m ST
  CLOSE_BASED_EXIT  — exit only when candle CLOSES beyond ST line (reduces whipsaws)

ChoCh EXIT CONFIG:
  CHOCH_EXIT_ENABLED — master toggle for ChoCh exit (True / False)
  CHOCH_LOOKBACK     — number of candles to look back when identifying
                       the swing High (for longs) or swing Low (for shorts)
                       that forms the "previous Higher High / Lower Low"
                       A higher value = fewer but more significant swing points
                       Recommended: 3–10 for 1m data

OUTPUTS:
  Console  : trade log + per-day P&L + time-slot P&L + holding time + summary
  Excel    : sheets — Supertrend Data, Trade Log, Daily P&L,
             Time Slot P&L, Holding Time Summary, Holding By Time Slot, Summary
  HTML     : interactive candlestick + ST chart

Install: pip install pandas plotly openpyxl pytz numpy yfinance
"""

import pandas as pd
import numpy as np
import plotly.graph_objects as go
import pytz
import os
from datetime import time as dtime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

# ── DATA SOURCE TOGGLE ─────────────────────────────────────────────────────────
DATA_SOURCE = "yfinance"          # "csv"      → load from local CSV files
                                  # "yfinance" → download live data from Yahoo Finance

# ── CSV settings  (used when DATA_SOURCE = "csv") ─────────────────────────────
BASE_CSV = "NIFTY 50_minute.csv"
HTF_CSV  = "NIFTY 50_15minute.csv"   # set to None to disable HTF filter

# ── yfinance settings  (used when DATA_SOURCE = "yfinance") ───────────────────
YF_TICKER        = "^NSEI"    # Yahoo Finance ticker symbol
YF_BASE_INTERVAL = "1m"       # base TF interval: "1m","2m","5m","15m","30m","1h"
YF_HTF_INTERVAL  = "15m"      # HTF interval: "5m","15m","30m","1h"  or None
YF_PERIOD        = "5d"       # lookback period: "1d","5d","7d","30d","60d","max"
YF_START         = None       # e.g. "2024-01-01" — set both to use a date range
YF_END           = None       # e.g. "2024-06-30" — if YF_START=None, YF_PERIOD used

# ── Label shown in all outputs ─────────────────────────────────────────────────
SYMBOL   = "NIFTY"

ST_PERIOD         = 14
ST_MULTIPLIER     = 4.0
HTF_ST_PERIOD     = 10
HTF_ST_MULTIPLIER = 3.0

HARD_SL_ENABLED = True
HARD_SL_PCT     = 0.25

EXIT_USE_HTF_ST  = False
CLOSE_BASED_EXIT = True

# ── ChoCh Exit Settings ────────────────────────────────────────────────────────
# Change of Character exit:
#   Long  → tracks the highest High seen AFTER entry.
#            If a subsequent candle's Close breaks BELOW that peak → exit.
#   Short → tracks the lowest Low seen AFTER entry.
#            If a subsequent candle's Close breaks ABOVE that trough → exit.
#
# CHOCH_LOOKBACK: minimum number of candles that must pass before a new
#   swing extreme is "confirmed" as the reference HH / LL.
#   Set to 1 to react on every new extreme immediately.
#   Set to 3–5 to wait for a minor pullback before locking in the swing point.
CHOCH_EXIT_ENABLED = True
CHOCH_LOOKBACK     = 5        # candles; 1 = immediate, 3-5 = swing-confirmed

TRADE_WINDOW_ENABLED = True
TRADE_WINDOWS = [
    (dtime(9, 30), dtime(11, 30)),
    (dtime(14,  0), dtime(14, 45)),
]

CONFIRM_CANDLES = 10

MIN_GAP_ENABLED = True
MIN_GAP_PCT     = 0.20

VOLUME_FILTER_ENABLED = False
VOLUME_MULTIPLIER     = 1.5
VOLUME_LOOKBACK       = 25

MAX_TRADES_PER_DAY_ENABLED = True
MAX_TRADES_PER_DAY         = 5

MARKET_OPEN  = dtime(9, 15)
MARKET_CLOSE = dtime(15, 15)
IST          = pytz.timezone("Asia/Kolkata")

CHART_DAYS = 5

# ═══════════════════════════════════════════════════════════════════════════════
#  END OF CONFIG
# ═══════════════════════════════════════════════════════════════════════════════


def fmt(ts):
    return ts.strftime("%Y-%m-%d %H:%M") if pd.notna(ts) else "—"


def mins_to_hhmm(m):
    """Convert integer minutes to HH:MM string."""
    m = int(round(m))
    return f"{m // 60}H:{m % 60:02d}M"


# ───────────────────────────────────────────────────────────────────────────────
#  SUPERTREND ENGINE
# ───────────────────────────────────────────────────────────────────────────────
def compute_supertrend(df, period, multiplier):
    high  = df["High"].values.astype(float)
    low   = df["Low"].values.astype(float)
    close = df["Close"].values.astype(float)
    n     = len(df)

    tr = np.maximum(high[1:] - low[1:],
         np.maximum(np.abs(high[1:] - close[:-1]),
                    np.abs(low[1:]  - close[:-1])))

    atr = np.full(n, np.nan)
    if n > period:
        atr[period] = tr[:period].mean()
        for j in range(period + 1, n):
            atr[j] = (atr[j-1] * (period - 1) + tr[j-1]) / period

    hl2         = (high + low) / 2.0
    upper_basic = hl2 + multiplier * atr
    lower_basic = hl2 - multiplier * atr

    upper = np.full(n, np.nan)
    lower = np.full(n, np.nan)
    bull  = np.full(n, True, dtype=bool)

    for j in range(period, n):
        if np.isnan(upper[j-1]):
            upper[j] = upper_basic[j];  lower[j] = lower_basic[j]
            bull[j]  = close[j] >= lower[j];  continue
        upper[j] = (upper_basic[j]
                    if upper_basic[j] < upper[j-1] or close[j-1] > upper[j-1]
                    else upper[j-1])
        lower[j] = (lower_basic[j]
                    if lower_basic[j] > lower[j-1] or close[j-1] < lower[j-1]
                    else lower[j-1])
        bull[j] = (close[j] >= lower[j]) if bull[j-1] else (close[j] > upper[j])

    df = df.copy()
    df["ST_val"]       = np.where(bull, lower, upper)
    df["ST_bull"]      = bull
    df["ST_direction"] = np.where(bull, "Bullish", "Bearish")
    return df


# ───────────────────────────────────────────────────────────────────────────────
#  CSV LOADER
# ───────────────────────────────────────────────────────────────────────────────
_DT_CANDS = ["datetime","date","time","timestamp","date/time","date time"]
_O_CANDS  = ["open","o"]
_H_CANDS  = ["high","h"]
_L_CANDS  = ["low","l"]
_C_CANDS  = ["close","ltp","last","c"]
_V_CANDS  = ["volume","vol","v"]


def _find_col(columns, candidates):
    lmap = {c.lower().strip(): c for c in columns}
    for cand in candidates:
        if cand in lmap:
            return lmap[cand]
    return None


def load_csv(path, label="CSV"):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"\n  ❌  File not found : '{path}'\n"
            f"      ↳ Check the BASE_CSV / HTF_CSV paths in the CONFIG section.\n"
        )

    print(f"\n  Loading {label}: {path} …")
    df = pd.read_csv(path)
    print(f"    Shape: {df.shape}  |  Columns: {list(df.columns)}")

    dt_col = _find_col(df.columns, _DT_CANDS)
    o_col  = _find_col(df.columns, _O_CANDS)
    h_col  = _find_col(df.columns, _H_CANDS)
    l_col  = _find_col(df.columns, _L_CANDS)
    c_col  = _find_col(df.columns, _C_CANDS)
    v_col  = _find_col(df.columns, _V_CANDS)

    print(f"    Mapped  → dt='{dt_col}' | open='{o_col}' | high='{h_col}' | "
          f"low='{l_col}' | close='{c_col}' | volume='{v_col or 'not found'}'")

    missing = [n for n, c in [("datetime", dt_col), ("open", o_col),
                                ("high", h_col), ("low", l_col), ("close", c_col)]
               if c is None]
    if missing:
        raise ValueError(
            f"\n  ❌  Could not detect required columns: {missing}\n"
            f"      Columns in file : {list(df.columns)}\n"
            f"      Please rename to: datetime, open, high, low, close, volume\n"
        )

    rename = {dt_col: "__dt", o_col: "Open", h_col: "High",
              l_col: "Low",   c_col: "Close"}
    if v_col:
        rename[v_col] = "Volume"

    df = df[list(rename.keys())].rename(columns=rename)

    formats_to_try = [
        "%d-%m-%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%dT%H:%M:%S",
    ]

    parsed = None
    for date_fmt in formats_to_try:
        attempt = pd.to_datetime(df["__dt"], format=date_fmt, errors="coerce")
        if attempt.notna().mean() > 0.95:
            parsed = attempt
            print(f"    DateTime format : '{date_fmt}'")
            break

    if parsed is None:
        parsed = pd.to_datetime(df["__dt"], errors="coerce")
        print(f"    ⚠  Format not detected — using auto-parse")

    df["__dt"] = parsed
    df = df.dropna(subset=["__dt"]).copy()

    df = df.set_index("__dt")
    df.index.name = "Datetime"

    if df.index.tz is None:
        df.index = df.index.tz_localize(IST, ambiguous="infer",
                                         nonexistent="shift_forward")
    else:
        df.index = df.index.tz_convert(IST)

    df.sort_index(inplace=True)

    for col in ["Open", "High", "Low", "Close"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    if "Volume" in df.columns:
        df["Volume"] = pd.to_numeric(df["Volume"], errors="coerce").fillna(0)
    else:
        df["Volume"] = 0.0

    df.dropna(subset=["Open","High","Low","Close"], inplace=True)

    df = df[df.index.notna()].copy()
    tod = df.index.time
    df  = df[(tod >= MARKET_OPEN) & (tod <= MARKET_CLOSE)].copy()

    print(f"    Clean rows: {len(df)}  |  "
          f"{fmt(df.index[0])} → {fmt(df.index[-1])}")
    return df


# ───────────────────────────────────────────────────────────────────────────────
#  YFINANCE LOADER
# ───────────────────────────────────────────────────────────────────────────────
def _normalise_yf_df(raw):
    if isinstance(raw.columns, pd.MultiIndex):
        raw.columns = raw.columns.get_level_values(0)
    raw.columns = [str(c).strip().title() for c in raw.columns]
    if "Adj Close" in raw.columns and "Close" not in raw.columns:
        raw = raw.rename(columns={"Adj Close": "Close"})
    keep = [c for c in ["Open","High","Low","Close","Volume"] if c in raw.columns]
    df = raw[keep].copy()
    if "Volume" not in df.columns:
        df["Volume"] = 0.0
    return df


def load_yfinance(ticker, interval, label="yfinance"):
    try:
        import yfinance as yf
    except ImportError:
        raise ImportError(
            "\n  ❌  yfinance is not installed.\n"
            "      Run:  pip install yfinance\n"
        )

    print(f"\n  Downloading {label}: ticker={ticker}  interval={interval} …")

    dl_kwargs = dict(
        tickers     = ticker,
        interval    = interval,
        auto_adjust = True,
        progress    = False,
    )
    if YF_START is not None:
        dl_kwargs["start"] = YF_START
        if YF_END is not None:
            dl_kwargs["end"] = YF_END
        print(f"    Date range  : {YF_START} → {YF_END or 'today'}")
    else:
        dl_kwargs["period"] = YF_PERIOD
        print(f"    Period      : {YF_PERIOD}")

    raw = yf.download(**dl_kwargs)

    if raw is None or raw.empty:
        raise ValueError(
            f"\n  ❌  yfinance returned no data for ticker='{ticker}' "
            f"interval='{interval}'.\n"
        )

    print(f"    Raw shape   : {raw.shape}")
    df = _normalise_yf_df(raw)
    df.index.name = "Datetime"

    if df.index.tz is None:
        df.index = df.index.tz_localize("UTC", ambiguous="infer",
                                         nonexistent="shift_forward")
    df.index = df.index.tz_convert(IST)

    df.sort_index(inplace=True)

    for col in ["Open", "High", "Low", "Close"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["Volume"] = pd.to_numeric(df["Volume"], errors="coerce").fillna(0)
    df.dropna(subset=["Open","High","Low","Close"], inplace=True)

    df = df[df.index.notna()].copy()
    tod = df.index.time
    df  = df[(tod >= MARKET_OPEN) & (tod <= MARKET_CLOSE)].copy()

    if df.empty:
        raise ValueError(
            f"\n  ❌  No data left after market-hours filter for {ticker}.\n"
        )

    print(f"    Clean rows  : {len(df)}  |  "
          f"{fmt(df.index[0])} → {fmt(df.index[-1])}")
    return df


# ───────────────────────────────────────────────────────────────────────────────
#  FETCH / PREPARE ALL DATA
# ───────────────────────────────────────────────────────────────────────────────
def fetch_data():
    src = DATA_SOURCE.strip().lower()

    if src == "csv":
        print(f"\n  ℹ  DATA_SOURCE = 'csv'")
        df = load_csv(BASE_CSV, label="Base TF CSV")
        htf_source_path = HTF_CSV
    elif src == "yfinance":
        print(f"\n  ℹ  DATA_SOURCE = 'yfinance'  |  Ticker: {YF_TICKER}")
        df = load_yfinance(YF_TICKER, YF_BASE_INTERVAL, label="Base TF (yfinance)")
        htf_source_path = YF_HTF_INTERVAL
    else:
        raise ValueError(
            f"\n  ❌  Unknown DATA_SOURCE='{DATA_SOURCE}'.\n"
            f"      Set it to 'csv' or 'yfinance' in the CONFIG section.\n"
        )

    df = compute_supertrend(df, ST_PERIOD, ST_MULTIPLIER)
    df.dropna(inplace=True)

    htf_enabled = False

    if htf_source_path is not None:
        try:
            if src == "csv":
                df_htf = load_csv(htf_source_path, label="HTF CSV")
            else:
                df_htf = load_yfinance(YF_TICKER, YF_HTF_INTERVAL,
                                       label="HTF (yfinance)")

            df_htf = compute_supertrend(df_htf, HTF_ST_PERIOD, HTF_ST_MULTIPLIER)
            df_htf.dropna(inplace=True)

            htf_bull = df_htf["ST_bull"].reindex(df.index, method="ffill")
            htf_val  = df_htf["ST_val"].reindex(df.index, method="ffill")
            coverage = htf_bull.notna().mean()

            if coverage < 0.5:
                print(f"\n  ⚠  HTF index overlap only {coverage:.0%} — "
                      f"HTF filter DISABLED (check date ranges match).")
            else:
                df["HTF_bull"] = htf_bull
                df["HTF_val"]  = htf_val
                df["HTF_bull"] = df["HTF_bull"].ffill()
                df["HTF_val"]  = df["HTF_val"].ffill()
                df.dropna(subset=["HTF_bull", "HTF_val"], inplace=True)
                htf_enabled = True
                exit_mode  = "HTF ST exit" if EXIT_USE_HTF_ST else "Base ST exit"
                close_mode = "close-based" if CLOSE_BASED_EXIT else "intracandle"
                print(f"\n  ✅  HTF ENABLED — ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER}) | "
                      f"Coverage: {coverage:.0%} | "
                      f"Exit mode: {exit_mode} | {close_mode} | "
                      f"Base candles: {len(df)}")

        except FileNotFoundError as e:
            print(e)
            print("  ⚠  HTF filter DISABLED (file not found).")
    else:
        if src == "csv":
            print("\n  ℹ  HTF_CSV = None  →  HTF filter DISABLED.")
        else:
            print("\n  ℹ  YF_HTF_INTERVAL = None  →  HTF filter DISABLED.")

    if not htf_enabled:
        df["HTF_bull"] = True
        df["HTF_val"]  = df["ST_val"]

    return df, htf_enabled


# ───────────────────────────────────────────────────────────────────────────────
#  CHOCH (CHANGE OF CHARACTER) DETECTOR
# ───────────────────────────────────────────────────────────────────────────────
def check_choch_exit(direction, candle_idx, highs, lows, closes,
                     choch_ref_price, choch_ref_candle_idx, current_idx):
    """
    Called once per candle while in a trade. Returns (triggered, new_ref, new_ref_idx).

    Logic
    ─────
    LONG trade
      • After entry we continuously track the running-highest High seen so far.
      • Once at least CHOCH_LOOKBACK candles have passed since the last time that
        peak was set (i.e. a minor pullback has occurred — confirming it as a
        'swing high'), we lock it in as the reference Higher High (HH).
      • If the current candle's Close breaks BELOW the locked HH → ChoCh triggered
        → exit the long.  (Price has failed to hold above the prior HH, signalling
        a potential trend reversal / change of character.)

    SHORT trade
      • Mirror image: track running-lowest Low.
      • After CHOCH_LOOKBACK candles of non-new-lows we lock in the swing low.
      • Close ABOVE locked LL → ChoCh triggered → exit the short.

    Parameters
    ──────────
    direction          : "long" or "short"
    candle_idx         : position in the array of the current candle
    highs / lows / closes : full numpy arrays for the bar data
    choch_ref_price    : last confirmed swing High (long) or swing Low (short)
    choch_ref_candle_idx : bar index when choch_ref_price was last updated
    current_idx        : current bar position

    Returns
    ───────
    (triggered: bool, updated_ref_price: float, updated_ref_candle_idx: int)
    """
    c_high  = highs[current_idx]
    c_low   = lows[current_idx]
    c_close = closes[current_idx]

    if direction == "long":
        # Update the running swing high if this candle made a new high
        if c_high > choch_ref_price:
            # New peak found — reset the reference and its timestamp
            new_ref       = c_high
            new_ref_idx   = current_idx
        else:
            new_ref       = choch_ref_price
            new_ref_idx   = choch_ref_candle_idx

        # Only treat it as a confirmed swing high once CHOCH_LOOKBACK candles
        # have elapsed without the peak being broken to the upside
        candles_since_peak = current_idx - new_ref_idx
        if candles_since_peak >= CHOCH_LOOKBACK:
            # Confirmed swing HH — check if price has broken below it
            if c_close < new_ref:
                return True, new_ref, new_ref_idx   # ← ChoCh triggered

        return False, new_ref, new_ref_idx

    else:  # direction == "short"
        # Update the running swing low if this candle made a new low
        if c_low < choch_ref_price:
            new_ref     = c_low
            new_ref_idx = current_idx
        else:
            new_ref     = choch_ref_price
            new_ref_idx = choch_ref_candle_idx

        candles_since_trough = current_idx - new_ref_idx
        if candles_since_trough >= CHOCH_LOOKBACK:
            if c_close > new_ref:
                return True, new_ref, new_ref_idx   # ← ChoCh triggered

        return False, new_ref, new_ref_idx


# ───────────────────────────────────────────────────────────────────────────────
#  BACKTEST ENGINE
# ───────────────────────────────────────────────────────────────────────────────
def run_backtest(df, htf_enabled):
    trades = []

    in_trade          = False
    direction         = None
    entry_price       = None
    entry_time        = None
    peak              = None
    hard_sl_price     = None
    flip_candle       = -999

    # ── ChoCh state variables ────────────────────────────────────────────────
    # choch_ref_price    : the current candidate swing High (long) or Low (short)
    # choch_ref_idx      : bar index where that extreme was last observed
    choch_ref_price    = None
    choch_ref_idx      = -999

    closes    = df["Close"].values.astype(float)
    highs     = df["High"].values.astype(float)
    lows      = df["Low"].values.astype(float)
    st_vals   = df["ST_val"].values.astype(float)
    htf_vals  = df["HTF_val"].values.astype(float)
    st_bulls  = df["ST_bull"].values.astype(bool)
    htf_bulls = df["HTF_bull"].values.astype(bool)
    volumes   = df["Volume"].values.astype(float)
    times     = df.index

    exit_st = htf_vals if (EXIT_USE_HTF_ST and htf_enabled) else st_vals

    vol_avg      = pd.Series(volumes).rolling(VOLUME_LOOKBACK).mean().values
    daily_trades = {}

    def in_window(t):
        if not TRADE_WINDOW_ENABLED:
            return True
        return any(s <= t <= e for s, e in TRADE_WINDOWS)

    def record(dir_, e_time, e_price, x_time, x_price, pk,
               sl_at_exit, reason, hard_sl):
        pnl = round(((x_price - e_price) / e_price * 100) if dir_ == "long"
                    else ((e_price - x_price) / e_price * 100), 4)
        hold_mins = int((x_time - e_time).total_seconds() // 60)
        hold_str  = f"{hold_mins // 60}H:{hold_mins % 60:02d}M"
        trades.append({
            "Date"            : e_time.strftime("%Y-%m-%d"),
            "Direction"       : "Long  ↑" if dir_ == "long" else "Short ↓",
            "Entry Time"      : fmt(e_time),
            "Entry Price"     : round(e_price, 2),
            "Hard SL Price"   : round(hard_sl, 2) if hard_sl is not None else "OFF",
            "Exit ST Line"    : round(sl_at_exit, 2),
            "Exit Time"       : fmt(x_time),
            "Exit Price"      : round(x_price, 2),
            "Peak"            : round(pk, 2),
            "Holding Mins"    : hold_mins,
            "Holding Time"    : hold_str,
            "Points Captured" : round(abs(x_price - e_price), 4),
            "P&L %"           : pnl,
            "Exit Reason"     : reason,
            "Result"          : "WIN" if pnl > 0 else "LOSS",
        })

    def open_trade(dir_, price, time_, high_, low_, date_key, bar_idx):
        nonlocal in_trade, direction, entry_price, entry_time, peak
        nonlocal hard_sl_price, choch_ref_price, choch_ref_idx
        in_trade      = True
        direction     = dir_
        entry_price   = price
        entry_time    = time_
        peak          = high_ if dir_ == "long" else low_
        hard_sl_price = (round(price * (1 - HARD_SL_PCT / 100), 4) if dir_ == "long"
                         else round(price * (1 + HARD_SL_PCT / 100), 4)) \
                        if HARD_SL_ENABLED else None
        daily_trades[date_key] = daily_trades.get(date_key, 0) + 1

        # ── Initialise ChoCh reference at entry bar ──────────────────────
        # For long  : start tracking from the entry candle's High
        # For short : start tracking from the entry candle's Low
        if CHOCH_EXIT_ENABLED:
            choch_ref_price = high_ if dir_ == "long" else low_
            choch_ref_idx   = bar_idx

    for i in range(1, len(df)):
        c_close    = closes[i];     c_high  = highs[i];      c_low  = lows[i]
        c_time     = times[i];      c_tod   = c_time.time(); c_st   = st_vals[i]
        c_exit_st  = exit_st[i]
        c_bull     = st_bulls[i];   prev_bull = st_bulls[i-1]
        c_htf_bull = htf_bulls[i];  date_key  = c_time.date()

        flipped_bull = (not prev_bull) and c_bull
        flipped_bear = prev_bull and (not c_bull)
        if flipped_bull or flipped_bear:
            flip_candle = i

        # ── EOD force exit ────────────────────────────────────────────────
        if in_trade and c_tod >= MARKET_CLOSE:
            pk_f = max(peak, c_high) if direction == "long" else min(peak, c_low)
            record(direction, entry_time, entry_price,
                   c_time, c_close, pk_f, c_exit_st, "EOD Exit", hard_sl_price)
            in_trade       = False
            choch_ref_price = None
            choch_ref_idx   = -999
            continue

        if c_tod < MARKET_OPEN or c_tod >= MARKET_CLOSE:
            continue

        # ── Manage open trade ─────────────────────────────────────────────
        if in_trade:
            if direction == "long":
                if c_high > peak: peak = c_high

                # --- Hard SL check (always intracandle) ---
                hard_hit = HARD_SL_ENABLED and hard_sl_price and c_low <= hard_sl_price

                # --- ST stop check ---
                if CLOSE_BASED_EXIT:
                    st_hit = (c_close < c_exit_st) or flipped_bear
                else:
                    st_hit = c_low <= c_exit_st or flipped_bear

                # --- ChoCh exit check (close-based only) ---
                choch_hit = False
                if CHOCH_EXIT_ENABLED and not hard_hit and not st_hit:
                    choch_hit, choch_ref_price, choch_ref_idx = check_choch_exit(
                        "long", i, highs, lows, closes,
                        choch_ref_price, choch_ref_idx, i)

                if hard_hit or st_hit or choch_hit:
                    if hard_hit:
                        reason  = f"Hard SL ({HARD_SL_PCT}%)"
                        exit_px = hard_sl_price
                    elif choch_hit:
                        reason  = "ChoCh Exit (HH Break)"
                        exit_px = c_close   # exit at the close that broke the HH
                    elif (CLOSE_BASED_EXIT and c_close < c_exit_st) or \
                         (not CLOSE_BASED_EXIT and c_low <= c_exit_st):
                        exit_label = ("HTF ST Exit" if (EXIT_USE_HTF_ST and htf_enabled)
                                      else "ST Stop Loss")
                        reason  = exit_label
                        exit_px = c_exit_st
                    else:
                        reason  = "ST Flip Bear"
                        exit_px = c_exit_st

                    record("long", entry_time, entry_price,
                           c_time, round(exit_px, 2), peak,
                           c_exit_st, reason, hard_sl_price)
                    in_trade        = False
                    choch_ref_price = None
                    choch_ref_idx   = -999

                    # Immediate reverse entry (only on ST-based exits, not ChoCh)
                    if (not hard_hit and not choch_hit and not c_bull
                            and (not htf_enabled or not c_htf_bull)
                            and in_window(c_tod)
                            and daily_trades.get(date_key, 0) < MAX_TRADES_PER_DAY):
                        open_trade("short", c_close, c_time, c_high, c_low,
                                   date_key, i)

            elif direction == "short":
                if c_low < peak: peak = c_low

                hard_hit = HARD_SL_ENABLED and hard_sl_price and c_high >= hard_sl_price

                if CLOSE_BASED_EXIT:
                    st_hit = (c_close > c_exit_st) or flipped_bull
                else:
                    st_hit = c_high >= c_exit_st or flipped_bull

                choch_hit = False
                if CHOCH_EXIT_ENABLED and not hard_hit and not st_hit:
                    choch_hit, choch_ref_price, choch_ref_idx = check_choch_exit(
                        "short", i, highs, lows, closes,
                        choch_ref_price, choch_ref_idx, i)

                if hard_hit or st_hit or choch_hit:
                    if hard_hit:
                        reason  = f"Hard SL ({HARD_SL_PCT}%)"
                        exit_px = hard_sl_price
                    elif choch_hit:
                        reason  = "ChoCh Exit (LL Break)"
                        exit_px = c_close   # exit at the close that broke the LL
                    elif (CLOSE_BASED_EXIT and c_close > c_exit_st) or \
                         (not CLOSE_BASED_EXIT and c_high >= c_exit_st):
                        exit_label = ("HTF ST Exit" if (EXIT_USE_HTF_ST and htf_enabled)
                                      else "ST Stop Loss")
                        reason  = exit_label
                        exit_px = c_exit_st
                    else:
                        reason  = "ST Flip Bull"
                        exit_px = c_exit_st

                    record("short", entry_time, entry_price,
                           c_time, round(exit_px, 2), peak,
                           c_exit_st, reason, hard_sl_price)
                    in_trade        = False
                    choch_ref_price = None
                    choch_ref_idx   = -999

                    if (not hard_hit and not choch_hit and c_bull
                            and (not htf_enabled or c_htf_bull)
                            and in_window(c_tod)
                            and daily_trades.get(date_key, 0) < MAX_TRADES_PER_DAY):
                        open_trade("long", c_close, c_time, c_high, c_low,
                                   date_key, i)
            continue

        # ── Entry filters ─────────────────────────────────────────────────
        if not in_window(c_tod):                                              continue
        if (i - flip_candle) < CONFIRM_CANDLES:                              continue
        if MIN_GAP_ENABLED and abs(c_close-c_st)/c_close*100 < MIN_GAP_PCT: continue
        if (VOLUME_FILTER_ENABLED and not np.isnan(vol_avg[i])
                and vol_avg[i] > 0
                and volumes[i] < VOLUME_MULTIPLIER * vol_avg[i]):            continue
        if (MAX_TRADES_PER_DAY_ENABLED
                and daily_trades.get(date_key, 0) >= MAX_TRADES_PER_DAY):   continue

        if c_bull  and (not htf_enabled or c_htf_bull):
            open_trade("long",  c_close, c_time, c_high, c_low, date_key, i)
        elif not c_bull and (not htf_enabled or not c_htf_bull):
            open_trade("short", c_close, c_time, c_high, c_low, date_key, i)

    return trades


# ───────────────────────────────────────────────────────────────────────────────
#  SUMMARY BUILDERS
# ───────────────────────────────────────────────────────────────────────────────
def build_daily_summary(trades):
    if not trades: return pd.DataFrame()
    df_t = pd.DataFrame(trades); rows = []
    for date, grp in df_t.groupby("Date"):
        total = len(grp); wins = (grp["P&L %"] > 0).sum(); pnl = grp["P&L %"].sum()
        wp = grp.loc[grp["P&L %"] > 0,  "Points Captured"].sum()
        lp = grp.loc[grp["P&L %"] <= 0, "Points Captured"].sum()
        rows.append({
            "Date"           : date,
            "Total Trades"   : total,
            "Longs"          : grp["Direction"].str.contains("Long").sum(),
            "Shorts"         : grp["Direction"].str.contains("Short").sum(),
            "Winners"        : wins,
            "Losers"         : total - wins,
            "Win Rate %"     : round(wins / total * 100, 1),
            "Total P&L %"    : round(pnl, 4),
            "Best Trade %"   : round(grp["P&L %"].max(), 4),
            "Worst Trade %"  : round(grp["P&L %"].min(), 4),
            "Points Captured": round(wp, 2),
            "Points Lost"    : round(lp, 2),
            "Net Points"     : round(wp - lp, 2),
            "Avg Hold Time"  : mins_to_hhmm(grp["Holding Mins"].mean()),
            "Day Result"     : "✅ Profit" if pnl > 0 else "❌ Loss" if pnl < 0 else "⚖ Flat",
        })
    return pd.DataFrame(rows)


def build_time_slot_summary(trades):
    if not trades: return pd.DataFrame()
    SLOTS = [
        ("09:15","09:30"),("09:30","10:00"),("10:00","10:30"),("10:30","11:00"),
        ("11:00","11:30"),("11:30","12:00"),("12:00","12:30"),("12:30","13:00"),
        ("13:00","13:30"),("13:30","14:00"),("14:00","14:30"),("14:30","15:00"),
        ("15:00","15:15"),
    ]
    df_t = pd.DataFrame(trades)
    df_t["Entry_tod"] = (pd.to_datetime(df_t["Entry Time"], format="%Y-%m-%d %H:%M")
                         .dt.strftime("%H:%M"))
    rows = []
    for s, e in SLOTS:
        grp = df_t[(df_t["Entry_tod"] >= s) & (df_t["Entry_tod"] < e)]
        if grp.empty:
            rows.append({"Time Slot": f"{s}–{e}", "Trades": 0, "Winners": 0,
                         "Losers": 0, "Win Rate %": "—", "Total P&L %": 0.0,
                         "Avg P&L %": 0.0, "Best %": 0.0, "Worst %": 0.0, "Verdict": "—"})
            continue
        total = len(grp); wins = (grp["P&L %"] > 0).sum()
        pnl = grp["P&L %"].sum(); wr = wins / total * 100
        rows.append({
            "Time Slot"  : f"{s}–{e}",
            "Trades"     : total,
            "Winners"    : wins,
            "Losers"     : total - wins,
            "Win Rate %" : round(wr, 1),
            "Total P&L %": round(pnl, 4),
            "Avg P&L %"  : round(grp["P&L %"].mean(), 4),
            "Best %"     : round(grp["P&L %"].max(), 4),
            "Worst %"    : round(grp["P&L %"].min(), 4),
            "Verdict"    : "🟢 Trade" if pnl > 0 and wr >= 50 else "🔴 Avoid",
        })
    return pd.DataFrame(rows)


# ───────────────────────────────────────────────────────────────────────────────
#  HOLDING TIME SUMMARY
# ───────────────────────────────────────────────────────────────────────────────
def build_holding_time_summary(trades):
    if not trades:
        return pd.DataFrame(), pd.DataFrame()

    df_t = pd.DataFrame(trades)
    if "Holding Mins" not in df_t.columns:
        return pd.DataFrame(), pd.DataFrame()

    overall_avg = mins_to_hhmm(df_t["Holding Mins"].mean())
    overall_min = mins_to_hhmm(df_t["Holding Mins"].min())
    overall_max = mins_to_hhmm(df_t["Holding Mins"].max())

    wins_mask   = df_t["P&L %"] > 0
    losses_mask = df_t["P&L %"] <= 0
    longs_mask  = df_t["Direction"].str.contains("Long")
    shorts_mask = df_t["Direction"].str.contains("Short")

    wins_avg   = mins_to_hhmm(df_t.loc[wins_mask,   "Holding Mins"].mean()) if wins_mask.any()   else "—"
    losses_avg = mins_to_hhmm(df_t.loc[losses_mask, "Holding Mins"].mean()) if losses_mask.any() else "—"
    longs_avg  = mins_to_hhmm(df_t.loc[longs_mask,  "Holding Mins"].mean()) if longs_mask.any()  else "—"
    shorts_avg = mins_to_hhmm(df_t.loc[shorts_mask, "Holding Mins"].mean()) if shorts_mask.any() else "—"

    bins   = [0, 5, 15, 30, 60, 120, 9999]
    labels = ["0–5 min", "6–15 min", "16–30 min", "31–60 min", "1–2 hrs", ">2 hrs"]
    df_t["Hold_Bucket"] = pd.cut(df_t["Holding Mins"], bins=bins, labels=labels, right=True)
    bucket_counts = df_t["Hold_Bucket"].value_counts().reindex(labels, fill_value=0)

    summary_rows = [
        ["OVERALL HOLDING TIME",   ""],
        ["Avg Holding Time",        overall_avg],
        ["Min Holding Time",        overall_min],
        ["Max Holding Time",        overall_max],
        ["",                        ""],
        ["BY RESULT",               ""],
        ["Avg Hold — Winners",      wins_avg],
        ["Avg Hold — Losers",       losses_avg],
        ["",                        ""],
        ["BY DIRECTION",            ""],
        ["Avg Hold — Longs",        longs_avg],
        ["Avg Hold — Shorts",       shorts_avg],
        ["",                        ""],
        ["DURATION DISTRIBUTION",   ""],
    ]
    for lbl in labels:
        cnt = int(bucket_counts[lbl])
        pct = cnt / len(df_t) * 100
        summary_rows.append([f"  {lbl}", f"{cnt} trades  ({pct:.1f}%)"])

    df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    SLOTS = [
        ("09:15","09:30"),("09:30","10:00"),("10:00","10:30"),("10:30","11:00"),
        ("11:00","11:30"),("11:30","12:00"),("12:00","12:30"),("12:30","13:00"),
        ("13:00","13:30"),("13:30","14:00"),("14:00","14:30"),("14:30","15:00"),
        ("15:00","15:15"),
    ]
    df_t["Entry_tod"] = (pd.to_datetime(df_t["Entry Time"], format="%Y-%m-%d %H:%M")
                         .dt.strftime("%H:%M"))

    slot_rows = []
    for s, e in SLOTS:
        grp = df_t[(df_t["Entry_tod"] >= s) & (df_t["Entry_tod"] < e)]
        if grp.empty:
            slot_rows.append({
                "Time Slot"    : f"{s}–{e}",
                "Trades"       : 0,
                "Avg Holding"  : "—",
                "Min Holding"  : "—",
                "Max Holding"  : "—",
                "Avg P&L %"    : "—",
                "Win Avg Hold" : "—",
                "Loss Avg Hold": "—",
            })
            continue
        w = grp[grp["P&L %"] > 0];  l = grp[grp["P&L %"] <= 0]
        slot_rows.append({
            "Time Slot"    : f"{s}–{e}",
            "Trades"       : len(grp),
            "Avg Holding"  : mins_to_hhmm(grp["Holding Mins"].mean()),
            "Min Holding"  : mins_to_hhmm(grp["Holding Mins"].min()),
            "Max Holding"  : mins_to_hhmm(grp["Holding Mins"].max()),
            "Avg P&L %"    : round(grp["P&L %"].mean(), 4),
            "Win Avg Hold" : mins_to_hhmm(w["Holding Mins"].mean()) if not w.empty else "—",
            "Loss Avg Hold": mins_to_hhmm(l["Holding Mins"].mean()) if not l.empty else "—",
        })
    df_slots = pd.DataFrame(slot_rows)

    return df_summary, df_slots


# ───────────────────────────────────────────────────────────────────────────────
#  ADVANCED STATS BUILDER
# ───────────────────────────────────────────────────────────────────────────────
LOW_POINTS_THRESHOLD = 20

def build_advanced_stats(trades, df_day):
    if not trades:
        return {}

    df_t = pd.DataFrame(trades)

    low_pt = df_t[(df_t["Points Captured"] < LOW_POINTS_THRESHOLD) & (df_t["P&L %"] > 0)].copy()
    low_pt_total = len(low_pt)

    if not low_pt.empty:
        low_pt_daily = (low_pt.groupby("Date")
                        .agg(
                            Low_Pt_Trades   = ("Points Captured", "count"),
                            Avg_Points      = ("Points Captured", lambda x: round(x.mean(), 2)),
                            Total_PnL_Pct   = ("P&L %", "sum"),
                            Winners         = ("Result", lambda x: (x == "WIN").sum()),
                            Losers          = ("Result", lambda x: (x == "LOSS").sum()),
                        )
                        .reset_index()
                        .rename(columns={
                            "Low_Pt_Trades": f"Trades < {LOW_POINTS_THRESHOLD}pts",
                            "Avg_Points"   : "Avg Points",
                            "Total_PnL_Pct": "Total P&L %",
                        }))
        low_pt_daily["Total P&L %"] = low_pt_daily["Total P&L %"].round(4)
    else:
        low_pt_daily = pd.DataFrame(columns=["Date", f"Trades < {LOW_POINTS_THRESHOLD}pts",
                                              "Avg Points","Total P&L %","Winners","Losers"])

    results = (df_t["P&L %"] > 0).astype(int).tolist()
    max_win_streak  = 0; cur_win  = 0
    max_loss_streak = 0; cur_loss = 0
    win_streak_end  = None; loss_streak_end = None

    for idx, r in enumerate(results):
        if r == 1:
            cur_win += 1; cur_loss = 0
            if cur_win > max_win_streak:
                max_win_streak = cur_win
                win_streak_end = df_t.iloc[idx]["Exit Time"]
        else:
            cur_loss += 1; cur_win = 0
            if cur_loss > max_loss_streak:
                max_loss_streak = cur_loss
                loss_streak_end = df_t.iloc[idx]["Exit Time"]

    total_longs  = df_t["Direction"].str.contains("Long").sum()
    total_shorts = df_t["Direction"].str.contains("Short").sum()
    long_wins    = df_t[(df_t["Direction"].str.contains("Long"))  & (df_t["P&L %"] > 0)].shape[0]
    short_wins   = df_t[(df_t["Direction"].str.contains("Short")) & (df_t["P&L %"] > 0)].shape[0]

    if df_day is not None and not df_day.empty:
        profit_days_pts = df_day.loc[df_day["Net Points"] > 0, "Net Points"]
        loss_days_pts   = df_day.loc[df_day["Net Points"] < 0, "Net Points"]
        avg_profit_per_day     = round(profit_days_pts.mean(), 2) if not profit_days_pts.empty else 0.0
        avg_loss_per_day       = round(loss_days_pts.mean(),   2) if not loss_days_pts.empty   else 0.0
        highest_profit_day_pts = round(df_day["Net Points"].max(), 2)
        highest_loss_day_pts   = round(df_day["Net Points"].min(), 2)
        highest_profit_day     = df_day.loc[df_day["Net Points"].idxmax(), "Date"]
        highest_loss_day       = df_day.loc[df_day["Net Points"].idxmin(), "Date"]
    else:
        avg_profit_per_day = avg_loss_per_day = 0.0
        highest_profit_day_pts = highest_loss_day_pts = 0.0
        highest_profit_day = highest_loss_day = "—"

    # ── ChoCh exit stats ─────────────────────────────────────────────────────
    choch_long_exits  = int(df_t["Exit Reason"].str.contains("ChoCh.*HH", na=False).sum())
    choch_short_exits = int(df_t["Exit Reason"].str.contains("ChoCh.*LL", na=False).sum())
    choch_total       = choch_long_exits + choch_short_exits
    if choch_total > 0:
        choch_wins = int(df_t[df_t["Exit Reason"].str.contains("ChoCh", na=False)
                              & (df_t["P&L %"] > 0)].shape[0])
        choch_wr   = round(choch_wins / choch_total * 100, 1)
        choch_avg_pnl = round(
            df_t.loc[df_t["Exit Reason"].str.contains("ChoCh", na=False), "P&L %"].mean(), 4)
    else:
        choch_wins = 0; choch_wr = 0.0; choch_avg_pnl = 0.0

    return {
        "low_pt_trades"         : low_pt,
        "low_pt_total"          : low_pt_total,
        "low_pt_daily"          : low_pt_daily,
        "max_win_streak"        : max_win_streak,
        "max_loss_streak"       : max_loss_streak,
        "win_streak_end"        : win_streak_end,
        "loss_streak_end"       : loss_streak_end,
        "total_longs"           : total_longs,
        "total_shorts"          : total_shorts,
        "long_wins"             : long_wins,
        "short_wins"            : short_wins,
        "avg_profit_per_day"    : avg_profit_per_day,
        "avg_loss_per_day"      : avg_loss_per_day,
        "highest_profit_day_pts": highest_profit_day_pts,
        "highest_loss_day_pts"  : highest_loss_day_pts,
        "highest_profit_day"    : highest_profit_day,
        "highest_loss_day"      : highest_loss_day,
        # ChoCh-specific
        "choch_total"           : choch_total,
        "choch_long_exits"      : choch_long_exits,
        "choch_short_exits"     : choch_short_exits,
        "choch_wins"            : choch_wins,
        "choch_wr"              : choch_wr,
        "choch_avg_pnl"         : choch_avg_pnl,
    }


# ───────────────────────────────────────────────────────────────────────────────
#  PRINT RESULTS
# ───────────────────────────────────────────────────────────────────────────────
def print_results(trades, htf_enabled):
    SEP  = "═" * 120
    DASH = "─" * 120
    htf_info     = (f"ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER}) from '{HTF_CSV}'"
                    if htf_enabled else "OFF (HTF_CSV = None or file not found)")
    exit_st_info = (f"HTF ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER})"
                    if (EXIT_USE_HTF_ST and htf_enabled) else f"Base ST({ST_PERIOD},{ST_MULTIPLIER})")
    close_info   = "Close-based (no wick exits)" if CLOSE_BASED_EXIT else "Intracandle (Low/High)"
    choch_info   = (f"ON (lookback={CHOCH_LOOKBACK} candles)" if CHOCH_EXIT_ENABLED else "OFF")

    print("\n" + SEP)
    print(f"  {SYMBOL}  |  Entry ST({ST_PERIOD},{ST_MULTIPLIER})  |  EOD: {MARKET_CLOSE}")
    print(f"  Base CSV  : {BASE_CSV}")
    print(f"  HTF       : {htf_info}")
    print(f"  Exit ST   : {exit_st_info}")
    print(f"  Exit Mode : {close_info}")
    print(f"  ChoCh Exit: {choch_info}")
    print(f"  Filters   : Window={'ON' if TRADE_WINDOW_ENABLED else 'OFF'}  "
          f"| ConfirmN={CONFIRM_CANDLES}  "
          f"| MinGap={'ON('+str(MIN_GAP_PCT)+'%)' if MIN_GAP_ENABLED else 'OFF'}  "
          f"| Volume={'ON' if VOLUME_FILTER_ENABLED else 'OFF'}  "
          f"| MaxTrades/Day={MAX_TRADES_PER_DAY if MAX_TRADES_PER_DAY_ENABLED else 'OFF'}")
    print(f"  Hard SL   : {'ON — '+str(HARD_SL_PCT)+'% from entry (always intracandle)' if HARD_SL_ENABLED else 'OFF'}")
    print(SEP)

    if not trades:
        print("  ⚠  No trades found. Try relaxing filters or checking CSV date range.")
        print(SEP)
        return None, None, None, None, pd.DataFrame(), pd.DataFrame()

    df_t   = pd.DataFrame(trades)
    df_day = build_daily_summary(trades)
    df_ts  = build_time_slot_summary(trades)
    adv    = build_advanced_stats(trades, df_day)
    df_hold_sum, df_hold_slots = build_holding_time_summary(trades)

    total    = len(df_t)
    wins     = (df_t["P&L %"] > 0).sum()
    losses   = total - wins
    win_pts  = df_t.loc[df_t["P&L %"] > 0,  "Points Captured"].sum()
    loss_pts = df_t.loc[df_t["P&L %"] <= 0, "Points Captured"].sum()
    p_days   = (df_day["Total P&L %"] > 0).sum()
    l_days   = (df_day["Total P&L %"] < 0).sum()
    best_d   = df_day.loc[df_day["Total P&L %"].idxmax()]
    wrst_d   = df_day.loc[df_day["Total P&L %"].idxmin()]
    best_s   = df_ts.loc[df_ts["Total P&L %"].idxmax()] if not df_ts.empty else None
    wrst_s   = df_ts.loc[df_ts["Total P&L %"].idxmin()] if not df_ts.empty else None

    # ── Trade Log ─────────────────────────────────────────────────────────
    print("\n  TRADE LOG"); print(DASH)
    cols = ["Date","Direction","Entry Time","Entry Price","Hard SL Price",
            "Exit ST Line","Exit Time","Exit Price","Holding Time",
            "Points Captured","P&L %","Exit Reason","Result"]
    print(df_t[cols].to_string(index=False))

    # ── Low-point trades ──────────────────────────────────────────────────
    print("\n\n" + SEP)
    print(f"  LOW-POINT TRADES  (Winning trades with Points Captured < {LOW_POINTS_THRESHOLD})")
    print(DASH)
    if adv["low_pt_total"] > 0:
        lp_cols = ["Date","Direction","Entry Time","Entry Price","Exit Time",
                   "Exit Price","Holding Time","Points Captured","P&L %","Exit Reason","Result"]
        print(adv["low_pt_trades"][lp_cols].to_string(index=False))
        print(f"\n  Total: {adv['low_pt_total']} of {total} ({adv['low_pt_total']/total*100:.1f}%)")
        print(f"\n  Per-day breakdown:")
        print(adv["low_pt_daily"].to_string(index=False))
    else:
        print(f"  ✅  No winning trades with Points Captured < {LOW_POINTS_THRESHOLD}")

    # ── Per-day P&L ───────────────────────────────────────────────────────
    print("\n\n" + SEP); print("  PER-DAY P&L BREAKDOWN"); print(DASH)
    print(df_day.to_string(index=False))

    # ── Time-slot P&L ─────────────────────────────────────────────────────
    print("\n\n" + SEP)
    print("  TIME-SLOT P&L ANALYSIS  (entry time grouped into 30-min windows)")
    print(DASH); print(df_ts.to_string(index=False))

    # ── Holding Time Report ───────────────────────────────────────────────
    print("\n\n" + SEP)
    print("  HOLDING TIME ANALYSIS")
    print(DASH)
    if not df_hold_sum.empty:
        print(df_hold_sum.to_string(index=False))
        print()
        print("  Avg Holding Time by Entry Time Slot:")
        print(DASH)
        print(df_hold_slots.to_string(index=False))

    # ── Overall Summary ───────────────────────────────────────────────────
    print("\n\n" + SEP); print("  OVERALL SUMMARY"); print(DASH)
    print(f"  {'EXIT CONFIGURATION':─<55}")
    print(f"  Exit ST Line        : {exit_st_info}")
    print(f"  Exit Trigger        : {close_info}")
    print(f"  Hard SL             : {'ON ('+str(HARD_SL_PCT)+'%) — always intracandle' if HARD_SL_ENABLED else 'OFF'}")
    print(f"  ChoCh Exit          : {choch_info}")
    print(DASH)
    print(f"  {'TRADE STATS':─<55}")
    print(f"  Total Trades        : {total}")
    print(f"  Winners             : {wins}  ({wins/total*100:.1f}%)")
    print(f"  Losers              : {losses}  ({losses/total*100:.1f}%)")
    htf_exit_count = df_t["Exit Reason"].str.contains("HTF ST Exit").sum()
    print(f"  Hard SL Exits       : {df_t['Exit Reason'].str.startswith('Hard SL').sum()}")
    print(f"  HTF ST Exits        : {htf_exit_count}")
    print(f"  Base ST SL Exits    : {(df_t['Exit Reason'] == 'ST Stop Loss').sum()}")
    print(f"  ST Flip Exits       : {df_t['Exit Reason'].str.startswith('ST Flip').sum()}")
    print(f"  EOD Force Exits     : {(df_t['Exit Reason'] == 'EOD Exit').sum()}")
    # ── NEW: ChoCh exit counts ─────────────────────────────────────────────
    print(f"  ChoCh Exits (Long)  : {adv['choch_long_exits']}  (HH break → exit long)")
    print(f"  ChoCh Exits (Short) : {adv['choch_short_exits']}  (LL break → exit short)")
    print(DASH)
    print(f"  {'CHOCH EXIT ANALYSIS':─<55}")
    if adv['choch_total'] > 0:
        print(f"  Total ChoCh Exits   : {adv['choch_total']}")
        print(f"  ChoCh Win Rate      : {adv['choch_wr']}%  ({adv['choch_wins']} wins)")
        print(f"  ChoCh Avg P&L %     : {adv['choch_avg_pnl']:.4f}%")
    else:
        print(f"  No ChoCh exits triggered in this backtest period.")
    print(DASH)
    print(f"  {'P&L STATS':─<55}")
    print(f"  Total P&L           : {df_t['P&L %'].sum():.4f}%")
    print(f"  Avg P&L / Trade     : {df_t['P&L %'].mean():.4f}%")
    print(f"  Best Single Trade   : {df_t['P&L %'].max():.4f}%")
    print(f"  Worst Single Trade  : {df_t['P&L %'].min():.4f}%")
    print(DASH)
    print(f"  {'POINTS':─<55}")
    print(f"  Points Captured     : {win_pts:.2f}  (winning trades)")
    print(f"  Points Lost         : {loss_pts:.2f}  (losing trades)")
    print(f"  Net Points          : {win_pts - loss_pts:.2f}")
    print(DASH)
    print(f"  {'HOLDING TIME':─<55}")
    if not df_hold_sum.empty:
        hs = df_hold_sum.set_index("Metric")["Value"].to_dict()
        print(f"  Avg Holding Time    : {hs.get('Avg Holding Time','—')}")
        print(f"  Avg Hold — Winners  : {hs.get('Avg Hold — Winners','—')}")
        print(f"  Avg Hold — Losers   : {hs.get('Avg Hold — Losers','—')}")
        print(f"  Avg Hold — Longs    : {hs.get('Avg Hold — Longs','—')}")
        print(f"  Avg Hold — Shorts   : {hs.get('Avg Hold — Shorts','—')}")
    print(DASH)
    print(f"  {'DAY STATS':─<55}")
    print(f"  Total Days          : {len(df_day)}")
    print(f"  Profit Days         : {p_days}  ({p_days/len(df_day)*100:.1f}%)")
    print(f"  Loss Days           : {l_days}  ({l_days/len(df_day)*100:.1f}%)")
    print(f"  Flat Days           : {len(df_day) - p_days - l_days}")
    print(f"  Best Day            : {best_d['Date']}  →  {best_d['Total P&L %']:.4f}%  |  Net Pts: {best_d['Net Points']:.2f}")
    print(f"  Worst Day           : {wrst_d['Date']}  →  {wrst_d['Total P&L %']:.4f}%  |  Net Pts: {wrst_d['Net Points']:.2f}")
    if best_s is not None:
        print(DASH)
        print(f"  {'TIME SLOT INSIGHTS':─<55}")
        print(f"  Best Slot           : {best_s['Time Slot']}  →  {best_s['Total P&L %']:.4f}%  |  WR: {best_s['Win Rate %']}%")
        print(f"  Worst Slot          : {wrst_s['Time Slot']}  →  {wrst_s['Total P&L %']:.4f}%  |  WR: {wrst_s['Win Rate %']}%")
    print(DASH)
    print(f"  {'DIRECTION BREAKDOWN':─<55}")
    if adv['total_longs']  > 0: print(f"  Total Longs         : {adv['total_longs']}  |  Won: {adv['long_wins']}  ({adv['long_wins']/adv['total_longs']*100:.1f}%)")
    if adv['total_shorts'] > 0: print(f"  Total Shorts        : {adv['total_shorts']}  |  Won: {adv['short_wins']}  ({adv['short_wins']/adv['total_shorts']*100:.1f}%)")
    print(DASH)
    print(f"  {'STREAK ANALYSIS':─<55}")
    print(f"  Max Win  Streak     : {adv['max_win_streak']}  consecutive wins   (ended at: {adv['win_streak_end'] or '—'})")
    print(f"  Max Loss Streak     : {adv['max_loss_streak']}  consecutive losses (ended at: {adv['loss_streak_end'] or '—'})")
    print(DASH)
    print(f"  {'LOW-POINT TRADE SUMMARY':─<55}")
    print(f"  Trades < {LOW_POINTS_THRESHOLD} pts       : {adv['low_pt_total']} of {total}  ({adv['low_pt_total']/total*100:.1f}%)")
    print(DASH)
    print(f"  {'DAILY POINT AVERAGES':─<55}")
    print(f"  Avg Profit / Day    : +{adv['avg_profit_per_day']:.2f} pts  (profit days only)")
    print(f"  Avg Loss   / Day    : {adv['avg_loss_per_day']:.2f} pts  (loss days only)")
    print(f"  Highest Profit Day  : {adv['highest_profit_day']}  →  +{adv['highest_profit_day_pts']:.2f} pts")
    print(f"  Highest Loss Day    : {adv['highest_loss_day']}  →  {adv['highest_loss_day_pts']:.2f} pts")
    print(SEP + "\n")

    return df_t, df_day, df_ts, adv, df_hold_sum, df_hold_slots


# ───────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ───────────────────────────────────────────────────────────────────────────────
def export_excel(df, df_trades, df_day, df_ts, htf_enabled, adv,
                 df_hold_sum=None, df_hold_slots=None):
    fname = f"{SYMBOL}_supertrend_v3.xlsx"

    df_st = df[["Open","High","Low","Close","Volume","ST_val","ST_direction"]].copy()
    df_st.index = df_st.index.strftime("%Y-%m-%d %H:%M")
    df_st.index.name = "DateTime (IST)"
    df_st.columns = ["Open","High","Low","Close","Volume",
                     f"ST({ST_PERIOD},{ST_MULTIPLIER})","Direction"]
    df_st = df_st.round(2)

    if df_trades is not None and not df_trades.empty:
        total    = len(df_trades); wins = (df_trades["P&L %"] > 0).sum()
        win_pts  = df_trades.loc[df_trades["P&L %"] > 0,  "Points Captured"].sum()
        loss_pts = df_trades.loc[df_trades["P&L %"] <= 0, "Points Captured"].sum()
        pd_      = int((df_day["Total P&L %"] > 0).sum()) if df_day is not None else 0
        ld_      = int((df_day["Total P&L %"] < 0).sum()) if df_day is not None else 0

        avg_hold_overall = "—"
        avg_hold_wins    = "—"
        avg_hold_losses  = "—"
        if "Holding Mins" in df_trades.columns:
            avg_hold_overall = mins_to_hhmm(df_trades["Holding Mins"].mean())
            w = df_trades[df_trades["P&L %"] > 0]
            l = df_trades[df_trades["P&L %"] <= 0]
            if not w.empty: avg_hold_wins   = mins_to_hhmm(w["Holding Mins"].mean())
            if not l.empty: avg_hold_losses = mins_to_hhmm(l["Holding Mins"].mean())

        choch_info_str = (f"ON (lookback={CHOCH_LOOKBACK} candles)"
                          if CHOCH_EXIT_ENABLED else "OFF")

        summary_rows = [
            ["STRATEGY INFO",  ""],
            ["Symbol",         SYMBOL],
            ["Base CSV",       BASE_CSV],
            ["HTF CSV",        HTF_CSV or "—  (HTF filter OFF)"],
            ["Base ST",        f"Period={ST_PERIOD}, Mult={ST_MULTIPLIER}"],
            ["HTF ST",         f"Period={HTF_ST_PERIOD}, Mult={HTF_ST_MULTIPLIER}"],
            ["HTF Active",     "YES" if htf_enabled else "NO"],
            ["EOD Exit",       str(MARKET_CLOSE)],
            ["", ""],
            ["EXIT CONFIG",    ""],
            ["Exit ST Line",   f"HTF ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER})" if (EXIT_USE_HTF_ST and htf_enabled) else f"Base ST({ST_PERIOD},{ST_MULTIPLIER})"],
            ["Exit Trigger",   "Close-based" if CLOSE_BASED_EXIT else "Intracandle"],
            ["Hard SL",        f"{'ON ('+str(HARD_SL_PCT)+'%)' if HARD_SL_ENABLED else 'OFF'}"],
            ["ChoCh Exit",     choch_info_str],
            ["ChoCh Lookback", f"{CHOCH_LOOKBACK} candles" if CHOCH_EXIT_ENABLED else "N/A"],
            ["", ""],
            ["FILTERS",        ""],
            ["HTF Filter",     "ON" if htf_enabled else "OFF"],
            ["Trade Window",   "ON" if TRADE_WINDOW_ENABLED else "OFF"],
            ["Confirm Candles",CONFIRM_CANDLES],
            ["Min Gap",        f"{'ON ('+str(MIN_GAP_PCT)+'%)' if MIN_GAP_ENABLED else 'OFF'}"],
            ["Volume Filter",  f"{'ON ('+str(VOLUME_MULTIPLIER)+'x)' if VOLUME_FILTER_ENABLED else 'OFF'}"],
            ["Max Trades/Day", MAX_TRADES_PER_DAY if MAX_TRADES_PER_DAY_ENABLED else "OFF"],
            ["", ""],
            ["TRADE STATS",    ""],
            ["Total Trades",   total],
            ["Winners",        f"{wins} ({wins/total*100:.1f}%)"],
            ["Losers",         f"{total-wins} ({(total-wins)/total*100:.1f}%)"],
            ["Hard SL Exits",  int(df_trades["Exit Reason"].str.startswith("Hard SL").sum())],
            ["HTF ST Exits",   int(df_trades["Exit Reason"].str.contains("HTF ST Exit").sum())],
            ["Base ST Exits",  int((df_trades["Exit Reason"]=="ST Stop Loss").sum())],
            ["ST Flip Exits",  int(df_trades["Exit Reason"].str.startswith("ST Flip").sum())],
            ["EOD Exits",      int((df_trades["Exit Reason"]=="EOD Exit").sum())],
            ["ChoCh Exits (Long)",  adv.get("choch_long_exits", 0)  if adv else 0],
            ["ChoCh Exits (Short)", adv.get("choch_short_exits", 0) if adv else 0],
            ["", ""],
            ["CHOCH EXIT ANALYSIS", ""],
            ["Total ChoCh Exits",   adv.get("choch_total", 0)   if adv else 0],
            ["ChoCh Win Rate",      f"{adv.get('choch_wr', 0.0):.1f}%" if adv else "—"],
            ["ChoCh Avg P&L %",     f"{adv.get('choch_avg_pnl', 0.0):.4f}%" if adv else "—"],
            ["", ""],
            ["P&L STATS",      ""],
            ["Total P&L %",    round(df_trades["P&L %"].sum(), 4)],
            ["Avg P&L %",      round(df_trades["P&L %"].mean(), 4)],
            ["Best Trade %",   round(df_trades["P&L %"].max(), 4)],
            ["Worst Trade %",  round(df_trades["P&L %"].min(), 4)],
            ["", ""],
            ["HOLDING TIME",   ""],
            ["Avg Hold Overall",  avg_hold_overall],
            ["Avg Hold Winners",  avg_hold_wins],
            ["Avg Hold Losers",   avg_hold_losses],
            ["", ""],
            ["POINTS",         ""],
            ["Points Captured",round(win_pts, 2)],
            ["Points Lost",    round(loss_pts, 2)],
            ["Net Points",     round(win_pts - loss_pts, 2)],
            ["", ""],
            ["DAY STATS",      ""],
            ["Total Days",     len(df_day) if df_day is not None else 0],
            ["Profit Days",    pd_],
            ["Loss Days",      ld_],
        ]
        if df_day is not None and not df_day.empty:
            bd = df_day.loc[df_day["Total P&L %"].idxmax()]
            wd = df_day.loc[df_day["Total P&L %"].idxmin()]
            summary_rows += [
                ["Best Day",  f"{bd['Date']}  →  {bd['Total P&L %']:.4f}%"],
                ["Worst Day", f"{wd['Date']}  →  {wd['Total P&L %']:.4f}%"],
            ]
        df_summary = pd.DataFrame(summary_rows, columns=["Metric","Value"])
    else:
        df_summary = pd.DataFrame([{"Metric":"No trades found","Value":""}])

    df_trades_export = df_trades.drop(columns=["Holding Mins"], errors="ignore") \
                       if df_trades is not None else None

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        df_st.to_excel(writer, sheet_name="Supertrend Data", index=True)
        if df_trades_export is not None and not df_trades_export.empty:
            df_trades_export.to_excel(writer, sheet_name="Trade Log", index=False)
        if df_day is not None and not df_day.empty:
            df_day.to_excel(writer, sheet_name="Daily P&L", index=False)
        if df_ts is not None and not df_ts.empty:
            df_ts.to_excel(writer, sheet_name="Time Slot P&L", index=False)
        if df_hold_sum is not None and not df_hold_sum.empty:
            df_hold_sum.to_excel(writer, sheet_name="Holding Time Summary", index=False)
        if df_hold_slots is not None and not df_hold_slots.empty:
            df_hold_slots.to_excel(writer, sheet_name="Holding By Time Slot", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        if adv:
            total_t = len(df_trades) if df_trades is not None else 1
            adv_rows = [
                ["DIRECTION BREAKDOWN", ""],
                ["Total Longs",         adv["total_longs"]],
                ["Long Wins",           f"{adv['long_wins']} ({adv['long_wins']/max(adv['total_longs'],1)*100:.1f}%)"],
                ["Total Shorts",        adv["total_shorts"]],
                ["Short Wins",          f"{adv['short_wins']} ({adv['short_wins']/max(adv['total_shorts'],1)*100:.1f}%)"],
                ["", ""],
                ["STREAK ANALYSIS", ""],
                ["Max Win Streak",      adv["max_win_streak"]],
                ["Win Streak Ended At", adv["win_streak_end"] or "—"],
                ["Max Loss Streak",     adv["max_loss_streak"]],
                ["Loss Streak Ended At",adv["loss_streak_end"] or "—"],
                ["", ""],
                ["CHOCH EXIT ANALYSIS", ""],
                ["ChoCh Exit Enabled",  "YES" if CHOCH_EXIT_ENABLED else "NO"],
                ["ChoCh Lookback",      f"{CHOCH_LOOKBACK} candles" if CHOCH_EXIT_ENABLED else "N/A"],
                ["Total ChoCh Exits",   adv["choch_total"]],
                ["  → From Long Trades (HH Break)", adv["choch_long_exits"]],
                ["  → From Short Trades (LL Break)", adv["choch_short_exits"]],
                ["ChoCh Winners",       f"{adv['choch_wins']} ({adv['choch_wr']:.1f}%)"],
                ["ChoCh Avg P&L %",     f"{adv['choch_avg_pnl']:.4f}%"],
                ["", ""],
                ["LOW-POINT TRADES", ""],
                ["Threshold (pts)",     LOW_POINTS_THRESHOLD],
                [f"Trades < {LOW_POINTS_THRESHOLD} pts", adv["low_pt_total"]],
                ["% of Total Trades",  f"{adv['low_pt_total']/total_t*100:.1f}%"],
                ["", ""],
                ["DAILY POINT AVERAGES", ""],
                ["Avg Profit / Day (pts)", adv["avg_profit_per_day"]],
                ["Avg Loss   / Day (pts)", adv["avg_loss_per_day"]],
                ["Highest Profit Day",  f"{adv['highest_profit_day']}  →  +{adv['highest_profit_day_pts']:.2f} pts"],
                ["Highest Loss Day",    f"{adv['highest_loss_day']}  →  {adv['highest_loss_day_pts']:.2f} pts"],
            ]
            pd.DataFrame(adv_rows, columns=["Metric","Value"]).to_excel(
                writer, sheet_name="Advanced Stats", index=False)

            if not adv["low_pt_trades"].empty:
                lp_cols = ["Date","Direction","Entry Time","Entry Price","Exit Time",
                           "Exit Price","Holding Time","Points Captured",
                           "P&L %","Exit Reason","Result"]
                existing = [c for c in lp_cols if c in adv["low_pt_trades"].columns]
                adv["low_pt_trades"][existing].to_excel(
                    writer, sheet_name="Low-Point Trades", index=False)

            if not adv["low_pt_daily"].empty:
                adv["low_pt_daily"].to_excel(writer, sheet_name="Low-Pt Per Day", index=False)

    _style_excel(fname, df_trades_export, df_day, df_ts, df_hold_sum, df_hold_slots)
    print(f"  Excel saved  → {fname}")
    return fname


def _style_excel(fname, df_trades, df_day, df_ts, df_hold_sum=None, df_hold_slots=None):
    GL="C6EFCE"; GD="1A5C38"; RL="FFC7CE"; RD="9C0006"
    BH="1F3864"; YS="FFD700"; GA="F2F2F2"; WH="FFFFFF"
    OR="FCE4D6"; TL="DDEBF7"; PU="EAD1DC"; CY="D6E4F0"  # CY = cyan-ish for ChoCh
    ts  = Side(style="thin", color="CCCCCC")
    bdr = Border(left=ts, right=ts, top=ts, bottom=ts)

    def hdr(ws, row=1, bg=BH, fg=WH):
        for c in ws[row]:
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(bold=True, color=fg, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr
        ws.row_dimensions[row].height = 22

    def aw(ws, cap=30):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+3, cap)

    wb = load_workbook(fname)

    # ST Data
    if "Supertrend Data" in wb.sheetnames:
        ws = wb["Supertrend Data"]; hdr(ws); ws.freeze_panes = "B2"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            v = str(row[-1].value or "")
            bg = GL if v=="Bullish" else RL if v=="Bearish" else WH
            fc = GD if v=="Bullish" else RD if v=="Bearish" else "000000"
            for cell in row:
                cell.fill=PatternFill("solid",fgColor=bg); cell.border=bdr
                cell.alignment=Alignment(horizontal="center")
            row[-1].font=Font(bold=True, color=fc)
        aw(ws)

    # Trade Log — ChoCh exits get a distinct cyan highlight
    if "Trade Log" in wb.sheetnames and df_trades is not None:
        ws = wb["Trade Log"]; hdr(ws); ws.freeze_panes = "A2"
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            res = str(row[-1].value or ""); rsn = str(row[-2].value or "")
            bg  = GL if res=="WIN" else RL if res=="LOSS" else (GA if i%2==0 else WH)
            if res=="WIN":  row[-1].font=Font(bold=True,color=GD)
            if res=="LOSS": row[-1].font=Font(bold=True,color=RD)
            if rsn.startswith("Hard SL"):
                row[-2].fill=PatternFill("solid",fgColor="FF0000")
                row[-2].font=Font(bold=True,color="FFFFFF")
            elif "ChoCh" in rsn:
                row[-2].fill=PatternFill("solid",fgColor="2980B9")
                row[-2].font=Font(bold=True,color="FFFFFF")
            elif "Stop Loss" in rsn:
                row[-2].fill=PatternFill("solid",fgColor=OR)
                row[-2].font=Font(bold=True,color="C55A11")
            elif "EOD" in rsn:
                row[-2].fill=PatternFill("solid",fgColor=TL)
                row[-2].font=Font(bold=True,color="2E75B6")
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000","FFFFFFFF"):
                    cell.fill=PatternFill("solid",fgColor=bg)
                cell.border=bdr; cell.alignment=Alignment(horizontal="center")
        aw(ws)

    # Daily P&L
    if "Daily P&L" in wb.sheetnames and df_day is not None:
        ws = wb["Daily P&L"]; hdr(ws); ws.freeze_panes = "A2"; rc = ws.max_column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            dr = str(row[rc-1].value or "")
            bg = GL if "Profit" in dr else RL if "Loss" in dr else GA
            if "Profit" in dr: row[rc-1].font=Font(bold=True,color=GD)
            elif "Loss"  in dr:row[rc-1].font=Font(bold=True,color=RD)
            for cell in row:
                cell.fill=PatternFill("solid",fgColor=bg); cell.border=bdr
                cell.alignment=Alignment(horizontal="center")
        aw(ws)

    # Time Slot P&L
    if "Time Slot P&L" in wb.sheetnames and df_ts is not None:
        ws = wb["Time Slot P&L"]; hdr(ws, bg="2E4057"); ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            try: pv = float(row[5].value) if row[5].value not in (None,"—") else 0
            except: pv = 0
            bg = GL if pv > 0 else RL if pv < 0 else GA
            vd = str(row[-1].value or "")
            if "🟢" in vd: row[-1].font=Font(bold=True,color=GD)
            elif "🔴" in vd:row[-1].font=Font(bold=True,color=RD)
            for cell in row:
                cell.fill=PatternFill("solid",fgColor=bg); cell.border=bdr
                cell.alignment=Alignment(horizontal="center")
        aw(ws)

    # Holding Time Summary
    if "Holding Time Summary" in wb.sheetnames:
        ws = wb["Holding Time Summary"]; hdr(ws, bg="2E4057")
        HOLD_SECS = {"OVERALL HOLDING TIME","BY RESULT","BY DIRECTION","DURATION DISTRIBUTION"}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            lbl = str(row[0].value or "")
            if lbl in HOLD_SECS:
                for c in row:
                    c.fill=PatternFill("solid",fgColor="2E4057")
                    c.font=Font(bold=True,color=WH,size=10); c.border=bdr
                continue
            for c in row:
                c.border=bdr; c.alignment=Alignment(horizontal="left")
            if len(row) > 1:
                if "Winner" in lbl or "Profit" in lbl:
                    row[1].fill=PatternFill("solid",fgColor=GL)
                elif "Loser" in lbl or "Loss" in lbl:
                    row[1].fill=PatternFill("solid",fgColor=RL)
        ws.column_dimensions["A"].width=32; ws.column_dimensions["B"].width=28

    # Holding By Time Slot
    if "Holding By Time Slot" in wb.sheetnames:
        ws = wb["Holding By Time Slot"]; hdr(ws, bg="2E4057"); ws.freeze_panes="A2"
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            bg = GA if i % 2 == 0 else WH
            for cell in row:
                cell.fill=PatternFill("solid",fgColor=bg); cell.border=bdr
                cell.alignment=Alignment(horizontal="center")
        aw(ws)

    # Summary
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]; hdr(ws)
        SECS = {"STRATEGY INFO","EXIT CONFIG","FILTERS","TRADE STATS","P&L STATS",
                "HOLDING TIME","POINTS","DAY STATS","CHOCH EXIT ANALYSIS"}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            lbl = str(row[0].value or "")
            if lbl in SECS:
                for c in row:
                    c.fill=PatternFill("solid",fgColor=YS)
                    c.font=Font(bold=True,size=10); c.border=bdr
                continue
            for c in row:
                c.border=bdr; c.alignment=Alignment(horizontal="left")
            if len(row)>1 and "Net Points" in lbl:
                try: v=float(row[1].value)
                except: v=0
                row[1].fill=PatternFill("solid",fgColor=GL if v>0 else RL)
                row[1].font=Font(bold=True,color=GD if v>0 else RD)
            # Highlight ChoCh rows in cyan
            if len(row)>1 and "ChoCh" in lbl:
                row[0].fill=PatternFill("solid",fgColor=CY)
                row[1].fill=PatternFill("solid",fgColor=CY)
        ws.column_dimensions["A"].width=32; ws.column_dimensions["B"].width=40

    # Advanced Stats — ChoCh section highlighted
    ADV_SECS = {"DIRECTION BREAKDOWN","STREAK ANALYSIS","LOW-POINT TRADES",
                "DAILY POINT AVERAGES","CHOCH EXIT ANALYSIS"}
    if "Advanced Stats" in wb.sheetnames:
        ws = wb["Advanced Stats"]; hdr(ws, bg="2E4057")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            lbl = str(row[0].value or "")
            if lbl in ADV_SECS:
                for c in row:
                    c.fill=PatternFill("solid",fgColor="2E4057")
                    c.font=Font(bold=True,color=WH,size=10); c.border=bdr
                continue
            for c in row:
                c.border=bdr; c.alignment=Alignment(horizontal="left")
            if len(row) > 1:
                if "Win Streak" in lbl:
                    row[1].fill=PatternFill("solid",fgColor=GL)
                    row[1].font=Font(bold=True,color=GD)
                elif "Loss Streak" in lbl:
                    row[1].fill=PatternFill("solid",fgColor=RL)
                    row[1].font=Font(bold=True,color=RD)
                elif "Avg Profit" in lbl or "Highest Profit" in lbl:
                    row[1].fill=PatternFill("solid",fgColor=GL)
                elif "Avg Loss" in lbl or "Highest Loss" in lbl:
                    row[1].fill=PatternFill("solid",fgColor=RL)
                elif "ChoCh" in lbl:
                    row[0].fill=PatternFill("solid",fgColor=CY)
                    row[1].fill=PatternFill("solid",fgColor=CY)
        ws.column_dimensions["A"].width=36; ws.column_dimensions["B"].width=36

    if "Low-Point Trades" in wb.sheetnames:
        ws = wb["Low-Point Trades"]; hdr(ws, bg="7B2D8B"); ws.freeze_panes="A2"
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            res = str(row[-1].value or "")
            bg  = GL if res=="WIN" else RL if res=="LOSS" else (GA if i%2==0 else WH)
            if res=="WIN":  row[-1].font=Font(bold=True,color=GD)
            if res=="LOSS": row[-1].font=Font(bold=True,color=RD)
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000","FFFFFFFF"):
                    cell.fill=PatternFill("solid",fgColor=bg)
                cell.border=bdr; cell.alignment=Alignment(horizontal="center")
        aw(ws)

    if "Low-Pt Per Day" in wb.sheetnames:
        ws = wb["Low-Pt Per Day"]; hdr(ws, bg="7B2D8B"); ws.freeze_panes="A2"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill=PatternFill("solid",fgColor=GA); cell.border=bdr
                cell.alignment=Alignment(horizontal="center")
        aw(ws)

    wb.save(fname)


# ───────────────────────────────────────────────────────────────────────────────
#  CHART
# ───────────────────────────────────────────────────────────────────────────────
def build_chart(df, trades, htf_enabled):
    if CHART_DAYS:
        udays  = sorted(df.index.normalize().unique())
        cutoff = udays[-CHART_DAYS] if len(udays) >= CHART_DAYS else udays[0]
        df_c   = df[df.index.normalize() >= cutoff].copy()
    else:
        df_c = df.copy()

    fig = go.Figure()
    fig.add_trace(go.Candlestick(
        x=df_c.index, open=df_c["Open"], high=df_c["High"],
        low=df_c["Low"], close=df_c["Close"], name="Price",
        increasing_line_color="#26a69a", decreasing_line_color="#ef5350"))
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["ST_val"].where(df_c["ST_bull"]),
        name=f"ST Bullish ({ST_PERIOD},{ST_MULTIPLIER})",
        mode="lines", line=dict(color="#22c55e", width=2.5), connectgaps=False))
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["ST_val"].where(~df_c["ST_bull"]),
        name=f"ST Bearish ({ST_PERIOD},{ST_MULTIPLIER})",
        mode="lines", line=dict(color="#ef4444", width=2.5), connectgaps=False))

    if htf_enabled and "HTF_bull" in df_c.columns:
        htf_bull_bool = df_c["HTF_bull"].astype(bool)
        mid = (df_c["High"] + df_c["Low"]) / 2
        fig.add_trace(go.Scatter(x=df_c.index, y=mid.where(htf_bull_bool),
            name="HTF Bullish", mode="lines",
            line=dict(color="#86efac", width=1.2, dash="dot"), connectgaps=False))
        fig.add_trace(go.Scatter(x=df_c.index, y=mid.where(~htf_bull_bool),
            name="HTF Bearish", mode="lines",
            line=dict(color="#fca5a5", width=1.2, dash="dot"), connectgaps=False))

    if trades:
        df_t = pd.DataFrame(trades); start = df_c.index[0]
        def parse(col):
            return pd.to_datetime(df_t[col], format="%Y-%m-%d %H:%M").dt.tz_localize(IST)
        et=parse("Entry Time"); xt=parse("Exit Time")
        ep=df_t["Entry Price"]; xp=df_t["Exit Price"]
        dr=df_t["Direction"];   rs=df_t["Result"]; ex=df_t["Exit Reason"]
        mask = et >= start

        # Separate ChoCh exits for special marker
        choch_mask = ex.str.contains("ChoCh", na=False)

        for xv,yv,nm,sym,col,ec in [
            (et[mask&dr.str.contains("Long")].tolist(),  ep[mask&dr.str.contains("Long")].tolist(),  "Long Entry",  "triangle-up",   "#22c55e","white"),
            (et[mask&dr.str.contains("Short")].tolist(), ep[mask&dr.str.contains("Short")].tolist(), "Short Entry", "triangle-down", "#ef4444","white"),
            (xt[mask&(rs=="WIN")&~choch_mask].tolist(),  xp[mask&(rs=="WIN")&~choch_mask].tolist(),  "Exit WIN",    "circle",   "#86efac","#16a34a"),
            (xt[mask&(rs=="LOSS")&~choch_mask].tolist(), xp[mask&(rs=="LOSS")&~choch_mask].tolist(), "Exit LOSS",   "x",        "#fca5a5","#dc2626"),
            (xt[mask&ex.str.startswith("Hard SL")].tolist(), xp[mask&ex.str.startswith("Hard SL")].tolist(), f"Hard SL ({HARD_SL_PCT}%)", "hexagram","#ff0000","white"),
        ]:
            if xv: fig.add_trace(go.Scatter(x=xv, y=yv, mode="markers", name=nm,
                       marker=dict(symbol=sym,size=13,color=col,line=dict(color=ec,width=1.5))))

        # ChoCh exits — distinct diamond marker in cyan
        choch_x = xt[mask & choch_mask].tolist()
        choch_y = xp[mask & choch_mask].tolist()
        if choch_x:
            fig.add_trace(go.Scatter(
                x=choch_x, y=choch_y, mode="markers",
                name="ChoCh Exit",
                marker=dict(symbol="diamond", size=14, color="#00BCD4",
                            line=dict(color="white", width=1.5))))

    htf_lbl = f"HTF ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER}) | " if htf_enabled else "HTF: OFF | "
    choch_lbl = f"ChoCh Exit: ON (lb={CHOCH_LOOKBACK})" if CHOCH_EXIT_ENABLED else "ChoCh: OFF"
    fig.update_layout(
        template="plotly_dark", height=820, xaxis_rangeslider_visible=False,
        legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="right", x=1, font=dict(size=9)),
        margin=dict(l=60,r=40,t=120,b=50), font=dict(family="monospace",size=11),
        title=dict(
            text=(f"<b>{SYMBOL}</b>  |  Supertrend({ST_PERIOD},{ST_MULTIPLIER})<br>"
                  f"<span style='font-size:10px;color:#94a3b8'>"
                  f"{htf_lbl}Window={'ON' if TRADE_WINDOW_ENABLED else 'OFF'} | "
                  f"ConfirmN={CONFIRM_CANDLES} | HardSL={HARD_SL_PCT}% | {choch_lbl}</span>"),
            x=0.5, xanchor="center"))
    fig.update_yaxes(title_text="Price", showgrid=True, gridcolor="#1e293b")
    fig.update_xaxes(showgrid=True, gridcolor="#1e293b",
        rangebreaks=[dict(bounds=["sat","mon"]), dict(bounds=[15.5,9.25], pattern="hour")])

    chart_name = f"{SYMBOL}_chart_v3.html"
    fig.write_html(chart_name)
    print(f"  Chart saved  → {chart_name}  (open in browser)")


# ───────────────────────────────────────────────────────────────────────────────
#  MAIN
# ───────────────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═"*60)
    print("  SUPERTREND BACKTEST v3.1  —  CSV / yfinance Mode")
    print("  + ChoCh (Change of Character) Exit")
    print("═"*60)
    try:
        df, htf_enabled = fetch_data()
    except (ValueError, FileNotFoundError, ImportError) as e:
        print(e); return

    trades = run_backtest(df, htf_enabled)
    df_trades, df_day, df_ts, adv, df_hold_sum, df_hold_slots = print_results(trades, htf_enabled)

    print("  Exporting Excel …")
    export_excel(df, df_trades, df_day, df_ts, htf_enabled, adv, df_hold_sum, df_hold_slots)

    print("  Building chart …")
    build_chart(df, trades, htf_enabled)
    print("\n  ✅  Done.\n")


if __name__ == "__main__":
    main()