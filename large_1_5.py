"""
Supertrend Intraday Backtesting Strategy — v3.0
─────────────────────────────────────────────────────────────────────────────
DATA SOURCE : CSV files only (no internet / yfinance required)

  BASE_CSV   — your 1m (or any TF) OHLC data
  HTF_CSV    — your higher timeframe OHLC data (e.g. 5m)
               Set HTF_CSV = None  →  HTF filter disabled automatically

CSV FORMAT  : Standard columns (case-insensitive, auto-detected)
  datetime, open, high, low, close, volume
  DateTime column accepts any standard format (auto-parsed)

FILTERS:
  1. HTF Filter        — only trade in direction of HTF Supertrend
                         (auto-disabled if HTF_CSV = None)
  2. Trade Window      — only enter during profitable time slots
  3. Confirm Candles   — wait N candles after ST flip before entering
  4. Min Gap           — skip if price too close to ST line (choppy)
  5. Volume            — require volume > N × 20-bar average
  6. Max Trades/Day    — hard cap on daily entries

ENTRIES:
  Long  : ST below price  →  enter at close
  Short : ST above price  →  enter at close

EXITS (whichever fires first):
  Hard SL      →  fixed % loss from entry
  ST SL        →  price crosses ST line (base TF or HTF ST depending on config)
  EOD          →  force exit at 15:15 IST

EXIT IMPROVEMENTS (v3.1):
  EXIT_USE_HTF_ST   — use the wider HTF ST line as exit instead of tight 1m ST
                      lets the trade breathe, captures bigger moves
  CLOSE_BASED_EXIT  — exit only when candle CLOSES beyond ST line, not just
                      when intracandle Low/High touches it (reduces whipsaws)

OUTPUTS:
  Console  : trade log + per-day P&L + time-slot P&L + summary
  Excel    : 5 sheets — Supertrend Data, Trade Log, Daily P&L,
             Time Slot P&L, Summary
  HTML     : interactive candlestick + ST chart

Install: pip install pandas plotly openpyxl pytz numpy
"""

import pandas as pd
import numpy as np
import plotly.graph_objects as go
import pytz
import os
from datetime import time as dtime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════════════════════
#  ██████╗ ██████╗ ███╗   ██╗███████╗██╗ ██████╗
#  ██╔════╝██╔═══██╗████╗  ██║██╔════╝██║██╔════╝
#  ██║     ██║   ██║██╔██╗ ██║█████╗  ██║██║  ███╗
#  ██║     ██║   ██║██║╚██╗██║██╔══╝  ██║██║   ██║
#  ╚██████╗╚██████╔╝██║ ╚████║██║     ██║╚██████╔╝
#   ╚═════╝ ╚═════╝ ╚═╝  ╚═══╝╚═╝     ╚═╝ ╚═════╝
# ═══════════════════════════════════════════════════════════════════════════════

# ── CSV File Paths ─────────────────────────────────────────────────────────────
BASE_CSV = "Book1min.csv"        # ← Your base timeframe CSV  (e.g. 1m data)
HTF_CSV  = "Book1.csv"         # ← Your HTF CSV             (e.g. 5m data)
                                  #   Set to None to auto-disable HTF filter:
                                  #   HTF_CSV = None

# Label shown in all outputs
SYMBOL   = "NIFTY"              # e.g. "NIFTY", "BANKNIFTY", "RELIANCE"

# ── Supertrend Settings ────────────────────────────────────────────────────────
ST_PERIOD         = 14           # Base TF supertrend period
ST_MULTIPLIER     = 3.5          # Base TF supertrend multiplier

HTF_ST_PERIOD     = 10           # HTF supertrend period
HTF_ST_MULTIPLIER = 3.0          # HTF supertrend multiplier

# ── Hard Stop Loss ─────────────────────────────────────────────────────────────
HARD_SL_ENABLED = True
HARD_SL_PCT     = 0.25           # Max loss % from entry before hard exit

# ── Exit Improvement 1: Use HTF ST as Exit Line ───────────────────────────────
EXIT_USE_HTF_ST = False          # True  → exit when price crosses the HTF ST line
                                 #         (wider buffer, lets winners run further)
                                 # False → exit when price crosses the base 1m ST line
                                 #         (original tight behaviour)
                                 # NOTE: requires HTF_CSV to be set. If HTF data is
                                 #       unavailable, automatically falls back to base ST.

# ── Exit Improvement 2: Close-Based Exit ──────────────────────────────────────
CLOSE_BASED_EXIT = True         # True  → exit only when candle CLOSES beyond ST line
                                 #         (ignores intracandle wicks, fewer whipsaws)
                                 # False → exit when candle Low/High touches ST line
                                 #         (original intracandle behaviour)
                                 # NOTE: Hard SL always remains intracandle (Low/High)
                                 #       regardless of this setting — it is a safety net.

# ── Filter 1: HTF Confirmation ─────────────────────────────────────────────────
#    Automatically ON  when HTF_CSV is a valid path
#    Automatically OFF when HTF_CSV = None  (no extra toggle needed)

# ── Filter 2: Trade Window ─────────────────────────────────────────────────────
TRADE_WINDOW_ENABLED = True
TRADE_WINDOWS = [                # Only enter trades inside these windows
    (dtime(9, 30), dtime(11, 30)),    # Morning momentum  ← best window
    (dtime(14,  0), dtime(15, 10)),   # Pre-close momentum
]
# Set TRADE_WINDOW_ENABLED = False to trade the full session

# ── Filter 3: Confirmation Candles ────────────────────────────────────────────
CONFIRM_CANDLES = 10              # Wait this many candles after an ST flip
                                  # before entering. Set to 1 to enter immediately.

# ── Filter 4: Min Gap (anti-chop) ─────────────────────────────────────────────
MIN_GAP_ENABLED = True
MIN_GAP_PCT     = 0.15           # Skip entry if |price − ST| / price < this %

# ── Filter 5: Volume ──────────────────────────────────────────────────────────
VOLUME_FILTER_ENABLED = False    # Set True if your CSV has meaningful volume
VOLUME_MULTIPLIER     = 1.5      # Volume must exceed this × 20-bar average
VOLUME_LOOKBACK       = 20

# ── Filter 6: Max Trades Per Day ───────────────────────────────────────────────
MAX_TRADES_PER_DAY_ENABLED = True
MAX_TRADES_PER_DAY         = 5

# ── Intraday Session (NSE) ─────────────────────────────────────────────────────
MARKET_OPEN  = dtime(9, 15)
MARKET_CLOSE = dtime(15, 15)
IST          = pytz.timezone("Asia/Kolkata")

# ── Chart ──────────────────────────────────────────────────────────────────────
CHART_DAYS = 5                   # Recent trading days to plot (None = all)

# ═══════════════════════════════════════════════════════════════════════════════
#  END OF CONFIG
# ═══════════════════════════════════════════════════════════════════════════════


def fmt(ts):
    return ts.strftime("%Y-%m-%d %H:%M") if pd.notna(ts) else "—"


# ───────────────────────────────────────────────────────────────────────────────
#  SUPERTREND ENGINE
# ───────────────────────────────────────────────────────────────────────────────
def compute_supertrend(df, period, multiplier):
    """Wilder ATR Supertrend. Adds ST_val, ST_bull, ST_direction columns."""
    high  = df["High"].values.astype(float)
    low   = df["Low"].values.astype(float)
    close = df["Close"].values.astype(float)
    n     = len(df)

    tr = np.maximum(high[1:] - low[1:],
         np.maximum(np.abs(high[1:] - close[:-1]),
                    np.abs(low[1:]  - close[:-1])))

    atr = np.full(n, np.nan)
    if n > period:
        atr[period] = tr[:period].mean()
        for j in range(period + 1, n):
            atr[j] = (atr[j-1] * (period - 1) + tr[j-1]) / period

    hl2         = (high + low) / 2.0
    upper_basic = hl2 + multiplier * atr
    lower_basic = hl2 - multiplier * atr

    upper = np.full(n, np.nan)
    lower = np.full(n, np.nan)
    bull  = np.full(n, True, dtype=bool)

    for j in range(period, n):
        if np.isnan(upper[j-1]):
            upper[j] = upper_basic[j];  lower[j] = lower_basic[j]
            bull[j]  = close[j] >= lower[j];  continue
        upper[j] = (upper_basic[j]
                    if upper_basic[j] < upper[j-1] or close[j-1] > upper[j-1]
                    else upper[j-1])
        lower[j] = (lower_basic[j]
                    if lower_basic[j] > lower[j-1] or close[j-1] < lower[j-1]
                    else lower[j-1])
        bull[j] = (close[j] >= lower[j]) if bull[j-1] else (close[j] > upper[j])

    df = df.copy()
    df["ST_val"]       = np.where(bull, lower, upper)
    df["ST_bull"]      = bull
    df["ST_direction"] = np.where(bull, "Bullish", "Bearish")
    return df


# ───────────────────────────────────────────────────────────────────────────────
#  CSV LOADER
# ───────────────────────────────────────────────────────────────────────────────
_DT_CANDS = ["datetime","date","time","timestamp","date/time","date time"]
_O_CANDS  = ["open","o"]
_H_CANDS  = ["high","h"]
_L_CANDS  = ["low","l"]
_C_CANDS  = ["close","ltp","last","c"]
_V_CANDS  = ["volume","vol","v"]


def _find_col(columns, candidates):
    lmap = {c.lower().strip(): c for c in columns}
    for cand in candidates:
        if cand in lmap:
            return lmap[cand]
    return None


def load_csv(path, label="CSV"):
    """
    Load OHLC data from a CSV file.
    - Auto-detects column names (case-insensitive)
    - Auto-parses datetime in any standard format
    - Localizes index to IST
    - Filters to market hours
    Returns a clean DataFrame ready for Supertrend computation.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"\n  ❌  File not found : '{path}'\n"
            f"      ↳ Check the BASE_CSV / HTF_CSV paths in the CONFIG section.\n"
        )

    print(f"\n  Loading {label}: {path} …")
    df = pd.read_csv(path)
    print(f"    Shape: {df.shape}  |  Columns: {list(df.columns)}")

    dt_col = _find_col(df.columns, _DT_CANDS)
    o_col  = _find_col(df.columns, _O_CANDS)
    h_col  = _find_col(df.columns, _H_CANDS)
    l_col  = _find_col(df.columns, _L_CANDS)
    c_col  = _find_col(df.columns, _C_CANDS)
    v_col  = _find_col(df.columns, _V_CANDS)

    print(f"    Mapped  → dt='{dt_col}' | open='{o_col}' | high='{h_col}' | "
          f"low='{l_col}' | close='{c_col}' | volume='{v_col or 'not found'}'")

    missing = [n for n, c in [("datetime", dt_col), ("open", o_col),
                                ("high", h_col), ("low", l_col), ("close", c_col)]
               if c is None]
    if missing:
        raise ValueError(
            f"\n  ❌  Could not detect required columns: {missing}\n"
            f"      Columns in file : {list(df.columns)}\n"
            f"      Please rename to: datetime, open, high, low, close, volume\n"
        )

    rename = {dt_col: "__dt", o_col: "Open", h_col: "High",
              l_col: "Low",   c_col: "Close"}
    if v_col:
        rename[v_col] = "Volume"

    df = df[list(rename.keys())].rename(columns=rename)

    # Parse and localize datetime
    df["__dt"] = pd.to_datetime(
    df["__dt"],
    format="%d-%m-%Y %H:%M",
    errors="coerce"
)
    df = df.set_index("__dt")
    df.index.name = "Datetime"

    if df.index.tz is None:
        df.index = df.index.tz_localize(IST, ambiguous="infer",
                                         nonexistent="shift_forward")
    else:
        df.index = df.index.tz_convert(IST)

    df.sort_index(inplace=True)

    # Numeric coercion
    for col in ["Open", "High", "Low", "Close"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    if "Volume" in df.columns:
        df["Volume"] = pd.to_numeric(df["Volume"], errors="coerce").fillna(0)
    else:
        df["Volume"] = 0.0

    df.dropna(subset=["Open","High","Low","Close"], inplace=True)

    # Market hours filter
    tod = df.index.time
    df  = df[(tod >= MARKET_OPEN) & (tod <= MARKET_CLOSE)].copy()

    print(f"    Clean rows: {len(df)}  |  "
          f"{fmt(df.index[0])} → {fmt(df.index[-1])}")
    return df


# ───────────────────────────────────────────────────────────────────────────────
#  FETCH / PREPARE ALL DATA
# ───────────────────────────────────────────────────────────────────────────────
def fetch_data():
    # Load & compute ST on base TF
    df = load_csv(BASE_CSV, label="Base TF CSV")
    df = compute_supertrend(df, ST_PERIOD, ST_MULTIPLIER)
    df.dropna(inplace=True)

    htf_enabled = False

    if HTF_CSV is not None:
        try:
            df_htf = load_csv(HTF_CSV, label="HTF CSV")
            df_htf = compute_supertrend(df_htf, HTF_ST_PERIOD, HTF_ST_MULTIPLIER)
            df_htf.dropna(inplace=True)

            # Forward-fill HTF ST_bull and ST_val onto the base TF index
            htf_bull = df_htf["ST_bull"].reindex(df.index, method="ffill")
            htf_val  = df_htf["ST_val"].reindex(df.index, method="ffill")
            coverage = htf_bull.notna().mean()

            if coverage < 0.5:
                print(f"\n  ⚠  HTF index overlap only {coverage:.0%} — "
                      f"HTF filter DISABLED (check date ranges match).")
            else:
                df["HTF_bull"] = htf_bull
                df["HTF_val"]  = htf_val          # ← HTF ST price level for exit
                df["HTF_bull"] = df["HTF_bull"].ffill()
                df["HTF_val"]  = df["HTF_val"].ffill()
                df.dropna(subset=["HTF_bull", "HTF_val"], inplace=True)
                htf_enabled = True
                exit_mode = "HTF ST exit" if EXIT_USE_HTF_ST else "Base ST exit"
                close_mode = "close-based" if CLOSE_BASED_EXIT else "intracandle"
                print(f"\n  ✅  HTF ENABLED — ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER}) | "
                      f"Coverage: {coverage:.0%} | "
                      f"Exit mode: {exit_mode} | {close_mode} | "
                      f"Base candles: {len(df)}")

        except FileNotFoundError as e:
            print(e)
            print("  ⚠  HTF filter DISABLED (file not found).")
    else:
        print("\n  ℹ  HTF_CSV = None  →  HTF filter DISABLED.")

    # When HTF filter is off, fill columns with neutral values
    if not htf_enabled:
        df["HTF_bull"] = True
        df["HTF_val"]  = df["ST_val"]   # fallback: use base ST as exit line

    return df, htf_enabled


# ───────────────────────────────────────────────────────────────────────────────
#  BACKTEST ENGINE
# ───────────────────────────────────────────────────────────────────────────────
def run_backtest(df, htf_enabled):
    trades = []

    in_trade      = False
    direction     = None
    entry_price   = None
    entry_time    = None
    peak          = None
    hard_sl_price = None
    flip_candle   = -999

    closes    = df["Close"].values.astype(float)
    highs     = df["High"].values.astype(float)
    lows      = df["Low"].values.astype(float)
    st_vals   = df["ST_val"].values.astype(float)
    htf_vals  = df["HTF_val"].values.astype(float)   # HTF ST price level
    st_bulls  = df["ST_bull"].values.astype(bool)
    htf_bulls = df["HTF_bull"].values.astype(bool)
    volumes   = df["Volume"].values.astype(float)
    times     = df.index

    # Decide which ST line to use for exits
    # EXIT_USE_HTF_ST=True  → use wider HTF ST value
    # EXIT_USE_HTF_ST=False → use tight base 1m ST value
    exit_st = htf_vals if (EXIT_USE_HTF_ST and htf_enabled) else st_vals

    vol_avg      = pd.Series(volumes).rolling(VOLUME_LOOKBACK).mean().values
    daily_trades = {}   # {date: count}

    def in_window(t):
        if not TRADE_WINDOW_ENABLED:
            return True
        return any(s <= t <= e for s, e in TRADE_WINDOWS)

    def record(dir_, e_time, e_price, x_time, x_price, pk, sl_at_exit, reason, hard_sl):
        pnl = round(((x_price - e_price) / e_price * 100) if dir_ == "long"
                    else ((e_price - x_price) / e_price * 100), 4)
        trades.append({
            "Date"            : e_time.strftime("%Y-%m-%d"),
            "Direction"       : "Long  ↑" if dir_ == "long" else "Short ↓",
            "Entry Time"      : fmt(e_time),
            "Entry Price"     : round(e_price, 2),
            "Hard SL Price"   : round(hard_sl, 2) if hard_sl is not None else "OFF",
            "Exit ST Line"    : round(sl_at_exit, 2),
            "Exit Time"       : fmt(x_time),
            "Exit Price"      : round(x_price, 2),
            "Peak"            : round(pk, 2),
            "Points Captured" : round(abs(x_price - e_price), 4),
            "P&L %"           : pnl,
            "Exit Reason"     : reason,
            "Result"          : "WIN" if pnl > 0 else "LOSS",
        })

    def open_trade(dir_, price, time_, high_, low_, date_key):
        nonlocal in_trade, direction, entry_price, entry_time, peak, hard_sl_price
        in_trade      = True
        direction     = dir_
        entry_price   = price
        entry_time    = time_
        peak          = high_ if dir_ == "long" else low_
        hard_sl_price = (round(price * (1 - HARD_SL_PCT / 100), 4) if dir_ == "long"
                         else round(price * (1 + HARD_SL_PCT / 100), 4)) \
                        if HARD_SL_ENABLED else None
        daily_trades[date_key] = daily_trades.get(date_key, 0) + 1

    for i in range(1, len(df)):
        c_close    = closes[i];     c_high  = highs[i];      c_low  = lows[i]
        c_time     = times[i];      c_tod   = c_time.time(); c_st   = st_vals[i]
        c_exit_st  = exit_st[i]     # ← exit line: HTF ST or base ST depending on config
        c_bull     = st_bulls[i];   prev_bull = st_bulls[i-1]
        c_htf_bull = htf_bulls[i];  date_key  = c_time.date()

        flipped_bull = (not prev_bull) and c_bull
        flipped_bear = prev_bull and (not c_bull)
        if flipped_bull or flipped_bear:
            flip_candle = i

        # ── EOD force exit ─────────────────────────────────────────────────────
        if in_trade and c_tod >= MARKET_CLOSE:
            pk_f = max(peak, c_high) if direction == "long" else min(peak, c_low)
            record(direction, entry_time, entry_price,
                   c_time, c_close, pk_f, c_exit_st, "EOD Exit", hard_sl_price)
            in_trade = False
            continue

        if c_tod < MARKET_OPEN or c_tod >= MARKET_CLOSE:
            continue

        # ── Manage open trade ──────────────────────────────────────────────────
        if in_trade:
            if direction == "long":
                if c_high > peak: peak = c_high

                # Hard SL: always intracandle (safety net — never delayed)
                hard_hit = HARD_SL_ENABLED and hard_sl_price and c_low <= hard_sl_price

                # ST SL: use exit_st line (HTF or base), respect CLOSE_BASED_EXIT
                if CLOSE_BASED_EXIT:
                    # Exit only if candle CLOSES below the exit ST line
                    st_cross  = c_close < c_exit_st
                    st_flip   = flipped_bear   # base TF flip still counts
                    st_hit    = st_cross or st_flip
                else:
                    # Original: exit when intracandle Low touches the exit ST line
                    st_hit    = c_low <= c_exit_st or flipped_bear

                if hard_hit or st_hit:
                    if hard_hit:
                        reason = f"Hard SL ({HARD_SL_PCT}%)"; exit_px = hard_sl_price
                    elif (CLOSE_BASED_EXIT and c_close < c_exit_st) or (not CLOSE_BASED_EXIT and c_low <= c_exit_st):
                        exit_label = "HTF ST Exit" if (EXIT_USE_HTF_ST and htf_enabled) else "ST Stop Loss"
                        exit_price_level = c_exit_st
                        reason = exit_label; exit_px = exit_price_level
                    else:
                        reason = "ST Flip Bear"; exit_px = c_exit_st
                    record("long", entry_time, entry_price,
                           c_time, round(exit_px, 2), peak, c_exit_st, reason, hard_sl_price)
                    in_trade = False
                    # Re-enter short on clean base TF flip only
                    if (not hard_hit and not c_bull
                            and (not htf_enabled or not c_htf_bull)
                            and in_window(c_tod)
                            and daily_trades.get(date_key, 0) < MAX_TRADES_PER_DAY):
                        open_trade("short", c_close, c_time, c_high, c_low, date_key)

            elif direction == "short":
                if c_low < peak: peak = c_low

                # Hard SL: always intracandle
                hard_hit = HARD_SL_ENABLED and hard_sl_price and c_high >= hard_sl_price

                # ST SL: use exit_st line, respect CLOSE_BASED_EXIT
                if CLOSE_BASED_EXIT:
                    st_cross  = c_close > c_exit_st
                    st_flip   = flipped_bull
                    st_hit    = st_cross or st_flip
                else:
                    st_hit    = c_high >= c_exit_st or flipped_bull

                if hard_hit or st_hit:
                    if hard_hit:
                        reason = f"Hard SL ({HARD_SL_PCT}%)"; exit_px = hard_sl_price
                    elif (CLOSE_BASED_EXIT and c_close > c_exit_st) or (not CLOSE_BASED_EXIT and c_high >= c_exit_st):
                        exit_label = "HTF ST Exit" if (EXIT_USE_HTF_ST and htf_enabled) else "ST Stop Loss"
                        reason = exit_label; exit_px = c_exit_st
                    else:
                        reason = "ST Flip Bull"; exit_px = c_exit_st
                    record("short", entry_time, entry_price,
                           c_time, round(exit_px, 2), peak, c_exit_st, reason, hard_sl_price)
                    in_trade = False
                    # Re-enter long on clean base TF flip only
                    if (not hard_hit and c_bull
                            and (not htf_enabled or c_htf_bull)
                            and in_window(c_tod)
                            and daily_trades.get(date_key, 0) < MAX_TRADES_PER_DAY):
                        open_trade("long", c_close, c_time, c_high, c_low, date_key)
            continue

        # ── Entry filters (not in trade) ───────────────────────────────────────
        if not in_window(c_tod):                                              continue
        if (i - flip_candle) < CONFIRM_CANDLES:                              continue
        if MIN_GAP_ENABLED and abs(c_close-c_st)/c_close*100 < MIN_GAP_PCT: continue
        if (VOLUME_FILTER_ENABLED and not np.isnan(vol_avg[i])
                and vol_avg[i] > 0
                and volumes[i] < VOLUME_MULTIPLIER * vol_avg[i]):            continue
        if (MAX_TRADES_PER_DAY_ENABLED
                and daily_trades.get(date_key, 0) >= MAX_TRADES_PER_DAY):   continue

        if c_bull  and (not htf_enabled or c_htf_bull):
            open_trade("long",  c_close, c_time, c_high, c_low, date_key)
        elif not c_bull and (not htf_enabled or not c_htf_bull):
            open_trade("short", c_close, c_time, c_high, c_low, date_key)

    return trades


# ───────────────────────────────────────────────────────────────────────────────
#  SUMMARY BUILDERS
# ───────────────────────────────────────────────────────────────────────────────
def build_daily_summary(trades):
    if not trades: return pd.DataFrame()
    df_t = pd.DataFrame(trades); rows = []
    for date, grp in df_t.groupby("Date"):
        total = len(grp); wins = (grp["P&L %"] > 0).sum(); pnl = grp["P&L %"].sum()
        wp = grp.loc[grp["P&L %"] > 0,  "Points Captured"].sum()
        lp = grp.loc[grp["P&L %"] <= 0, "Points Captured"].sum()
        rows.append({
            "Date"           : date,
            "Total Trades"   : total,
            "Longs"          : grp["Direction"].str.contains("Long").sum(),
            "Shorts"         : grp["Direction"].str.contains("Short").sum(),
            "Winners"        : wins,
            "Losers"         : total - wins,
            "Win Rate %"     : round(wins / total * 100, 1),
            "Total P&L %"    : round(pnl, 4),
            "Best Trade %"   : round(grp["P&L %"].max(), 4),
            "Worst Trade %"  : round(grp["P&L %"].min(), 4),
            "Points Captured": round(wp, 2),
            "Points Lost"    : round(lp, 2),
            "Net Points"     : round(wp - lp, 2),
            "Day Result"     : "✅ Profit" if pnl > 0 else "❌ Loss" if pnl < 0 else "⚖ Flat",
        })
    return pd.DataFrame(rows)


def build_time_slot_summary(trades):
    if not trades: return pd.DataFrame()
    SLOTS = [
        ("09:15","09:30"),("09:30","10:00"),("10:00","10:30"),("10:30","11:00"),
        ("11:00","11:30"),("11:30","12:00"),("12:00","12:30"),("12:30","13:00"),
        ("13:00","13:30"),("13:30","14:00"),("14:00","14:30"),("14:30","15:00"),
        ("15:00","15:15"),
    ]
    df_t = pd.DataFrame(trades)
    df_t["Entry_tod"] = (pd.to_datetime(df_t["Entry Time"], format="%Y-%m-%d %H:%M")
                         .dt.strftime("%H:%M"))
    rows = []
    for s, e in SLOTS:
        grp = df_t[(df_t["Entry_tod"] >= s) & (df_t["Entry_tod"] < e)]
        if grp.empty:
            rows.append({"Time Slot": f"{s}–{e}", "Trades": 0, "Winners": 0,
                         "Losers": 0, "Win Rate %": "—", "Total P&L %": 0.0,
                         "Avg P&L %": 0.0, "Best %": 0.0, "Worst %": 0.0, "Verdict": "—"})
            continue
        total = len(grp); wins = (grp["P&L %"] > 0).sum()
        pnl = grp["P&L %"].sum(); wr = wins / total * 100
        rows.append({
            "Time Slot"  : f"{s}–{e}",
            "Trades"     : total,
            "Winners"    : wins,
            "Losers"     : total - wins,
            "Win Rate %" : round(wr, 1),
            "Total P&L %": round(pnl, 4),
            "Avg P&L %"  : round(grp["P&L %"].mean(), 4),
            "Best %"     : round(grp["P&L %"].max(), 4),
            "Worst %"    : round(grp["P&L %"].min(), 4),
            "Verdict"    : "🟢 Trade" if pnl > 0 and wr >= 50 else "🔴 Avoid",
        })
    return pd.DataFrame(rows)


# ───────────────────────────────────────────────────────────────────────────────
#  PRINT RESULTS
# ───────────────────────────────────────────────────────────────────────────────
def print_results(trades, htf_enabled):
    SEP = "═" * 120; DASH = "─" * 120
    htf_info = (f"ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER}) from '{HTF_CSV}'"
                if htf_enabled else "OFF (HTF_CSV = None or file not found)")

    exit_st_info = (f"HTF ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER})" 
                    if (EXIT_USE_HTF_ST and htf_enabled) else f"Base ST({ST_PERIOD},{ST_MULTIPLIER})")
    close_info   = "Close-based (no wick exits)" if CLOSE_BASED_EXIT else "Intracandle (Low/High)"

    print("\n" + SEP)
    print(f"  {SYMBOL}  |  Entry ST({ST_PERIOD},{ST_MULTIPLIER})  |  EOD: {MARKET_CLOSE}")
    print(f"  Base CSV  : {BASE_CSV}")
    print(f"  HTF       : {htf_info}")
    print(f"  Exit ST   : {exit_st_info}  ← exit triggered against this ST line")
    print(f"  Exit Mode : {close_info}")
    print(f"  Filters   : Window={'ON' if TRADE_WINDOW_ENABLED else 'OFF'}  "
          f"| ConfirmN={CONFIRM_CANDLES}  "
          f"| MinGap={'ON('+str(MIN_GAP_PCT)+'%)' if MIN_GAP_ENABLED else 'OFF'}  "
          f"| Volume={'ON' if VOLUME_FILTER_ENABLED else 'OFF'}  "
          f"| MaxTrades/Day={MAX_TRADES_PER_DAY if MAX_TRADES_PER_DAY_ENABLED else 'OFF'}")
    print(f"  Hard SL   : {'ON — '+str(HARD_SL_PCT)+'% from entry (always intracandle)' if HARD_SL_ENABLED else 'OFF'}")
    print(SEP)

    if not trades:
        print("  ⚠  No trades found. Try relaxing filters or checking CSV date range.")
        print(SEP); return None, None, None

    df_t   = pd.DataFrame(trades)
    df_day = build_daily_summary(trades)
    df_ts  = build_time_slot_summary(trades)

    print("\n  TRADE LOG"); print(DASH)
    cols = ["Date","Direction","Entry Time","Entry Price","Hard SL Price",
            "Exit ST Line","Exit Time","Exit Price","Points Captured",
            "P&L %","Exit Reason","Result"]
    print(df_t[cols].to_string(index=False))

    print("\n\n" + SEP); print("  PER-DAY P&L BREAKDOWN"); print(DASH)
    print(df_day.to_string(index=False))

    print("\n\n" + SEP)
    print("  TIME-SLOT P&L ANALYSIS  (entry time grouped into 30-min windows)")
    print(DASH); print(df_ts.to_string(index=False))

    total = len(df_t); wins = (df_t["P&L %"] > 0).sum(); losses = total - wins
    win_pts  = df_t.loc[df_t["P&L %"] > 0,  "Points Captured"].sum()
    loss_pts = df_t.loc[df_t["P&L %"] <= 0, "Points Captured"].sum()
    p_days   = (df_day["Total P&L %"] > 0).sum()
    l_days   = (df_day["Total P&L %"] < 0).sum()
    best_d   = df_day.loc[df_day["Total P&L %"].idxmax()]
    wrst_d   = df_day.loc[df_day["Total P&L %"].idxmin()]
    best_s   = df_ts.loc[df_ts["Total P&L %"].idxmax()] if not df_ts.empty else None
    wrst_s   = df_ts.loc[df_ts["Total P&L %"].idxmin()] if not df_ts.empty else None

    print("\n\n" + SEP); print("  OVERALL SUMMARY"); print(DASH)
    print(f"  {'EXIT CONFIGURATION':─<55}")
    print(f"  Exit ST Line        : {exit_st_info}")
    print(f"  Exit Trigger        : {close_info}")
    print(f"  Hard SL             : {'ON ('+str(HARD_SL_PCT)+'%) — always intracandle' if HARD_SL_ENABLED else 'OFF'}")
    print(DASH)
    print(f"  {'TRADE STATS':─<55}")
    print(f"  Total Trades        : {total}")
    print(f"  Winners             : {wins}  ({wins/total*100:.1f}%)")
    print(f"  Losers              : {losses}  ({losses/total*100:.1f}%)")
    htf_exit_count = df_t["Exit Reason"].str.contains("HTF ST Exit").sum()
    print(f"  Hard SL Exits       : {df_t['Exit Reason'].str.startswith('Hard SL').sum()}")
    print(f"  HTF ST Exits        : {htf_exit_count}")
    print(f"  Base ST SL Exits    : {(df_t['Exit Reason'] == 'ST Stop Loss').sum()}")
    print(f"  ST Flip Exits       : {df_t['Exit Reason'].str.startswith('ST Flip').sum()}")
    print(f"  EOD Force Exits     : {(df_t['Exit Reason'] == 'EOD Exit').sum()}")
    print(DASH)
    print(f"  {'P&L STATS':─<55}")
    print(f"  Total P&L           : {df_t['P&L %'].sum():.4f}%")
    print(f"  Avg P&L / Trade     : {df_t['P&L %'].mean():.4f}%")
    print(f"  Best Single Trade   : {df_t['P&L %'].max():.4f}%")
    print(f"  Worst Single Trade  : {df_t['P&L %'].min():.4f}%")
    print(DASH)
    print(f"  {'POINTS':─<55}")
    print(f"  Points Captured     : {win_pts:.2f}  (winning trades)")
    print(f"  Points Lost         : {loss_pts:.2f}  (losing trades)")
    print(f"  Net Points          : {win_pts - loss_pts:.2f}")
    print(DASH)
    print(f"  {'DAY STATS':─<55}")
    print(f"  Total Days          : {len(df_day)}")
    print(f"  Profit Days         : {p_days}  ({p_days/len(df_day)*100:.1f}%)")
    print(f"  Loss Days           : {l_days}  ({l_days/len(df_day)*100:.1f}%)")
    print(f"  Flat Days           : {len(df_day) - p_days - l_days}")
    print(f"  Best Day            : {best_d['Date']}  →  {best_d['Total P&L %']:.4f}%  |  Net Pts: {best_d['Net Points']:.2f}")
    print(f"  Worst Day           : {wrst_d['Date']}  →  {wrst_d['Total P&L %']:.4f}%  |  Net Pts: {wrst_d['Net Points']:.2f}")
    if best_s is not None:
        print(DASH)
        print(f"  {'TIME SLOT INSIGHTS':─<55}")
        print(f"  Best Slot           : {best_s['Time Slot']}  →  {best_s['Total P&L %']:.4f}%  |  WR: {best_s['Win Rate %']}%")
        print(f"  Worst Slot          : {wrst_s['Time Slot']}  →  {wrst_s['Total P&L %']:.4f}%  |  WR: {wrst_s['Win Rate %']}%")
    print(SEP + "\n")
    return df_t, df_day, df_ts


# ───────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ───────────────────────────────────────────────────────────────────────────────
def export_excel(df, df_trades, df_day, df_ts, htf_enabled):
    fname = f"{SYMBOL}_supertrend_v3.xlsx"

    df_st = df[["Open","High","Low","Close","Volume","ST_val","ST_direction"]].copy()
    df_st.index = df_st.index.strftime("%Y-%m-%d %H:%M")
    df_st.index.name = "DateTime (IST)"
    df_st.columns = ["Open","High","Low","Close","Volume",
                     f"ST({ST_PERIOD},{ST_MULTIPLIER})","Direction"]
    df_st = df_st.round(2)

    if df_trades is not None and not df_trades.empty:
        total    = len(df_trades); wins = (df_trades["P&L %"] > 0).sum()
        win_pts  = df_trades.loc[df_trades["P&L %"] > 0,  "Points Captured"].sum()
        loss_pts = df_trades.loc[df_trades["P&L %"] <= 0, "Points Captured"].sum()
        pd_      = int((df_day["Total P&L %"] > 0).sum()) if df_day is not None else 0
        ld_      = int((df_day["Total P&L %"] < 0).sum()) if df_day is not None else 0

        summary_rows = [
            ["STRATEGY INFO",  ""],
            ["Symbol",         SYMBOL],
            ["Base CSV",       BASE_CSV],
            ["HTF CSV",        HTF_CSV or "—  (HTF filter OFF)"],
            ["Base ST",        f"Period={ST_PERIOD}, Mult={ST_MULTIPLIER}"],
            ["HTF ST",         f"Period={HTF_ST_PERIOD}, Mult={HTF_ST_MULTIPLIER}"],
            ["HTF Active",     "YES" if htf_enabled else "NO"],
            ["EOD Exit",       str(MARKET_CLOSE)],
            ["", ""],
            ["EXIT CONFIG",    ""],
            ["Exit ST Line",   f"HTF ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER})" if (EXIT_USE_HTF_ST and htf_enabled) else f"Base ST({ST_PERIOD},{ST_MULTIPLIER})"],
            ["Exit Trigger",   "Close-based (candle must CLOSE beyond ST)" if CLOSE_BASED_EXIT else "Intracandle (Low/High touches ST)"],
            ["Hard SL",        f"{'ON ('+str(HARD_SL_PCT)+'%) — always intracandle' if HARD_SL_ENABLED else 'OFF'}"],
            ["", ""],
            ["FILTERS",        ""],
            ["HTF Filter",     "ON" if htf_enabled else "OFF"],
            ["Trade Window",   "ON" if TRADE_WINDOW_ENABLED else "OFF"],
            ["Confirm Candles",CONFIRM_CANDLES],
            ["Min Gap",        f"{'ON ('+str(MIN_GAP_PCT)+'%)' if MIN_GAP_ENABLED else 'OFF'}"],
            ["Volume Filter",  f"{'ON ('+str(VOLUME_MULTIPLIER)+'x)' if VOLUME_FILTER_ENABLED else 'OFF'}"],
            ["Max Trades/Day", MAX_TRADES_PER_DAY if MAX_TRADES_PER_DAY_ENABLED else "OFF"],
            ["", ""],
            ["TRADE STATS",    ""],
            ["Total Trades",   total],
            ["Winners",        f"{wins} ({wins/total*100:.1f}%)"],
            ["Losers",         f"{total-wins} ({(total-wins)/total*100:.1f}%)"],
            ["Hard SL Exits",  int(df_trades["Exit Reason"].str.startswith("Hard SL").sum())],
            ["HTF ST Exits",   int(df_trades["Exit Reason"].str.contains("HTF ST Exit").sum())],
            ["Base ST Exits",  int((df_trades["Exit Reason"]=="ST Stop Loss").sum())],
            ["ST Flip Exits",  int(df_trades["Exit Reason"].str.startswith("ST Flip").sum())],
            ["EOD Exits",      int((df_trades["Exit Reason"]=="EOD Exit").sum())],
            ["", ""],
            ["P&L STATS",      ""],
            ["Total P&L %",    round(df_trades["P&L %"].sum(), 4)],
            ["Avg P&L %",      round(df_trades["P&L %"].mean(), 4)],
            ["Best Trade %",   round(df_trades["P&L %"].max(), 4)],
            ["Worst Trade %",  round(df_trades["P&L %"].min(), 4)],
            ["", ""],
            ["POINTS",         ""],
            ["Points Captured",round(win_pts, 2)],
            ["Points Lost",    round(loss_pts, 2)],
            ["Net Points",     round(win_pts - loss_pts, 2)],
            ["", ""],
            ["DAY STATS",      ""],
            ["Total Days",     len(df_day) if df_day is not None else 0],
            ["Profit Days",    pd_],
            ["Loss Days",      ld_],
        ]
        if df_day is not None and not df_day.empty:
            bd = df_day.loc[df_day["Total P&L %"].idxmax()]
            wd = df_day.loc[df_day["Total P&L %"].idxmin()]
            summary_rows += [
                ["Best Day",  f"{bd['Date']}  →  {bd['Total P&L %']:.4f}%"],
                ["Worst Day", f"{wd['Date']}  →  {wd['Total P&L %']:.4f}%"],
            ]
        df_summary = pd.DataFrame(summary_rows, columns=["Metric","Value"])
    else:
        df_summary = pd.DataFrame([{"Metric":"No trades found","Value":""}])

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        df_st.to_excel(writer, sheet_name="Supertrend Data", index=True)
        if df_trades is not None and not df_trades.empty:
            df_trades.to_excel(writer, sheet_name="Trade Log", index=False)
        if df_day is not None and not df_day.empty:
            df_day.to_excel(writer, sheet_name="Daily P&L", index=False)
        if df_ts is not None and not df_ts.empty:
            df_ts.to_excel(writer, sheet_name="Time Slot P&L", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    _style_excel(fname, df_trades, df_day, df_ts)
    print(f"  Excel saved  → {fname}")
    return fname


def _style_excel(fname, df_trades, df_day, df_ts):
    GL = "C6EFCE"; GD = "1A5C38"; RL = "FFC7CE"; RD = "9C0006"
    BH = "1F3864"; YS = "FFD700"; GA = "F2F2F2"; WH = "FFFFFF"
    OR = "FCE4D6"; TL = "DDEBF7"
    ts  = Side(style="thin", color="CCCCCC")
    bdr = Border(left=ts, right=ts, top=ts, bottom=ts)

    def hdr(ws, row=1, bg=BH, fg=WH):
        for c in ws[row]:
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(bold=True, color=fg, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr
        ws.row_dimensions[row].height = 22

    def aw(ws, cap=30):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+3, cap)

    wb = load_workbook(fname)

    # ST Data
    ws = wb["Supertrend Data"]; hdr(ws); ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        v = str(row[-1].value or "")
        bg = GL if v=="Bullish" else RL if v=="Bearish" else WH
        fc = GD if v=="Bullish" else RD if v=="Bearish" else "000000"
        for cell in row:
            cell.fill=PatternFill("solid",fgColor=bg); cell.border=bdr
            cell.alignment=Alignment(horizontal="center")
        row[-1].font=Font(bold=True, color=fc)
    aw(ws)

    # Trade Log
    if "Trade Log" in wb.sheetnames and df_trades is not None:
        ws = wb["Trade Log"]; hdr(ws); ws.freeze_panes = "A2"
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            res = str(row[-1].value or ""); rsn = str(row[-2].value or "")
            bg  = GL if res=="WIN" else RL if res=="LOSS" else (GA if i%2==0 else WH)
            if res=="WIN":  row[-1].font=Font(bold=True,color=GD)
            if res=="LOSS": row[-1].font=Font(bold=True,color=RD)
            if rsn.startswith("Hard SL"):
                row[-2].fill=PatternFill("solid",fgColor="FF0000")
                row[-2].font=Font(bold=True,color="FFFFFF")
            elif "Stop Loss" in rsn:
                row[-2].fill=PatternFill("solid",fgColor=OR)
                row[-2].font=Font(bold=True,color="C55A11")
            elif "EOD" in rsn:
                row[-2].fill=PatternFill("solid",fgColor=TL)
                row[-2].font=Font(bold=True,color="2E75B6")
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000","FFFFFFFF"):
                    cell.fill=PatternFill("solid",fgColor=bg)
                cell.border=bdr; cell.alignment=Alignment(horizontal="center")
        aw(ws)

    # Daily P&L
    if "Daily P&L" in wb.sheetnames and df_day is not None:
        ws = wb["Daily P&L"]; hdr(ws); ws.freeze_panes = "A2"; rc = ws.max_column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            dr = str(row[rc-1].value or "")
            bg = GL if "Profit" in dr else RL if "Loss" in dr else GA
            if "Profit" in dr: row[rc-1].font=Font(bold=True,color=GD)
            elif "Loss"  in dr:row[rc-1].font=Font(bold=True,color=RD)
            for cell in row:
                cell.fill=PatternFill("solid",fgColor=bg); cell.border=bdr
                cell.alignment=Alignment(horizontal="center")
        aw(ws)

    # Time Slot P&L
    if "Time Slot P&L" in wb.sheetnames and df_ts is not None:
        ws = wb["Time Slot P&L"]; hdr(ws, bg="2E4057"); ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            try: pv = float(row[5].value) if row[5].value not in (None,"—") else 0
            except: pv = 0
            bg = GL if pv > 0 else RL if pv < 0 else GA
            vd = str(row[-1].value or "")
            if "🟢" in vd: row[-1].font=Font(bold=True,color=GD)
            elif "🔴" in vd:row[-1].font=Font(bold=True,color=RD)
            for cell in row:
                cell.fill=PatternFill("solid",fgColor=bg); cell.border=bdr
                cell.alignment=Alignment(horizontal="center")
        aw(ws)

    # Summary
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]; hdr(ws)
        SECS = {"STRATEGY INFO","EXIT CONFIG","FILTERS","TRADE STATS","P&L STATS","POINTS","DAY STATS"}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            lbl = str(row[0].value or "")
            if lbl in SECS:
                for c in row:
                    c.fill=PatternFill("solid",fgColor=YS)
                    c.font=Font(bold=True,size=10); c.border=bdr
                continue
            for c in row:
                c.border=bdr; c.alignment=Alignment(horizontal="left")
            if len(row)>1 and "Net Points" in lbl:
                try: v=float(row[1].value)
                except: v=0
                row[1].fill=PatternFill("solid",fgColor=GL if v>0 else RL)
                row[1].font=Font(bold=True,color=GD if v>0 else RD)
        ws.column_dimensions["A"].width=32; ws.column_dimensions["B"].width=40
    wb.save(fname)


# ───────────────────────────────────────────────────────────────────────────────
#  CHART
# ───────────────────────────────────────────────────────────────────────────────
def build_chart(df, trades, htf_enabled):
    if CHART_DAYS:
        udays  = sorted(df.index.normalize().unique())
        cutoff = udays[-CHART_DAYS] if len(udays) >= CHART_DAYS else udays[0]
        df_c   = df[df.index.normalize() >= cutoff].copy()
    else:
        df_c = df.copy()

    fig = go.Figure()
    fig.add_trace(go.Candlestick(
        x=df_c.index, open=df_c["Open"], high=df_c["High"],
        low=df_c["Low"], close=df_c["Close"], name="Price",
        increasing_line_color="#26a69a", decreasing_line_color="#ef5350"))
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["ST_val"].where(df_c["ST_bull"]),
        name=f"ST Bullish ({ST_PERIOD},{ST_MULTIPLIER})",
        mode="lines", line=dict(color="#22c55e", width=2.5), connectgaps=False))
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["ST_val"].where(~df_c["ST_bull"]),
        name=f"ST Bearish ({ST_PERIOD},{ST_MULTIPLIER})",
        mode="lines", line=dict(color="#ef4444", width=2.5), connectgaps=False))

    if htf_enabled and "HTF_bull" in df_c.columns:
        mid = (df_c["High"] + df_c["Low"]) / 2
        fig.add_trace(go.Scatter(x=df_c.index, y=mid.where(df_c["HTF_bull"]),
            name="HTF Bullish", mode="lines",
            line=dict(color="#86efac", width=1.2, dash="dot"), connectgaps=False))
        fig.add_trace(go.Scatter(x=df_c.index, y=mid.where(~df_c["HTF_bull"]),
            name="HTF Bearish", mode="lines",
            line=dict(color="#fca5a5", width=1.2, dash="dot"), connectgaps=False))

    if trades:
        df_t = pd.DataFrame(trades); start = df_c.index[0]
        def parse(col):
            return pd.to_datetime(df_t[col], format="%Y-%m-%d %H:%M").dt.tz_localize(IST)
        et=parse("Entry Time"); xt=parse("Exit Time")
        ep=df_t["Entry Price"]; xp=df_t["Exit Price"]
        dr=df_t["Direction"];   rs=df_t["Result"]; ex=df_t["Exit Reason"]
        mask = et >= start
        for xv,yv,nm,sym,col,ec in [
            (et[mask&dr.str.contains("Long")].tolist(),  ep[mask&dr.str.contains("Long")].tolist(),  "Long Entry",  "triangle-up",   "#22c55e","white"),
            (et[mask&dr.str.contains("Short")].tolist(), ep[mask&dr.str.contains("Short")].tolist(), "Short Entry", "triangle-down", "#ef4444","white"),
            (xt[mask&(rs=="WIN")].tolist(),  xp[mask&(rs=="WIN")].tolist(),  "Exit WIN",  "circle",   "#86efac","#16a34a"),
            (xt[mask&(rs=="LOSS")].tolist(), xp[mask&(rs=="LOSS")].tolist(), "Exit LOSS", "x",        "#fca5a5","#dc2626"),
            (xt[mask&ex.str.startswith("Hard SL")].tolist(), xp[mask&ex.str.startswith("Hard SL")].tolist(), f"Hard SL ({HARD_SL_PCT}%)", "hexagram","#ff0000","white"),
        ]:
            if xv: fig.add_trace(go.Scatter(x=xv, y=yv, mode="markers", name=nm,
                       marker=dict(symbol=sym,size=13,color=col,line=dict(color=ec,width=1.5))))

    htf_lbl = f"HTF ST({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER}) | " if htf_enabled else "HTF: OFF | "
    fig.update_layout(
        template="plotly_dark", height=820, xaxis_rangeslider_visible=False,
        legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="right", x=1, font=dict(size=9)),
        margin=dict(l=60,r=40,t=120,b=50), font=dict(family="monospace",size=11),
        title=dict(
            text=(f"<b>{SYMBOL}</b>  |  Supertrend({ST_PERIOD},{ST_MULTIPLIER})<br>"
                  f"<span style='font-size:10px;color:#94a3b8'>"
                  f"{htf_lbl}Window={'ON' if TRADE_WINDOW_ENABLED else 'OFF'} | "
                  f"ConfirmN={CONFIRM_CANDLES} | HardSL={HARD_SL_PCT}%</span>"),
            x=0.5, xanchor="center"))
    fig.update_yaxes(title_text="Price", showgrid=True, gridcolor="#1e293b")
    fig.update_xaxes(showgrid=True, gridcolor="#1e293b",
        rangebreaks=[dict(bounds=["sat","mon"]), dict(bounds=[15.5,9.25], pattern="hour")])

    chart_name = f"{SYMBOL}_chart_v3.html"
    fig.write_html(chart_name)
    print(f"  Chart saved  → {chart_name}  (open in browser)")


# ───────────────────────────────────────────────────────────────────────────────
#  MAIN
# ───────────────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═"*60)
    print("  SUPERTREND BACKTEST v3.0  —  CSV Mode")
    print("═"*60)
    try:
        df, htf_enabled = fetch_data()
    except (ValueError, FileNotFoundError) as e:
        print(e); return

    trades = run_backtest(df, htf_enabled)
    df_trades, df_day, df_ts = print_results(trades, htf_enabled)

    print("  Exporting Excel …")
    export_excel(df, df_trades, df_day, df_ts, htf_enabled)

    print("  Building chart …")
    build_chart(df, trades, htf_enabled)
    print("\n  ✅  Done.\n")


if __name__ == "__main__":
    main()