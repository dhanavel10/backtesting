"""
Supertrend Intraday Backtesting Strategy — ENHANCED with ADX + ATR Filters
───────────────────────────────────────────────────────────────────────────
FILTERS ADDED:
  1. ADX Filter   : Only trade when ADX > ADX_THRESHOLD (default 25)
                    ADX < threshold = sideways/ranging → SKIP trade
  2. ATR Filter   : Only trade when ATR > ATR_MA (default 20-period MA of ATR)
                    ATR < ATR_MA = range compressing → SKIP trade
  Both filters can be toggled ON/OFF independently via config flags.

LONG ENTRY  : ST is BELOW close price  (ST bull → price > ST line)
              + ADX > threshold (if enabled)
              + ATR > ATR_MA   (if enabled)
              ⚠ No entry on the 1st candle of each trading day (9:15 candle)
LONG EXIT   : Price touches / crosses the ST line from above
              (ST line acts as the dynamic stop loss)

SHORT ENTRY : ST is ABOVE close price  (ST bear → price < ST line)
              + ADX > threshold (if enabled)
              + ATR > ATR_MA   (if enabled)
              ⚠ No entry on the 1st candle of each trading day (9:15 candle)
SHORT EXIT  : Price touches / crosses the ST line from below
              (ST line acts as the dynamic stop loss)
ADDITIONAL EXITS:
  - Hard Stop Loss   : Fixed % loss limit from entry (default 0.3%)
                       Fires when candle LOW (long) or HIGH (short) breaches the hard SL price.
                       Takes priority over all other SL checks.
                       Can be toggled OFF to preserve original strategy behaviour.
  - Rolling Trail SL : Activates once profit reaches TRAIL_ACTIVATE_PTS (100) points.
                       Trail SL is placed TRAIL_OFFSET_PTS (70) below the peak high (long)
                       or above the trough low (short).
                       Every TRAIL_STEP_PTS (10) additional points the peak advances,
                       the trail moves up by the same step — locking in more profit.
                       Trail SL can ONLY move in the profitable direction, never back.
                       Re-entry is blocked after a Trail SL exit.
                       Set TRAIL_SL_ENABLED = False to disable entirely.
                       Priority order: Hard SL → Trail SL → ST SL (first trigger wins).
  - EOD force exit at 15:15 IST

OUTPUTS:
  - Console: trade log + per-day P&L + overall P&L summary (with filter stats)
  - Excel   : Supertrend data sheet + Trade Log + Daily P&L + Summary
  - HTML    : Interactive candlestick + Supertrend + ADX + ATR chart

Install: pip install yfinance pandas pandas-ta plotly openpyxl
"""

import yfinance as yf
import pandas as pd
import pandas_ta as ta
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pytz
import numpy as np
from datetime import datetime, timezone, time as dtime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════
DEFAULT_TICKER        = "^NSEI"
DEFAULT_INTERVAL      = "5m"
DEFAULT_PERIOD        = "60d"
DEFAULT_START         = None
DEFAULT_END           = None
DEFAULT_ST_PERIOD     = 10
DEFAULT_ST_MULTIPLIER = 3.0

# ── ADX Filter ──────────────────────────────────────────────────
ADX_FILTER_ENABLED = True          # Set False to disable ADX filter
ADX_PERIOD         = 14            # Standard ADX period
ADX_THRESHOLD      = 25            # Only trade when ADX > this value
                                   # 20 = more trades, 25 = balanced, 30 = only strong trends

# ── ATR Filter ──────────────────────────────────────────────────
ATR_FILTER_ENABLED = True          # Set False to disable ATR filter
ATR_MA_PERIOD      = 20            # MA period for ATR baseline
                                   # Only trade when ATR > its own MA (expanding range)

# ── Hard Stop Loss ──────────────────────────────────────────────
HARD_SL_ENABLED = True             # Set False to disable
HARD_SL_PCT     = 0.3              # Max loss allowed from entry price (%)
                                   # Fires BEFORE the ST line is touched.
                                   # Whichever triggers first (Hard SL vs Trail SL vs ST SL) wins.

# ── Rolling Trailing Stop Loss ──────────────────────────────────
TRAIL_SL_ENABLED    = True         # Set False to disable completely
TRAIL_ACTIVATE_PTS  = 150          # Profit in points at which trail SL activates
                                   # Trail only kicks in once trade profit ≥ this value
TRAIL_OFFSET_PTS    = 70           # Trail SL is placed this many points BELOW the peak (long)
                                   # or ABOVE the trough (short)
                                   # e.g. peak = entry + 100 → trail SL = peak − 70
TRAIL_STEP_PTS      = 10           # Every time peak advances by this many additional points,
                                   # trail SL moves up by the same amount
                                   # e.g. peak hits +110 → trail = peak − 70 = +40 from entry
                                   #      peak hits +120 → trail = peak − 70 = +50 from entry
                                   # Trail SL can ONLY move toward profit, never back.

# Intraday session (NSE)
MARKET_OPEN   = dtime(9, 15)
MARKET_CLOSE  = dtime(15, 15)
IST           = pytz.timezone("Asia/Kolkata")

# Chart: how many recent trading days to show (None = all)
CHART_DAYS    = 5
# ═══════════════════════════════════════════════════════════════

MAX_DAYS = {"1m": 7, "2m": 60, "5m": 60, "15m": 60,
            "30m": 60, "60m": 60, "90m": 60, "1h": 60}


def fmt(ts):
    return ts.strftime("%Y-%m-%d %H:%M") if pd.notna(ts) else "—"


def fmtd(ts):
    return ts.strftime("%Y-%m-%d") if pd.notna(ts) else "—"


# ───────────────────────────────────────────────────────────────
def validate(interval=DEFAULT_INTERVAL, start=DEFAULT_START):
    if interval not in MAX_DAYS:
        return
    max_days = MAX_DAYS[interval]
    if start:
        try:
            start_dt = datetime.strptime(start, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise ValueError(f"START='{start}' must be YYYY-MM-DD format.")
        days_ago = (datetime.now(timezone.utc) - start_dt).days
        if days_ago > max_days:
            raise ValueError(
                f"\n  ❌  '{interval}' only covers the last {max_days} days.\n"
                f"      START='{start}' is {days_ago} days ago.\n\n"
                f"  ✅  FIX:\n"
                f"      • Use a recent START within last {max_days} days, OR\n"
                f"      • Set START=None and PERIOD='60d', OR\n"
                f"      • Use INTERVAL='1d' for any historical range\n"
            )


# ───────────────────────────────────────────────────────────────
def compute_supertrend(df, st_period=DEFAULT_ST_PERIOD, st_multiplier=DEFAULT_ST_MULTIPLIER):
    """Wilder ATR-based Supertrend."""
    high  = df["High"].values
    low   = df["Low"].values
    close = df["Close"].values
    n, p, m = len(df), st_period, st_multiplier

    tr = np.maximum(high[1:] - low[1:],
         np.maximum(np.abs(high[1:] - close[:-1]),
                    np.abs(low[1:]  - close[:-1])))

    atr = np.full(n, np.nan)
    atr[p] = tr[:p].mean()
    for j in range(p + 1, n):
        atr[j] = (atr[j-1] * (p - 1) + tr[j-1]) / p

    hl2         = (high + low) / 2.0
    upper_basic = hl2 + m * atr
    lower_basic = hl2 - m * atr

    upper = np.full(n, np.nan)
    lower = np.full(n, np.nan)
    bull  = np.full(n, True, dtype=bool)

    for j in range(p, n):
        if np.isnan(upper[j-1]):
            upper[j] = upper_basic[j]
            lower[j] = lower_basic[j]
            bull[j]  = close[j] >= lower[j]
            continue
        upper[j] = (upper_basic[j]
                    if upper_basic[j] < upper[j-1] or close[j-1] > upper[j-1]
                    else upper[j-1])
        lower[j] = (lower_basic[j]
                    if lower_basic[j] > lower[j-1] or close[j-1] < lower[j-1]
                    else lower[j-1])
        bull[j] = (close[j] >= lower[j]) if bull[j-1] else (close[j] > upper[j])

    df = df.copy()
    df["ST_val"]       = np.where(bull, lower, upper)
    df["ST_bull"]      = bull
    df["ST_direction"] = np.where(bull, "Bullish", "Bearish")
    df["ST_atr"]       = atr
    return df


# ───────────────────────────────────────────────────────────────
def compute_adx(df, period=ADX_PERIOD):
    """Wilder-smoothed ADX. Adds: ADX, DI_plus, DI_minus."""
    high  = df["High"].values.astype(float)
    low   = df["Low"].values.astype(float)
    close = df["Close"].values.astype(float)
    n     = len(df)

    tr = np.full(n, np.nan)
    for i in range(1, n):
        tr[i] = max(high[i] - low[i],
                    abs(high[i] - close[i-1]),
                    abs(low[i]  - close[i-1]))

    dm_plus  = np.full(n, 0.0)
    dm_minus = np.full(n, 0.0)
    for i in range(1, n):
        up   = high[i] - high[i-1]
        down = low[i-1] - low[i]
        dm_plus[i]  = up   if (up > down and up > 0)   else 0.0
        dm_minus[i] = down if (down > up and down > 0) else 0.0

    atr_w = np.full(n, np.nan)
    dmp_w = np.full(n, np.nan)
    dmm_w = np.full(n, np.nan)

    if n > period:
        atr_w[period] = np.nansum(tr[1:period+1])
        dmp_w[period] = np.sum(dm_plus[1:period+1])
        dmm_w[period] = np.sum(dm_minus[1:period+1])
        for i in range(period+1, n):
            atr_w[i] = atr_w[i-1] - (atr_w[i-1] / period) + tr[i]
            dmp_w[i] = dmp_w[i-1] - (dmp_w[i-1] / period) + dm_plus[i]
            dmm_w[i] = dmm_w[i-1] - (dmm_w[i-1] / period) + dm_minus[i]

    di_plus  = np.where(atr_w > 0, 100 * dmp_w / atr_w, 0.0)
    di_minus = np.where(atr_w > 0, 100 * dmm_w / atr_w, 0.0)
    di_diff  = np.abs(di_plus - di_minus)
    di_sum   = di_plus + di_minus
    dx       = np.where(di_sum > 0, 100 * di_diff / di_sum, 0.0)

    adx = np.full(n, np.nan)
    start = period * 2
    if n > start:
        adx[start] = np.nanmean(dx[period:start+1])
        for i in range(start+1, n):
            if not np.isnan(adx[i-1]):
                adx[i] = (adx[i-1] * (period - 1) + dx[i]) / period

    df = df.copy()
    df["ADX"]      = adx
    df["DI_plus"]  = di_plus
    df["DI_minus"] = di_minus
    return df


# ───────────────────────────────────────────────────────────────
def compute_atr_filter(df, ma_period=ATR_MA_PERIOD):
    """ATR filter: trade only when ATR > its rolling MA (expanding range)."""
    df = df.copy()
    df["ATR_MA"]       = df["ST_atr"].rolling(window=ma_period, min_periods=ma_period).mean()
    df["ATR_expanding"] = df["ST_atr"] > df["ATR_MA"]
    return df


# ───────────────────────────────────────────────────────────────
def fetch_data(ticker=DEFAULT_TICKER, interval=DEFAULT_INTERVAL, period=DEFAULT_PERIOD,
               start=DEFAULT_START, end=DEFAULT_END,
               st_period=DEFAULT_ST_PERIOD, st_multiplier=DEFAULT_ST_MULTIPLIER):
    validate(interval, start)
    rng = f"{start} → {end}" if start and end else period
    print(f"\nFetching {ticker} | {interval} | {rng} …")

    kw = dict(interval=interval, auto_adjust=True, progress=True)
    df = (yf.download(ticker, start=start, end=end, **kw)
          if start and end else
          yf.download(ticker, period=period, **kw))

    if df.empty:
        raise ValueError(
            f"\n  ❌  No data for '{ticker}'.\n"
            f"      Nifty50='^NSEI'  |  Reliance='RELIANCE.NS'\n"
            f"      Check ticker spelling and date range.\n"
        )

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    df.index = pd.to_datetime(df.index)
    if df.index.tz is None:
        df.index = df.index.tz_localize("UTC")
    df.index = df.index.tz_convert(IST)

    df = compute_supertrend(df, st_period, st_multiplier)
    df = compute_adx(df, ADX_PERIOD)
    df = compute_atr_filter(df, ATR_MA_PERIOD)
    df.dropna(inplace=True)

    print(f"  {len(df)} candles  |  {fmt(df.index[0])} → {fmt(df.index[-1])}")
    print(f"\n  ── FILTER CONFIG ──────────────────────────────────────")
    print(f"  ADX Filter   : {'ENABLED' if ADX_FILTER_ENABLED else 'DISABLED'}  "
          f"| Period={ADX_PERIOD}  | Threshold={ADX_THRESHOLD}")
    print(f"  ATR Filter   : {'ENABLED' if ATR_FILTER_ENABLED else 'DISABLED'}  "
          f"| MA Period={ATR_MA_PERIOD}")
    print(f"  Hard SL      : {'ENABLED' if HARD_SL_ENABLED else 'DISABLED'}  "
          f"| {HARD_SL_PCT}% from entry")
    print(f"  Trail SL     : {'ENABLED' if TRAIL_SL_ENABLED else 'DISABLED'}  "
          f"| Activate +{TRAIL_ACTIVATE_PTS}pts | Offset {TRAIL_OFFSET_PTS}pts | Step {TRAIL_STEP_PTS}pts")
    print(f"  ───────────────────────────────────────────────────────\n")
    return df


# ───────────────────────────────────────────────────────────────
def run_backtest(df):
    """
    Enhanced backtest with ADX + ATR filters + Hard SL + Rolling Trail SL.

    Filter logic:
      - ADX / ATR filters gate NEW entries and re-entries only.
      - Exits always run regardless of filter state.
      - No entry on first candle of each day (9:15).

    Stop Loss priority on every candle (first hit wins):
      1. Hard SL  : fixed % from entry price
      2. Trail SL : activates at +TRAIL_ACTIVATE_PTS profit;
                    sits TRAIL_OFFSET_PTS below peak (long) / above trough (short);
                    advances by TRAIL_STEP_PTS for every new step the peak makes.
                    Can only move toward profit, never back.
      3. ST SL    : dynamic ST line value

    Re-entry rules:
      - After ST flip : re-entry allowed (if filters pass + not first candle)
      - After Hard SL : re-entry BLOCKED
      - After Trail SL: re-entry BLOCKED

    Returns: list of trade dicts, filtered_out count
    """
    trades       = []
    filtered_out = 0

    in_trade        = False
    direction       = None
    entry_price     = None
    entry_time      = None
    peak            = None          # highest high (long) / lowest low (short)
    hard_sl_price   = None          # fixed SL from entry; None if disabled
    trail_sl_price  = None          # current trail SL level; None = not yet activated
    trail_peak_step = None          # last peak anchor used to advance trail

    closes      = df["Close"].to_numpy(dtype=float)
    highs       = df["High"].to_numpy(dtype=float)
    lows        = df["Low"].to_numpy(dtype=float)
    st_vals     = df["ST_val"].to_numpy(dtype=float)
    st_bulls    = df["ST_bull"].to_numpy(dtype=bool)
    adx_vals    = df["ADX"].to_numpy(dtype=float)
    atr_vals    = df["ST_atr"].to_numpy(dtype=float)
    atr_ma_vals = df["ATR_MA"].to_numpy(dtype=float)
    times       = df.index

    current_day       = None
    first_candle_done = False

    # Stored at entry time for the record() call
    entry_adx    = np.nan
    entry_atr    = np.nan
    entry_atr_ma = np.nan

    def filters_pass(i):
        reasons = []
        if ADX_FILTER_ENABLED:
            adx = adx_vals[i]
            if np.isnan(adx) or adx <= ADX_THRESHOLD:
                reasons.append(f"ADX={adx:.1f}≤{ADX_THRESHOLD}")
        if ATR_FILTER_ENABLED:
            atr    = atr_vals[i]
            atr_ma = atr_ma_vals[i]
            if np.isnan(atr_ma) or atr <= atr_ma:
                reasons.append(f"ATR({atr:.1f})≤ATR_MA({atr_ma:.1f})")
        return (False, " | ".join(reasons)) if reasons else (True, "")

    def record(dir_, e_time, e_price, x_time, x_price, pk,
               sl_at_exit, reason, adx_e, atr_e, atr_ma_e, hard_sl, trail_sl):
        pnl = round((x_price - e_price) / e_price * 100, 4) if dir_ == "long" \
              else round((e_price - x_price) / e_price * 100, 4)
        pts = round(abs(x_price - e_price), 4)
        trades.append({
            "Date"            : e_time.strftime("%Y-%m-%d"),
            "Direction"       : "Long  ↑" if dir_ == "long" else "Short ↓",
            "Entry Time"      : fmt(e_time),
            "Entry Price"     : round(e_price, 2),
            "Hard SL Price"   : round(hard_sl, 2) if hard_sl is not None else "OFF",
            "Trail SL Price"  : round(trail_sl, 2) if trail_sl is not None else "—",
            "ST Stop Loss"    : round(sl_at_exit, 2),
            "Exit Time"       : fmt(x_time),
            "Exit Price"      : round(x_price, 2),
            "Peak"            : round(pk, 2),
            "Points Captured" : pts,
            "P&L %"           : pnl,
            "Exit Reason"     : reason,
            "Result"          : "WIN" if pnl > 0 else "LOSS",
            "ADX @ Entry"     : round(adx_e, 2) if not np.isnan(adx_e) else None,
            "ATR @ Entry"     : round(atr_e, 4) if not np.isnan(atr_e) else None,
            "ATR_MA @ Entry"  : round(atr_ma_e, 4) if not np.isnan(atr_ma_e) else None,
        })

    def reset_trail():
        """Return fresh (trail_sl_price, trail_peak_step) for a new entry."""
        return None, None

    for i in range(1, len(df)):
        c_close   = closes[i]
        c_high    = highs[i]
        c_low     = lows[i]
        c_time    = times[i]
        c_tod     = c_time.time()
        c_st      = st_vals[i]
        c_bull    = st_bulls[i]
        prev_bull = st_bulls[i-1]

        flipped_bull = (not prev_bull) and c_bull   # bear → bull flip
        flipped_bear = prev_bull and (not c_bull)   # bull → bear flip

        # ── EOD FORCE EXIT ──────────────────────────────────────
        if in_trade and c_tod >= MARKET_CLOSE:
            pk_final = max(peak, c_high) if direction == "long" else min(peak, c_low)
            record(direction, entry_time, entry_price,
                   c_time, c_close, pk_final, c_st, "EOD Exit",
                   entry_adx, entry_atr, entry_atr_ma, hard_sl_price, trail_sl_price)
            in_trade = False
            continue

        # ── Skip outside market hours ───────────────────────────
        if c_tod < MARKET_OPEN or c_tod >= MARKET_CLOSE:
            continue

        # ── First-candle-of-day tracking ────────────────────────
        c_date = c_time.date()
        if c_date != current_day:
            current_day       = c_date
            first_candle_done = False
        else:
            first_candle_done = True

        # ── MANAGE OPEN TRADE ───────────────────────────────────
        if in_trade:
            if direction == "long":
                if c_high > peak: peak = c_high

                # ── Update Trail SL ──────────────────────────────
                if TRAIL_SL_ENABLED:
                    profit_pts = peak - entry_price
                    if profit_pts >= TRAIL_ACTIVATE_PTS:
                        if trail_sl_price is None:
                            # First activation — snap to nearest completed step
                            steps_done      = int((profit_pts - TRAIL_ACTIVATE_PTS) / TRAIL_STEP_PTS)
                            trail_peak_step = entry_price + TRAIL_ACTIVATE_PTS + steps_done * TRAIL_STEP_PTS
                            trail_sl_price  = trail_peak_step - TRAIL_OFFSET_PTS
                        else:
                            # Advance trail for every new step the peak clears
                            while peak >= trail_peak_step + TRAIL_STEP_PTS:
                                trail_peak_step += TRAIL_STEP_PTS
                                trail_sl_price  += TRAIL_STEP_PTS

                # ── Check exits: Hard SL → Trail SL → ST SL ─────
                hard_sl_hit  = (HARD_SL_ENABLED and
                                hard_sl_price is not None and
                                c_low <= hard_sl_price)
                trail_sl_hit = (TRAIL_SL_ENABLED and
                                trail_sl_price is not None and
                                c_low <= trail_sl_price)
                st_sl_hit    = c_low <= c_st or flipped_bear

                if hard_sl_hit or trail_sl_hit or st_sl_hit:
                    if hard_sl_hit:
                        exit_px = round(hard_sl_price, 2)
                        reason  = f"Hard SL ({HARD_SL_PCT}%)"
                    elif trail_sl_hit:
                        exit_px = round(trail_sl_price, 2)
                        reason  = "Trail SL"
                    else:
                        exit_px = round(c_st, 2)
                        reason  = "ST Stop Loss" if c_low <= c_st else "ST Flip Bear"

                    record("long", entry_time, entry_price,
                           c_time, exit_px, peak, c_st, reason,
                           entry_adx, entry_atr, entry_atr_ma, hard_sl_price, trail_sl_price)
                    in_trade = False

                    # Re-entry into short — only on clean ST flip (not Hard/Trail SL)
                    if not hard_sl_hit and not trail_sl_hit and not c_bull and first_candle_done:
                        ok, _ = filters_pass(i)
                        if ok:
                            entry_price      = c_close; entry_time = c_time
                            peak             = c_low;   direction  = "short"; in_trade = True
                            hard_sl_price    = round(entry_price * (1 + HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
                            trail_sl_price, trail_peak_step = reset_trail()
                            entry_adx = adx_vals[i]; entry_atr = atr_vals[i]
                            entry_atr_ma = atr_ma_vals[i]
                        else:
                            filtered_out += 1

            elif direction == "short":
                if c_low < peak: peak = c_low

                # ── Update Trail SL ──────────────────────────────
                if TRAIL_SL_ENABLED:
                    profit_pts = entry_price - peak   # positive when short is in profit
                    if profit_pts >= TRAIL_ACTIVATE_PTS:
                        if trail_sl_price is None:
                            steps_done      = int((profit_pts - TRAIL_ACTIVATE_PTS) / TRAIL_STEP_PTS)
                            trail_peak_step = entry_price - TRAIL_ACTIVATE_PTS - steps_done * TRAIL_STEP_PTS
                            trail_sl_price  = trail_peak_step + TRAIL_OFFSET_PTS
                        else:
                            while peak <= trail_peak_step - TRAIL_STEP_PTS:
                                trail_peak_step -= TRAIL_STEP_PTS
                                trail_sl_price  -= TRAIL_STEP_PTS

                # ── Check exits: Hard SL → Trail SL → ST SL ─────
                hard_sl_hit  = (HARD_SL_ENABLED and
                                hard_sl_price is not None and
                                c_high >= hard_sl_price)
                trail_sl_hit = (TRAIL_SL_ENABLED and
                                trail_sl_price is not None and
                                c_high >= trail_sl_price)
                st_sl_hit    = c_high >= c_st or flipped_bull

                if hard_sl_hit or trail_sl_hit or st_sl_hit:
                    if hard_sl_hit:
                        exit_px = round(hard_sl_price, 2)
                        reason  = f"Hard SL ({HARD_SL_PCT}%)"
                    elif trail_sl_hit:
                        exit_px = round(trail_sl_price, 2)
                        reason  = "Trail SL"
                    else:
                        exit_px = round(c_st, 2)
                        reason  = "ST Stop Loss" if c_high >= c_st else "ST Flip Bull"

                    record("short", entry_time, entry_price,
                           c_time, exit_px, peak, c_st, reason,
                           entry_adx, entry_atr, entry_atr_ma, hard_sl_price, trail_sl_price)
                    in_trade = False

                    # Re-entry into long — only on clean ST flip (not Hard/Trail SL)
                    if not hard_sl_hit and not trail_sl_hit and c_bull and first_candle_done:
                        ok, _ = filters_pass(i)
                        if ok:
                            entry_price      = c_close; entry_time = c_time
                            peak             = c_high;  direction  = "long"; in_trade = True
                            hard_sl_price    = round(entry_price * (1 - HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
                            trail_sl_price, trail_peak_step = reset_trail()
                            entry_adx = adx_vals[i]; entry_atr = atr_vals[i]
                            entry_atr_ma = atr_ma_vals[i]
                        else:
                            filtered_out += 1
            continue

        # ── NOT IN TRADE — CHECK ENTRY ──────────────────────────
        if not first_candle_done:
            continue

        ok, _ = filters_pass(i)
        if not ok:
            filtered_out += 1
            continue

        if c_bull:      # LONG
            entry_price      = c_close; entry_time = c_time
            peak             = c_high;  direction  = "long"; in_trade = True
            hard_sl_price    = round(entry_price * (1 - HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
            trail_sl_price, trail_peak_step = reset_trail()
            entry_adx = adx_vals[i]; entry_atr = atr_vals[i]
            entry_atr_ma = atr_ma_vals[i]

        elif not c_bull:  # SHORT
            entry_price      = c_close; entry_time = c_time
            peak             = c_low;   direction  = "short"; in_trade = True
            hard_sl_price    = round(entry_price * (1 + HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
            trail_sl_price, trail_peak_step = reset_trail()
            entry_adx = adx_vals[i]; entry_atr = atr_vals[i]
            entry_atr_ma = atr_ma_vals[i]

    return trades, filtered_out


# ───────────────────────────────────────────────────────────────
def build_daily_summary(trades):
    if not trades:
        return pd.DataFrame()

    df_t = pd.DataFrame(trades)
    daily = []

    for date, grp in df_t.groupby("Date"):
        total   = len(grp)
        wins    = (grp["P&L %"] > 0).sum()
        losses  = total - wins
        pnl_sum = grp["P&L %"].sum()
        win_pts = grp.loc[grp["P&L %"] > 0, "Points Captured"].sum()
        los_pts = grp.loc[grp["P&L %"] <= 0, "Points Captured"].sum()
        longs   = grp["Direction"].str.contains("Long").sum()
        shorts  = grp["Direction"].str.contains("Short").sum()
        avg_adx = grp["ADX @ Entry"].mean() if "ADX @ Entry" in grp.columns else None
        daily.append({
            "Date"           : date,
            "Total Trades"   : total,
            "Longs"          : longs,
            "Shorts"         : shorts,
            "Winners"        : wins,
            "Losers"         : losses,
            "Win Rate %"     : round(wins / total * 100, 1),
            "Total P&L %"    : round(pnl_sum, 4),
            "Best Trade %"   : round(grp["P&L %"].max(), 4),
            "Worst Trade %"  : round(grp["P&L %"].min(), 4),
            "Points Captured": round(win_pts, 2),
            "Points Lost"    : round(los_pts, 2),
            "Net Points"     : round(win_pts - los_pts, 2),
            "Avg ADX @ Entry": round(avg_adx, 2) if avg_adx is not None and not np.isnan(avg_adx) else None,
            "Day Result"     : "✅ Profit" if pnl_sum > 0 else "❌ Loss" if pnl_sum < 0 else "⚖ Flat",
        })

    return pd.DataFrame(daily)


# ───────────────────────────────────────────────────────────────
def print_results(trades, filtered_out=0,
                  ticker=DEFAULT_TICKER, interval=DEFAULT_INTERVAL,
                  st_period=DEFAULT_ST_PERIOD, st_multiplier=DEFAULT_ST_MULTIPLIER):
    sep  = "═" * 130
    dash = "─" * 130

    print("\n" + sep)
    print(f"  Supertrend({st_period},{st_multiplier})  |  {ticker}  |  {interval}  |  EOD Exit: {MARKET_CLOSE}")
    print(f"  Long : ST below price  →  Stop Loss = ST line  →  Exit when Low ≤ ST line")
    print(f"  Short: ST above price  →  Stop Loss = ST line  →  Exit when High ≥ ST line")
    print(f"  ⚠  No new entries on the 1st candle of each trading day")
    print(f"  🔍 ADX Filter : {'ON ' if ADX_FILTER_ENABLED else 'OFF'}  | "
          f"ADX({ADX_PERIOD}) > {ADX_THRESHOLD}  |  "
          f"ATR Filter : {'ON ' if ATR_FILTER_ENABLED else 'OFF'}  | "
          f"ATR > ATR_MA({ATR_MA_PERIOD})")
    print(f"  🛑 Hard SL    : {'ON ' if HARD_SL_ENABLED else 'OFF'}  | "
          f"{HARD_SL_PCT}% from entry  (priority 1)")
    print(f"  📈 Trail SL   : {'ON ' if TRAIL_SL_ENABLED else 'OFF'}  | "
          f"Activate +{TRAIL_ACTIVATE_PTS}pts | Offset {TRAIL_OFFSET_PTS}pts | "
          f"Step {TRAIL_STEP_PTS}pts  (priority 2)")
    print(sep)

    if not trades:
        print("  ⚠  No trades found. Try a longer PERIOD, different INTERVAL, or relax filters.")
        if filtered_out > 0:
            print(f"  ℹ  {filtered_out} signals were filtered out by ADX/ATR filters.")
        print(sep)
        return None, None

    df_t   = pd.DataFrame(trades)
    df_day = build_daily_summary(trades)

    # ── TRADE LOG ───────────────────────────────────────────────
    print("\n  TRADE LOG")
    print(dash)
    cols = ["Date", "Direction", "Entry Time", "Entry Price",
            "Hard SL Price", "Trail SL Price", "ST Stop Loss",
            "Exit Time", "Exit Price", "Points Captured", "P&L %",
            "ADX @ Entry", "Exit Reason", "Result"]
    print(df_t[cols].to_string(index=False))

    # ── PER-DAY P&L ─────────────────────────────────────────────
    print("\n\n" + sep)
    print("  PER-DAY P&L BREAKDOWN")
    print(dash)
    day_cols = ["Date", "Total Trades", "Longs", "Shorts", "Winners", "Losers",
                "Win Rate %", "Total P&L %", "Best Trade %", "Worst Trade %",
                "Points Captured", "Points Lost", "Net Points",
                "Avg ADX @ Entry", "Day Result"]
    print(df_day[day_cols].to_string(index=False))

    # ── OVERALL SUMMARY ─────────────────────────────────────────
    total        = len(df_t)
    wins         = (df_t["P&L %"] > 0).sum()
    losses       = total - wins
    hard_exits   = df_t["Exit Reason"].str.startswith("Hard SL").sum()
    trail_exits  = (df_t["Exit Reason"] == "Trail SL").sum()
    sl_exits     = (df_t["Exit Reason"] == "ST Stop Loss").sum()
    st_exits     = df_t["Exit Reason"].str.startswith("ST Flip").sum()
    eod_exits    = (df_t["Exit Reason"] == "EOD Exit").sum()
    win_pts      = df_t.loc[df_t["P&L %"] > 0, "Points Captured"].sum()
    loss_pts     = df_t.loc[df_t["P&L %"] <= 0, "Points Captured"].sum()
    net_pts      = win_pts - loss_pts

    profit_days = (df_day["Total P&L %"] > 0).sum()
    loss_days   = (df_day["Total P&L %"] < 0).sum()
    flat_days   = len(df_day) - profit_days - loss_days
    best_day    = df_day.loc[df_day["Total P&L %"].idxmax()]
    worst_day   = df_day.loc[df_day["Total P&L %"].idxmin()]
    avg_adx     = df_t["ADX @ Entry"].mean() if "ADX @ Entry" in df_t.columns else None

    print("\n\n" + sep)
    print("  OVERALL SUMMARY")
    print(dash)
    print(f"  {'FILTER & SL CONFIG':─<60}")
    print(f"  ADX Filter            : {'ENABLED  (ADX > ' + str(ADX_THRESHOLD) + ')' if ADX_FILTER_ENABLED else 'DISABLED'}")
    print(f"  ATR Filter            : {'ENABLED  (ATR > ATR_MA(' + str(ATR_MA_PERIOD) + '))' if ATR_FILTER_ENABLED else 'DISABLED'}")
    print(f"  Hard Stop Loss        : {'ENABLED  (' + str(HARD_SL_PCT) + '% from entry)' if HARD_SL_ENABLED else 'DISABLED'}")
    print(f"  Rolling Trail SL      : ", end="")
    if TRAIL_SL_ENABLED:
        print(f"ENABLED  | Activate +{TRAIL_ACTIVATE_PTS}pts | Offset {TRAIL_OFFSET_PTS}pts | Step {TRAIL_STEP_PTS}pts")
    else:
        print("DISABLED")
    print(f"  Signals Filtered Out  : {filtered_out}")
    print(f"  Trades Taken          : {total}")
    if avg_adx is not None and not np.isnan(avg_adx):
        print(f"  Avg ADX at Entry      : {avg_adx:.2f}")
    print(dash)
    print(f"  {'TRADE STATS':─<60}")
    print(f"  Total Trades          : {total}")
    print(f"  Winners               : {wins}  ({wins/total*100:.1f}%)")
    print(f"  Losers                : {losses}  ({losses/total*100:.1f}%)")
    print(f"  Hard SL Exits         : {hard_exits}  (priority 1 — fixed % cap)")
    print(f"  Trail SL Exits        : {trail_exits}  (priority 2 — profit lock triggered)")
    print(f"  ST Stop Loss Exits    : {sl_exits}  (priority 3 — ST line touched)")
    print(f"  ST Flip Exits         : {st_exits}")
    print(f"  EOD Force Exits       : {eod_exits}")
    print(dash)
    print(f"  {'P&L STATS':─<60}")
    print(f"  Total P&L             : {df_t['P&L %'].sum():.4f}%")
    print(f"  Average P&L / Trade   : {df_t['P&L %'].mean():.4f}%")
    print(f"  Best Single Trade     : {df_t['P&L %'].max():.4f}%")
    print(f"  Worst Single Trade    : {df_t['P&L %'].min():.4f}%")
    print(dash)
    print(f"  {'POINTS':─<60}")
    print(f"  Total Points Captured : {win_pts:.2f}  (winning trades only)")
    print(f"  Points Lost           : {loss_pts:.2f}  (losing trades)")
    print(f"  Net Points            : {net_pts:.2f}")
    print(dash)
    print(f"  {'DAY STATS':─<60}")
    print(f"  Total Trading Days    : {len(df_day)}")
    print(f"  Profit Days           : {profit_days}  ({profit_days/len(df_day)*100:.1f}%)")
    print(f"  Loss Days             : {loss_days}  ({loss_days/len(df_day)*100:.1f}%)")
    print(f"  Flat Days             : {flat_days}")
    print(f"  Best Day              : {best_day['Date']}  →  {best_day['Total P&L %']:.4f}%  |  Net Pts: {best_day['Net Points']:.2f}")
    print(f"  Worst Day             : {worst_day['Date']}  →  {worst_day['Total P&L %']:.4f}%  |  Net Pts: {worst_day['Net Points']:.2f}")
    print(sep + "\n")

    return df_t, df_day


# ───────────────────────────────────────────────────────────────
def export_excel(df, trades, df_trades, df_day,
                 filtered_out=0,
                 ticker=DEFAULT_TICKER, interval=DEFAULT_INTERVAL,
                 st_period=DEFAULT_ST_PERIOD, st_multiplier=DEFAULT_ST_MULTIPLIER):
    ticker_clean = ticker.replace("^", "").replace(".", "_")
    fname = f"{ticker_clean}_{interval}_supertrend_enhanced.xlsx"

    df_st = df[["Open", "High", "Low", "Close", "Volume",
                "ST_val", "ST_direction",
                "ADX", "DI_plus", "DI_minus",
                "ST_atr", "ATR_MA"]].copy()
    df_st.index = df_st.index.strftime("%Y-%m-%d %H:%M")
    df_st.index.name = "DateTime (IST)"
    df_st.columns = ["Open", "High", "Low", "Close", "Volume",
                     f"ST Value ({st_period},{st_multiplier})", "ST Direction",
                     f"ADX({ADX_PERIOD})", "DI+", "DI-",
                     f"ATR({st_period})", f"ATR_MA({ATR_MA_PERIOD})"]
    df_st = df_st.round(4)

    if df_trades is not None and not df_trades.empty:
        total       = len(df_trades)
        wins        = (df_trades["P&L %"] > 0).sum()
        losses      = total - wins
        win_pts     = df_trades.loc[df_trades["P&L %"] > 0, "Points Captured"].sum()
        loss_pts    = df_trades.loc[df_trades["P&L %"] <= 0, "Points Captured"].sum()
        profit_days = int((df_day["Total P&L %"] > 0).sum()) if df_day is not None else 0
        loss_days_n = int((df_day["Total P&L %"] < 0).sum()) if df_day is not None else 0
        avg_adx     = df_trades["ADX @ Entry"].mean() if "ADX @ Entry" in df_trades.columns else None

        summary_rows = [
            ["STRATEGY INFO",       ""],
            ["Ticker",              ticker],
            ["Interval",            interval],
            ["Supertrend",          f"Period={st_period}, Multiplier={st_multiplier}"],
            ["Stop Loss",           "Dynamic — ST line value"],
            ["EOD Exit",            str(MARKET_CLOSE)],
            ["1st Candle Entry",    "Blocked — no entry on 9:15 candle"],
            ["", ""],
            ["FILTER & SL CONFIG",  ""],
            ["ADX Filter",          f"{'ENABLED' if ADX_FILTER_ENABLED else 'DISABLED'} — ADX({ADX_PERIOD}) > {ADX_THRESHOLD}"],
            ["ATR Filter",          f"{'ENABLED' if ATR_FILTER_ENABLED else 'DISABLED'} — ATR > ATR_MA({ATR_MA_PERIOD})"],
            ["Hard Stop Loss",      f"{'ENABLED' if HARD_SL_ENABLED else 'DISABLED'} — {HARD_SL_PCT}% max loss from entry"],
            ["Rolling Trail SL",    f"{'ENABLED' if TRAIL_SL_ENABLED else 'DISABLED'} — Activate +{TRAIL_ACTIVATE_PTS}pts | Offset {TRAIL_OFFSET_PTS}pts | Step {TRAIL_STEP_PTS}pts"],
            ["Signals Filtered Out",filtered_out],
            ["Trades Taken",        total],
            ["Avg ADX @ Entry",     round(avg_adx, 2) if avg_adx and not np.isnan(avg_adx) else "N/A"],
            ["", ""],
            ["TRADE STATS",         ""],
            ["Total Trades",        total],
            ["Winners",             f"{wins} ({wins/total*100:.1f}%)"],
            ["Losers",              f"{losses} ({losses/total*100:.1f}%)"],
            ["Hard SL Exits",       int(df_trades["Exit Reason"].str.startswith("Hard SL").sum())],
            ["Trail SL Exits",      int((df_trades["Exit Reason"] == "Trail SL").sum())],
            ["ST Stop Loss Exits",  int((df_trades["Exit Reason"] == "ST Stop Loss").sum())],
            ["ST Flip Exits",       int(df_trades["Exit Reason"].str.startswith("ST Flip").sum())],
            ["EOD Exits",           int((df_trades["Exit Reason"] == "EOD Exit").sum())],
            ["", ""],
            ["P&L STATS",           ""],
            ["Total P&L %",         round(df_trades["P&L %"].sum(), 4)],
            ["Avg P&L % / Trade",   round(df_trades["P&L %"].mean(), 4)],
            ["Best Trade %",        round(df_trades["P&L %"].max(), 4)],
            ["Worst Trade %",       round(df_trades["P&L %"].min(), 4)],
            ["", ""],
            ["POINTS",              ""],
            ["Points Captured (wins)", round(win_pts, 2)],
            ["Points Lost",         round(loss_pts, 2)],
            ["Net Points",          round(win_pts - loss_pts, 2)],
            ["", ""],
            ["DAY STATS",           ""],
            ["Total Trading Days",  len(df_day) if df_day is not None else 0],
            ["Profit Days",         profit_days],
            ["Loss Days",           loss_days_n],
        ]
        if df_day is not None and not df_day.empty:
            best  = df_day.loc[df_day["Total P&L %"].idxmax()]
            worst = df_day.loc[df_day["Total P&L %"].idxmin()]
            summary_rows += [
                ["Best Day",  f"{best['Date']}  →  {best['Total P&L %']:.4f}%"],
                ["Worst Day", f"{worst['Date']}  →  {worst['Total P&L %']:.4f}%"],
            ]
        df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value"])
    else:
        df_summary = pd.DataFrame([{"Metric": "No trades found", "Value": ""}])

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        df_st.to_excel(writer, sheet_name="Indicator Data", index=True)
        if df_trades is not None and not df_trades.empty:
            df_trades.to_excel(writer, sheet_name="Trade Log", index=False)
        if df_day is not None and not df_day.empty:
            df_day.to_excel(writer, sheet_name="Daily P&L", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    _style_excel(fname, df_trades, df_day)
    print(f"  Excel saved  → {fname}")
    return fname


def _style_excel(fname, df_trades, df_day):
    GREEN_DARK   = "1A5C38";  GREEN_LIGHT  = "C6EFCE"
    RED_DARK     = "9C0006";  RED_LIGHT    = "FFC7CE"
    BLUE_HDR     = "1F3864";  YELLOW_SEC   = "FFD700"
    GRAY_ALT     = "F2F2F2";  WHITE        = "FFFFFF"
    ORANGE       = "FCE4D6";  TEAL_LIGHT   = "DDEBF7"
    PURPLE_LIGHT = "E2EFDA";  PURPLE_TRAIL = "7B2D8B"

    thin_s = Side(style="thin", color="CCCCCC")
    bdr    = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    def hdr(ws, bg=BLUE_HDR, fg=WHITE):
        for cell in ws[1]:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(bold=True, color=fg, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = bdr
        ws.row_dimensions[1].height = 22

    def autowidth(ws, cap=28):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, cap)

    wb = load_workbook(fname)

    # ── Sheet 1: Indicator Data ──────────────────────────────────
    ws = wb["Indicator Data"]
    hdr(ws)
    ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        dir_cell = row[6]
        val      = str(dir_cell.value or "")
        bg     = GREEN_LIGHT if val == "Bullish" else RED_LIGHT if val == "Bearish" else WHITE
        fg_col = GREEN_DARK  if val == "Bullish" else RED_DARK  if val == "Bearish" else "000000"
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="center")
        dir_cell.font = Font(bold=True, color=fg_col)
    autowidth(ws)

    # ── Sheet 2: Trade Log ───────────────────────────────────────
    if "Trade Log" in wb.sheetnames and df_trades is not None:
        ws = wb["Trade Log"]
        hdr(ws)
        ws.freeze_panes = "A2"
        # Find Exit Reason column index dynamically
        hdr_row = [str(c.value or "") for c in ws[1]]
        reason_col_idx = next((i for i, h in enumerate(hdr_row) if "Exit Reason" in h), -4)

        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            result_cell = row[-1]
            result      = str(result_cell.value or "")
            reason_cell = row[reason_col_idx]
            reason      = str(reason_cell.value or "")

            if result == "WIN":
                bg = GREEN_LIGHT;  result_cell.font = Font(bold=True, color=GREEN_DARK)
            elif result == "LOSS":
                bg = RED_LIGHT;    result_cell.font = Font(bold=True, color=RED_DARK)
            else:
                bg = GRAY_ALT if i % 2 == 0 else WHITE

            # Exit reason colour coding
            if reason.startswith("Hard SL"):
                reason_cell.fill = PatternFill("solid", fgColor="FF0000")
                reason_cell.font = Font(bold=True, color="FFFFFF")
            elif reason == "Trail SL":
                reason_cell.fill = PatternFill("solid", fgColor=PURPLE_TRAIL)
                reason_cell.font = Font(bold=True, color="FFFFFF")
            elif "Stop Loss" in reason:
                reason_cell.fill = PatternFill("solid", fgColor=ORANGE)
                reason_cell.font = Font(bold=True, color="C55A11")
            elif "EOD" in reason:
                reason_cell.fill = PatternFill("solid", fgColor=TEAL_LIGHT)
                reason_cell.font = Font(bold=True, color="2E75B6")

            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = PatternFill("solid", fgColor=bg)
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    # ── Sheet 3: Daily P&L ───────────────────────────────────────
    if "Daily P&L" in wb.sheetnames and df_day is not None:
        ws = wb["Daily P&L"]
        hdr(ws)
        ws.freeze_panes = "A2"
        result_col = ws.max_column
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            day_result_cell = row[result_col - 1]
            day_result      = str(day_result_cell.value or "")
            if "Profit" in day_result:
                bg = GREEN_LIGHT;  day_result_cell.font = Font(bold=True, color=GREEN_DARK)
            elif "Loss" in day_result:
                bg = RED_LIGHT;    day_result_cell.font = Font(bold=True, color=RED_DARK)
            else:
                bg = GRAY_ALT
            for cell in row:
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    # ── Sheet 4: Summary ─────────────────────────────────────────
    ws = wb["Summary"]
    hdr(ws)
    SECTION_LABELS = {"STRATEGY INFO", "FILTER & SL CONFIG", "TRADE STATS", "P&L STATS", "POINTS", "DAY STATS"}
    FILTER_LABELS  = {"ADX Filter", "ATR Filter", "Hard Stop Loss", "Rolling Trail SL"}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        label = str(row[0].value or "")
        if label in SECTION_LABELS:
            for cell in row:
                cell.fill   = PatternFill("solid", fgColor=YELLOW_SEC)
                cell.font   = Font(bold=True, color="000000", size=10)
                cell.border = bdr
            continue
        if label in FILTER_LABELS:
            for cell in row:
                cell.fill   = PatternFill("solid", fgColor=PURPLE_LIGHT)
                cell.font   = Font(bold=True, color="375623")
                cell.border = bdr
            continue
        val_cell = row[1] if len(row) > 1 else None
        for cell in row:
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="left")
        if val_cell and "Net Points" in label:
            v = val_cell.value
            color = GREEN_LIGHT if isinstance(v, (int, float)) and v > 0 else RED_LIGHT
            val_cell.fill = PatternFill("solid", fgColor=color)
            val_cell.font = Font(bold=True,
                color=GREEN_DARK if isinstance(v, (int, float)) and v > 0 else RED_DARK)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34
    wb.save(fname)


# ───────────────────────────────────────────────────────────────
def build_chart(df, trades, filtered_out=0,
                ticker=DEFAULT_TICKER, interval=DEFAULT_INTERVAL,
                st_period=DEFAULT_ST_PERIOD, st_multiplier=DEFAULT_ST_MULTIPLIER):
    if CHART_DAYS:
        unique_days = sorted(df.index.normalize().unique())
        cutoff = unique_days[-CHART_DAYS] if len(unique_days) >= CHART_DAYS else unique_days[0]
        df_c   = df[df.index.normalize() >= cutoff].copy()
    else:
        df_c = df.copy()

    st_green = df_c["ST_val"].where(df_c["ST_bull"])
    st_red   = df_c["ST_val"].where(~df_c["ST_bull"])

    fig = make_subplots(
        rows=3, cols=1,
        shared_xaxes=True,
        row_heights=[0.60, 0.22, 0.18],
        vertical_spacing=0.03,
        subplot_titles=("", f"ADX({ADX_PERIOD})  — threshold {ADX_THRESHOLD}",
                        f"ATR vs ATR_MA({ATR_MA_PERIOD})")
    )

    # ── Row 1: Candlestick + ST lines ───────────────────────────
    fig.add_trace(go.Candlestick(
        x=df_c.index, open=df_c["Open"], high=df_c["High"],
        low=df_c["Low"], close=df_c["Close"], name="Price",
        increasing_line_color="#26a69a", decreasing_line_color="#ef5350"
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=df_c.index, y=st_green, name="ST Bullish (Long SL)",
        mode="lines", line=dict(color="#22c55e", width=2.5), connectgaps=False
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=df_c.index, y=st_red, name="ST Bearish (Short SL)",
        mode="lines", line=dict(color="#ef4444", width=2.5), connectgaps=False
    ), row=1, col=1)

    # ── Entry / Exit markers ────────────────────────────────────
    if trades:
        df_t   = pd.DataFrame(trades)
        in_win = df_c.index[0]

        def parse(col):
            return pd.to_datetime(df_t[col], format="%Y-%m-%d %H:%M").dt.tz_localize(IST)

        et = parse("Entry Time");  xt = parse("Exit Time")
        ep = df_t["Entry Price"];  xp = df_t["Exit Price"]
        dr = df_t["Direction"];    rs = df_t["Result"];  ex = df_t["Exit Reason"]
        mask = et >= in_win

        def pts(cond): return (et[mask & cond].tolist(), ep[mask & cond].tolist())
        def xts(cond): return (xt[mask & cond].tolist(), xp[mask & cond].tolist())

        lx, ly   = pts(dr.str.contains("Long"))
        sx, sy   = pts(dr.str.contains("Short"))
        wx, wy   = xts(rs == "WIN")
        lox, loy = xts(rs == "LOSS")
        slx, sly = xts(ex == "ST Stop Loss")
        hsx, hsy = xts(ex.str.startswith("Hard SL"))
        tsx, tsy = xts(ex == "Trail SL")

        markers = [
            (lx, ly,   "triangle-up",   14, "#22c55e", "Long Entry"),
            (sx, sy,   "triangle-down", 14, "#ef4444", "Short Entry"),
            (wx, wy,   "circle",        11, "#86efac", "Exit WIN"),
            (lox, loy, "x",             12, "#fca5a5", "Exit LOSS"),
            (slx, sly, "diamond",       11, "#fb923c", "ST Stop Loss Hit"),
            (hsx, hsy, "hexagram",      14, "#ff0000", f"Hard SL Hit ({HARD_SL_PCT}%)"),
            (tsx, tsy, "star",          14, "#c084fc", "Trail SL Hit"),
        ]
        for x_, y_, sym, sz, col, name_ in markers:
            if x_:
                fig.add_trace(go.Scatter(
                    x=x_, y=y_, mode="markers", name=name_,
                    marker=dict(symbol=sym, size=sz, color=col,
                                line=dict(color="white", width=1.5))
                ), row=1, col=1)

    # ── Row 2: ADX ───────────────────────────────────────────────
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["ADX"],
        name=f"ADX({ADX_PERIOD})",
        mode="lines", line=dict(color="#a78bfa", width=1.8),
        fill="tozeroy", fillcolor="rgba(167,139,250,0.12)"
    ), row=2, col=1)
    fig.add_hline(y=ADX_THRESHOLD, line_dash="dash",
                  line_color="#f59e0b", line_width=1.5,
                  annotation_text=f"Threshold {ADX_THRESHOLD}",
                  annotation_position="right",
                  annotation_font_color="#f59e0b",
                  row=2, col=1)
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["DI_plus"], name="DI+",
        mode="lines", line=dict(color="#22c55e", width=1, dash="dot")
    ), row=2, col=1)
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["DI_minus"], name="DI-",
        mode="lines", line=dict(color="#ef4444", width=1, dash="dot")
    ), row=2, col=1)

    # ── Row 3: ATR vs ATR_MA ─────────────────────────────────────
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["ST_atr"],
        name=f"ATR({st_period})",
        mode="lines", line=dict(color="#38bdf8", width=1.6)
    ), row=3, col=1)
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["ATR_MA"],
        name=f"ATR_MA({ATR_MA_PERIOD})",
        mode="lines", line=dict(color="#fb923c", width=1.6, dash="dash")
    ), row=3, col=1)
    expanding = df_c["ST_atr"].where(df_c["ATR_expanding"])
    fig.add_trace(go.Scatter(
        x=df_c.index, y=expanding,
        name="ATR Expanding (trade zone)",
        mode="lines", line=dict(color="rgba(34,197,94,0.0)"),
        fill="tonexty", fillcolor="rgba(34,197,94,0.10)"
    ), row=3, col=1)

    filter_parts = []
    if ADX_FILTER_ENABLED:  filter_parts.append(f"ADX>{ADX_THRESHOLD}")
    if ATR_FILTER_ENABLED:  filter_parts.append(f"ATR>ATR_MA({ATR_MA_PERIOD})")
    if HARD_SL_ENABLED:     filter_parts.append(f"HardSL {HARD_SL_PCT}%")
    if TRAIL_SL_ENABLED:    filter_parts.append(f"TrailSL +{TRAIL_ACTIVATE_PTS}/{TRAIL_OFFSET_PTS}off/{TRAIL_STEP_PTS}step")
    filter_str = " + ".join(filter_parts) if filter_parts else "No filters/SL"

    fig.update_layout(
        template="plotly_dark", height=900,
        xaxis_rangeslider_visible=False,
        legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="right", x=1,
                    font=dict(size=9)),
        margin=dict(l=60, r=40, t=120, b=50),
        font=dict(family="monospace", size=11),
        title=dict(
            text=(f"<b>{ticker}</b> | {interval} | "
                  f"Supertrend({st_period},{st_multiplier})  |  {filter_str}<br>"
                  f"<span style='font-size:10px;color:#94a3b8'>"
                  f"🟢 Long: ST below price  |  🔴 Short: ST above price  |  "
                  f"⬡ Hard SL  |  ★ Trail SL  |  "
                  f"{filtered_out} signals filtered</span>"),
            x=0.5, xanchor="center"
        )
    )
    fig.update_yaxes(showgrid=True, gridcolor="#1e293b")
    fig.update_xaxes(showgrid=True, gridcolor="#1e293b",
        rangebreaks=[dict(bounds=["sat", "mon"]),
                     dict(bounds=[15.5, 9.25], pattern="hour")])
    fig.update_yaxes(title_text="Price", row=1, col=1)
    fig.update_yaxes(title_text="ADX",   row=2, col=1)
    fig.update_yaxes(title_text="ATR",   row=3, col=1)

    chart_name = f"{ticker.replace('^','').replace('.','_')}_{interval}_chart_enhanced.html"
    fig.write_html(chart_name)
    print(f"  Chart saved  → {chart_name}  (open in browser)")
    return chart_name


# ───────────────────────────────────────────────────────────────
def run_backtest_api(ticker=DEFAULT_TICKER, interval=DEFAULT_INTERVAL,
                     period=DEFAULT_PERIOD, start=DEFAULT_START, end=DEFAULT_END,
                     st_period=DEFAULT_ST_PERIOD, st_multiplier=DEFAULT_ST_MULTIPLIER):
    """API entry point. Returns JSON-serialisable dicts."""
    df               = fetch_data(ticker, interval, period, start, end, st_period, st_multiplier)
    trades, filtered = run_backtest(df)
    df_day           = build_daily_summary(trades)
    df_t             = pd.DataFrame(trades)
    total            = len(trades)

    if total > 0:
        wins         = int((df_t["P&L %"] > 0).sum())
        losses       = int(total - wins)
        hard_exits   = int(df_t["Exit Reason"].str.startswith("Hard SL").sum())
        trail_exits  = int((df_t["Exit Reason"] == "Trail SL").sum())
        sl_exits     = int((df_t["Exit Reason"] == "ST Stop Loss").sum())
        st_exits     = int(df_t["Exit Reason"].str.startswith("ST Flip").sum())
        eod_exits    = int((df_t["Exit Reason"] == "EOD Exit").sum())
        win_pts      = float(df_t.loc[df_t["P&L %"] > 0, "Points Captured"].sum())
        loss_pts     = float(df_t.loc[df_t["P&L %"] <= 0, "Points Captured"].sum())
        net_pts      = float(win_pts - loss_pts)
        total_pnl    = float(df_t["P&L %"].sum())
        avg_pnl      = float(df_t["P&L %"].mean())
        best_pnl     = float(df_t["P&L %"].max())
        worst_pnl    = float(df_t["P&L %"].min())
        avg_adx      = float(df_t["ADX @ Entry"].mean()) if "ADX @ Entry" in df_t.columns else None
        profit_days  = int((df_day["Total P&L %"] > 0).sum())
        loss_days    = int((df_day["Total P&L %"] < 0).sum())
        flat_days    = int(len(df_day) - profit_days - loss_days)
        b = df_day.loc[df_day["Total P&L %"].idxmax()]
        w = df_day.loc[df_day["Total P&L %"].idxmin()]
        best_day  = {"Date": b["Date"], "Total P&L %": float(b["Total P&L %"]), "Net Points": float(b["Net Points"])}
        worst_day = {"Date": w["Date"], "Total P&L %": float(w["Total P&L %"]), "Net Points": float(w["Net Points"])}
    else:
        wins = losses = hard_exits = trail_exits = sl_exits = st_exits = eod_exits = 0
        win_pts = loss_pts = net_pts = total_pnl = avg_pnl = best_pnl = worst_pnl = 0.0
        profit_days = loss_days = flat_days = 0
        best_day = worst_day = avg_adx = None

    overall_summary = {
        "Filter Config": {
            "ADX Filter"         : ADX_FILTER_ENABLED,
            "ADX Period"         : ADX_PERIOD,
            "ADX Threshold"      : ADX_THRESHOLD,
            "ATR Filter"         : ATR_FILTER_ENABLED,
            "ATR_MA Period"      : ATR_MA_PERIOD,
            "Hard SL Enabled"    : HARD_SL_ENABLED,
            "Hard SL Pct"        : HARD_SL_PCT,
            "Trail SL Enabled"   : TRAIL_SL_ENABLED,
            "Trail Activate Pts" : TRAIL_ACTIVATE_PTS,
            "Trail Offset Pts"   : TRAIL_OFFSET_PTS,
            "Trail Step Pts"     : TRAIL_STEP_PTS,
            "Signals Filtered"   : filtered,
        },
        "Total Trades"           : total,
        "Winners"                : wins,
        "Losers"                 : losses,
        "Win Rate %"             : round((wins / total * 100) if total > 0 else 0, 2),
        "Hard SL Exits"          : hard_exits,
        "Trail SL Exits"         : trail_exits,
        "ST Stop Loss Exits"     : sl_exits,
        "ST Flip Exits"          : st_exits,
        "EOD Force Exits"        : eod_exits,
        "Total P&L %"            : round(total_pnl, 4),
        "Average P&L %"          : round(avg_pnl, 4),
        "Best Trade %"           : round(best_pnl, 4),
        "Worst Trade %"          : round(worst_pnl, 4),
        "Total Points Captured"  : round(win_pts, 2),
        "Points Lost"            : round(loss_pts, 2),
        "Net Points"             : round(net_pts, 2),
        "Avg ADX at Entry"       : round(avg_adx, 2) if avg_adx and not np.isnan(avg_adx) else None,
        "Total Trading Days"     : len(df_day),
        "Profit Days"            : profit_days,
        "Loss Days"              : loss_days,
        "Flat Days"              : flat_days,
        "Best Day"               : best_day,
        "Worst Day"              : worst_day,
    }

    df_day = df_day.replace({np.nan: None})
    return {
        "trades"         : trades,
        "daily_summary"  : df_day.to_dict(orient="records"),
        "overall_summary": overall_summary,
    }


# ───────────────────────────────────────────────────────────────
def main():
    ticker   = DEFAULT_TICKER
    interval = DEFAULT_INTERVAL
    period   = DEFAULT_PERIOD
    start    = DEFAULT_START
    end      = DEFAULT_END
    st_p     = DEFAULT_ST_PERIOD
    st_m     = DEFAULT_ST_MULTIPLIER

    try:
        df = fetch_data(ticker, interval, period, start, end, st_p, st_m)
    except ValueError as e:
        print(e); return

    trades, filtered_out = run_backtest(df)
    df_trades, df_day    = print_results(trades, filtered_out, ticker, interval, st_p, st_m)

    print("  Exporting Excel …")
    export_excel(df, trades, df_trades, df_day, filtered_out, ticker, interval, st_p, st_m)

    print("  Building chart …")
    build_chart(df, trades, filtered_out, ticker, interval, st_p, st_m)

    print("\n  ✅  Done.\n")


if __name__ == "__main__":
    main()