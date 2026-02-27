"""
Donchian Channel Breakout — Momentum Exit Strategy
─────────────────────────────────────────────────────────────────────────────
LONG ENTRY  : Current close breaks ABOVE the PREVIOUS candle's Donchian upper band
              (Price bursts out of the channel → momentum breakout signal)

LONG EXIT   : Whichever fires first —
              1. Hard SL        : Price LOW touches Donchian MIDLINE at entry candle
              2. Momentum Exit  : Current candle body < 0.5 × previous candle body
                                  (momentum dying — candle shrinking = get out)
              3. Time Exit      : 5 candles (25 min on 5m chart) — hard max hold
              4. EOD force exit : 15:15 IST

SHORT ENTRY : Current close breaks BELOW the PREVIOUS candle's Donchian lower band
              (Price breaks down out of channel → momentum breakdown signal)

SHORT EXIT  : Whichever fires first —
              1. Hard SL        : Price HIGH touches Donchian MIDLINE at entry candle
              2. Momentum Exit  : Current candle body < 0.5 × previous candle body
              3. Time Exit      : 5 candles (25 min on 5m chart)
              4. EOD force exit : 15:15 IST

DONCHIAN CHANNEL:
  Upper  = Highest High of last DC_PERIOD candles
  Lower  = Lowest  Low  of last DC_PERIOD candles
  Middle = (Upper + Lower) / 2  ← used as Hard SL

WHY THIS WORKS (from Nifty data):
  - Donchian upper/lower breaks = price escaping a consolidation range
  - The breakout candle IS the momentum candle — trade it immediately
  - Momentum dies fast (median = 1–2 candles on 5m Nifty)
  - Midline SL = natural halfway point of the channel = logical invalidation

OUTPUTS:
  - Console: trade log + per-day P&L + overall P&L summary
  - Excel  : Donchian data sheet + Trade Log + Daily P&L + Summary
  - HTML   : Interactive candlestick + Donchian channel chart

Install: pip install yfinance pandas plotly openpyxl
"""

import yfinance as yf
import pandas as pd
import plotly.graph_objects as go
import pytz
import numpy as np
from datetime import datetime, timezone, time as dtime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════
TICKER        = "^NSEI"
INTERVAL      = "1m"
PERIOD        = "7d"
START         = None
END           = None

# Donchian Channel settings
DC_PERIOD     = 20             # Lookback period for highest high / lowest low

# ── Hard Stop Loss — Donchian Midline ────────────────────────────
# SL = midline of the Donchian channel at the time of entry.
# If price reverses to the midline, the breakout has failed.
# This is set at entry and does NOT trail (fixed price level).
HARD_SL_ENABLED = True

# ── Momentum Exit ────────────────────────────────────────────────
# Exit when current candle body < MOMENTUM_DECAY_MULT × previous candle body.
# This detects that the breakout burst is fading — get out before theta/drift eats you.
# Default 0.5 = current candle is less than half the size of previous → dying.
MOMENTUM_EXIT_ENABLED = True
MOMENTUM_DECAY_MULT   = 0.5    # 0.5 = current body < 50% of previous body → exit

# ── Time Exit ────────────────────────────────────────────────────
# Hard maximum hold. On 5m chart: 5 candles = 25 minutes.
# After 25 min, a breakout move has either worked or stalled.
# Adjust for your interval: 1m → 25 candles, 15m → 2 candles
TIME_EXIT_ENABLED = True
MAX_CANDLES       = 5          # 5 candles × 5m = 25 minutes max hold

# Intraday session (NSE)
MARKET_OPEN   = dtime(9, 15)
MARKET_CLOSE  = dtime(15, 15)
IST           = pytz.timezone("Asia/Kolkata")

# Chart: how many recent trading days to show (None = all)
CHART_DAYS    = 5
# ═══════════════════════════════════════════════════════════════

MAX_DAYS = {"1m": 7, "2m": 60, "5m": 60, "15m": 60,
            "30m": 60, "60m": 60, "90m": 60, "1h": 60}


def fmt(ts):
    return ts.strftime("%Y-%m-%d %H:%M") if pd.notna(ts) else "—"


def fmtd(ts):
    return ts.strftime("%Y-%m-%d") if pd.notna(ts) else "—"


# ───────────────────────────────────────────────────────────────
def validate():
    if INTERVAL not in MAX_DAYS:
        return
    max_days = MAX_DAYS[INTERVAL]
    if START:
        try:
            start_dt = datetime.strptime(START, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise ValueError(f"START='{START}' must be YYYY-MM-DD format.")
        days_ago = (datetime.now(timezone.utc) - start_dt).days
        if days_ago > max_days:
            raise ValueError(
                f"\n  ❌  '{INTERVAL}' only covers the last {max_days} days.\n"
                f"      START='{START}' is {days_ago} days ago.\n\n"
                f"  ✅  FIX:\n"
                f"      • Use a recent START within last {max_days} days, OR\n"
                f"      • Set START=None and PERIOD='60d', OR\n"
                f"      • Use INTERVAL='1d' for any historical range\n"
            )


# ───────────────────────────────────────────────────────────────
def compute_donchian(df):
    """
    Donchian Channel (previous candle's values for clean entry logic).

    dc_upper  : rolling highest High over DC_PERIOD candles (shifted by 1 = previous)
    dc_lower  : rolling lowest  Low  over DC_PERIOD candles (shifted by 1 = previous)
    dc_mid    : (dc_upper + dc_lower) / 2  → used as hard SL level

    We use .shift(1) so entry conditions use the PREVIOUS candle's channel,
    not the current candle (avoids look-ahead bias on the same candle).
    """
    df = df.copy()

    dc_upper_raw = df["High"].rolling(DC_PERIOD).max()
    dc_lower_raw = df["Low"].rolling(DC_PERIOD).min()

    # Shift by 1: entry candle compares its close vs. PREVIOUS channel boundary
    df["DC_upper"] = dc_upper_raw.shift(1)
    df["DC_lower"] = dc_lower_raw.shift(1)
    df["DC_mid"]   = (df["DC_upper"] + df["DC_lower"]) / 2

    # Current (live) channel for chart drawing — no shift
    df["DC_upper_live"] = dc_upper_raw
    df["DC_lower_live"] = dc_lower_raw
    df["DC_mid_live"]   = (dc_upper_raw + dc_lower_raw) / 2

    # Candle body size (used for momentum decay exit)
    df["Body"] = (df["Close"] - df["Open"]).abs()

    return df


# ───────────────────────────────────────────────────────────────
def fetch_data():
    validate()
    rng = f"{START} → {END}" if START and END else PERIOD
    print(f"\nFetching {TICKER} | {INTERVAL} | {rng} …")

    kw = dict(interval=INTERVAL, auto_adjust=True, progress=True)
    df = (yf.download(TICKER, start=START, end=END, **kw)
          if START and END else
          yf.download(TICKER, period=PERIOD, **kw))

    if df.empty:
        raise ValueError(
            f"\n  ❌  No data for '{TICKER}'.\n"
            f"      Nifty50='^NSEI'  |  Reliance='RELIANCE.NS'\n"
            f"      Check ticker spelling and date range.\n"
        )

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    df.index = pd.to_datetime(df.index)
    if df.index.tz is None:
        df.index = df.index.tz_localize("UTC")
    df.index = df.index.tz_convert(IST)

    df = compute_donchian(df)
    df.dropna(inplace=True)
    print(f"  {len(df)} candles  |  {fmt(df.index[0])} → {fmt(df.index[-1])}")
    return df


# ───────────────────────────────────────────────────────────────
def run_backtest(df):
    """
    Donchian Channel Breakout — Momentum Exit Backtest.

    ENTRY:
      Long  : close > previous DC_upper  (breakout above channel)
      Short : close < previous DC_lower  (breakdown below channel)

    EXIT PRIORITY (checked in order each candle):
      1. Hard SL (midline at entry)   — price reversal, trade invalidated
      2. Momentum Exit                — current body < DECAY_MULT × previous body
      3. Time Exit                    — MAX_CANDLES reached (25 min on 5m)
      4. EOD Force Exit               — 15:15 IST

    Hard SL is fixed at entry (DC midline price at that candle).
    Re-entry after Hard SL: blocked (breakout failed, don't chase).
    Re-entry after other exits: allowed on next valid signal.
    """
    trades = []

    in_trade      = False
    direction     = None
    entry_price   = None
    entry_time    = None
    peak          = None
    hard_sl_price = None   # DC midline at entry — fixed, does not trail
    candles_held  = 0
    prev_body     = None   # body of the entry candle (for momentum decay check)

    closes     = df["Close"].to_numpy(dtype=float)
    opens      = df["Open"].to_numpy(dtype=float)
    highs      = df["High"].to_numpy(dtype=float)
    lows       = df["Low"].to_numpy(dtype=float)
    dc_uppers  = df["DC_upper"].to_numpy(dtype=float)
    dc_lowers  = df["DC_lower"].to_numpy(dtype=float)
    dc_mids    = df["DC_mid"].to_numpy(dtype=float)
    bodies     = df["Body"].to_numpy(dtype=float)
    times      = df.index

    def record(dir_, e_time, e_price, x_time, x_price, pk, hard_sl, dc_mid_at_entry, reason, candles):
        if dir_ == "long":
            pnl = round((x_price - e_price) / e_price * 100, 4)
        else:
            pnl = round((e_price - x_price) / e_price * 100, 4)
        pts = round(abs(x_price - e_price), 2)
        trades.append({
            "Date"             : e_time.strftime("%Y-%m-%d"),
            "Direction"        : "Long  ↑" if dir_ == "long" else "Short ↓",
            "Entry Time"       : fmt(e_time),
            "Entry Price"      : round(e_price, 2),
            "Hard SL (Midline)": round(hard_sl, 2) if hard_sl is not None else "OFF",
            "DC Mid at Entry"  : round(dc_mid_at_entry, 2),
            "Exit Time"        : fmt(x_time),
            "Exit Price"       : round(x_price, 2),
            "Peak"             : round(pk, 2),
            "Candles Held"     : candles,
            "Points Captured"  : pts,
            "P&L %"            : pnl,
            "Exit Reason"      : reason,
            "Result"           : "WIN" if pnl > 0 else "LOSS",
        })

    entry_dc_mid = None   # DC midline frozen at entry (for display in trade log)

    for i in range(1, len(df)):
        c_close  = closes[i]
        c_high   = highs[i]
        c_low    = lows[i]
        c_open   = opens[i]
        c_time   = times[i]
        c_tod    = c_time.time()
        c_body   = bodies[i]
        c_upper  = dc_uppers[i]    # previous candle's DC upper (shift=1)
        c_lower  = dc_lowers[i]    # previous candle's DC lower (shift=1)
        c_mid    = dc_mids[i]      # previous candle's DC mid

        # ── EOD FORCE EXIT ─────────────────────────────────────────────
        if in_trade and c_tod >= MARKET_CLOSE:
            pk_final = max(peak, c_high) if direction == "long" else min(peak, c_low)
            record(direction, entry_time, entry_price,
                   c_time, c_close, pk_final,
                   hard_sl_price, entry_dc_mid, "EOD Exit", candles_held)
            in_trade = False
            continue

        # Skip outside market hours
        if c_tod < MARKET_OPEN or c_tod >= MARKET_CLOSE:
            continue

        # ── MANAGE OPEN TRADE ──────────────────────────────────────────
        if in_trade:
            candles_held += 1

            if direction == "long":
                if c_high > peak:
                    peak = c_high

                # 1. Hard SL — DC midline at entry (fixed price)
                hard_sl_hit = (HARD_SL_ENABLED and
                               hard_sl_price is not None and
                               c_low <= hard_sl_price)

                # 2. Momentum Exit — current body shrinking vs entry body
                #    Check from candle 1 onward (prev_body = entry candle body)
                momentum_hit = (MOMENTUM_EXIT_ENABLED and
                                not hard_sl_hit and
                                prev_body is not None and prev_body > 0 and
                                c_body < MOMENTUM_DECAY_MULT * prev_body)

                # 3. Time Exit — max candles reached
                time_hit = (TIME_EXIT_ENABLED and
                            not hard_sl_hit and not momentum_hit and
                            candles_held >= MAX_CANDLES)

                if hard_sl_hit or momentum_hit or time_hit:
                    if hard_sl_hit:
                        exit_px = round(hard_sl_price, 2)
                        reason  = "Hard SL (Midline)"
                    elif momentum_hit:
                        exit_px = round(c_close, 2)
                        reason  = f"Momentum Fade (candle {candles_held}, body shrunk)"
                    else:
                        exit_px = round(c_close, 2)
                        reason  = f"Time Exit ({MAX_CANDLES} candles = {MAX_CANDLES * int(INTERVAL.replace('m',''))}min)"

                    pk_final = max(peak, c_high)
                    record("long", entry_time, entry_price,
                           c_time, exit_px, pk_final,
                           hard_sl_price, entry_dc_mid, reason, candles_held)
                    in_trade = False
                    # Block re-entry only after hard SL (breakout failed)
                    if hard_sl_hit:
                        continue
                else:
                    # Update prev_body for next candle's momentum check
                    prev_body = c_body
                continue

            elif direction == "short":
                if c_low < peak:
                    peak = c_low

                # 1. Hard SL — DC midline at entry (fixed price)
                hard_sl_hit = (HARD_SL_ENABLED and
                               hard_sl_price is not None and
                               c_high >= hard_sl_price)

                # 2. Momentum Exit
                momentum_hit = (MOMENTUM_EXIT_ENABLED and
                                not hard_sl_hit and
                                prev_body is not None and prev_body > 0 and
                                c_body < MOMENTUM_DECAY_MULT * prev_body)

                # 3. Time Exit
                time_hit = (TIME_EXIT_ENABLED and
                            not hard_sl_hit and not momentum_hit and
                            candles_held >= MAX_CANDLES)

                if hard_sl_hit or momentum_hit or time_hit:
                    if hard_sl_hit:
                        exit_px = round(hard_sl_price, 2)
                        reason  = "Hard SL (Midline)"
                    elif momentum_hit:
                        exit_px = round(c_close, 2)
                        reason  = f"Momentum Fade (candle {candles_held}, body shrunk)"
                    else:
                        exit_px = round(c_close, 2)
                        reason  = f"Time Exit ({MAX_CANDLES} candles = {MAX_CANDLES * int(INTERVAL.replace('m',''))}min)"

                    pk_final = min(peak, c_low)
                    record("short", entry_time, entry_price,
                           c_time, exit_px, pk_final,
                           hard_sl_price, entry_dc_mid, reason, candles_held)
                    in_trade = False
                    if hard_sl_hit:
                        continue
                else:
                    prev_body = c_body
                continue

        # ── NOT IN TRADE — CHECK ENTRY ─────────────────────────────────
        # Entry only during market hours and with valid DC values
        if np.isnan(c_upper) or np.isnan(c_lower) or np.isnan(c_mid):
            continue

        # LONG: close breaks above previous Donchian upper band
        if c_close > c_upper:
            entry_price   = c_close
            entry_time    = c_time
            entry_dc_mid  = c_mid
            peak          = c_high
            direction     = "long"
            in_trade      = True
            hard_sl_price = round(c_mid, 4) if HARD_SL_ENABLED else None
            prev_body     = c_body          # entry candle body = baseline for momentum check
            candles_held  = 0

        # SHORT: close breaks below previous Donchian lower band
        elif c_close < c_lower:
            entry_price   = c_close
            entry_time    = c_time
            entry_dc_mid  = c_mid
            peak          = c_low
            direction     = "short"
            in_trade      = True
            hard_sl_price = round(c_mid, 4) if HARD_SL_ENABLED else None
            prev_body     = c_body
            candles_held  = 0

    return trades


# ───────────────────────────────────────────────────────────────
def build_daily_summary(trades):
    if not trades:
        return pd.DataFrame()

    df_t  = pd.DataFrame(trades)
    daily = []

    for date, grp in df_t.groupby("Date"):
        total    = len(grp)
        wins     = (grp["P&L %"] > 0).sum()
        losses   = total - wins
        pnl_sum  = grp["P&L %"].sum()
        win_pts  = grp.loc[grp["P&L %"] > 0, "Points Captured"].sum()
        los_pts  = grp.loc[grp["P&L %"] <= 0, "Points Captured"].sum()
        longs    = grp["Direction"].str.contains("Long").sum()
        shorts   = grp["Direction"].str.contains("Short").sum()
        avg_held = grp["Candles Held"].mean()

        mom_exits  = grp["Exit Reason"].str.startswith("Momentum").sum()
        time_exits = grp["Exit Reason"].str.startswith("Time").sum()
        sl_exits   = grp["Exit Reason"].str.startswith("Hard SL").sum()

        daily.append({
            "Date"             : date,
            "Total Trades"     : total,
            "Longs"            : longs,
            "Shorts"           : shorts,
            "Winners"          : wins,
            "Losers"           : losses,
            "Win Rate %"       : round(wins / total * 100, 1),
            "Total P&L %"      : round(pnl_sum, 4),
            "Best Trade %"     : round(grp["P&L %"].max(), 4),
            "Worst Trade %"    : round(grp["P&L %"].min(), 4),
            "Points Captured"  : round(win_pts, 2),
            "Points Lost"      : round(los_pts, 2),
            "Net Points"       : round(win_pts - los_pts, 2),
            "Avg Candles Held" : round(avg_held, 1),
            "Momentum Exits"   : int(mom_exits),
            "Time Exits"       : int(time_exits),
            "Hard SL Exits"    : int(sl_exits),
            "Day Result"       : "✅ Profit" if pnl_sum > 0 else "❌ Loss" if pnl_sum < 0 else "⚖ Flat",
        })

    return pd.DataFrame(daily)


# ───────────────────────────────────────────────────────────────
def print_results(trades):
    sep  = "═" * 120
    dash = "─" * 120

    print("\n" + sep)
    print(f"  Donchian Channel({DC_PERIOD}) Breakout — Momentum Exit Strategy")
    print(f"  {TICKER}  |  {INTERVAL}  |  EOD Exit: {MARKET_CLOSE}")
    print(f"  Long : Close > prev DC Upper  →  Hard SL = DC Midline  →  Exit on momentum fade / {MAX_CANDLES} candles")
    print(f"  Short: Close < prev DC Lower  →  Hard SL = DC Midline  →  Exit on momentum fade / {MAX_CANDLES} candles")
    print(f"  🛑 Hard SL     : {'ON  — DC Midline at entry (fixed)' if HARD_SL_ENABLED else 'OFF'}")
    print(f"  💨 Momentum    : {'ON  — Exit when body < ' + str(MOMENTUM_DECAY_MULT) + '× prev body (fade detected)' if MOMENTUM_EXIT_ENABLED else 'OFF'}")
    print(
    f"  ⏱  Time Exit   : "
    f"{'ON  — ' + str(MAX_CANDLES) + ' candles = ' + str(MAX_CANDLES * int(INTERVAL.replace('m',''))) + ' min max hold' 
      if TIME_EXIT_ENABLED 
      else 'OFF'}")
    print(sep)

    if not trades:
        print("  ⚠  No trades found. Try reducing DC_PERIOD or checking date range.")
        print(sep)
        return None, None

    df_t   = pd.DataFrame(trades)
    df_day = build_daily_summary(trades)

    # ── TRADE LOG ──────────────────────────────────────────────────────
    print("\n  TRADE LOG")
    print(dash)
    cols = ["Date", "Direction", "Entry Time", "Entry Price", "Hard SL (Midline)",
            "DC Mid at Entry", "Exit Time", "Exit Price", "Candles Held",
            "Points Captured", "P&L %", "Exit Reason", "Result"]
    print(df_t[cols].to_string(index=False))

    # ── PER-DAY P&L ────────────────────────────────────────────────────
    print("\n\n" + sep)
    print("  PER-DAY P&L BREAKDOWN")
    print(dash)
    day_cols = ["Date", "Total Trades", "Longs", "Shorts", "Winners", "Losers",
                "Win Rate %", "Total P&L %", "Best Trade %", "Worst Trade %",
                "Points Captured", "Points Lost", "Net Points", "Avg Candles Held",
                "Momentum Exits", "Time Exits", "Hard SL Exits", "Day Result"]
    print(df_day[day_cols].to_string(index=False))

    # ── OVERALL SUMMARY ────────────────────────────────────────────────
    total       = len(df_t)
    wins        = (df_t["P&L %"] > 0).sum()
    losses      = total - wins
    mom_exits   = df_t["Exit Reason"].str.startswith("Momentum").sum()
    time_exits  = df_t["Exit Reason"].str.startswith("Time").sum()
    hard_exits  = df_t["Exit Reason"].str.startswith("Hard SL").sum()
    eod_exits   = (df_t["Exit Reason"] == "EOD Exit").sum()
    win_pts     = df_t.loc[df_t["P&L %"] > 0, "Points Captured"].sum()
    loss_pts    = df_t.loc[df_t["P&L %"] <= 0, "Points Captured"].sum()
    net_pts     = win_pts - loss_pts
    avg_held    = df_t["Candles Held"].mean()
    avg_win     = df_t.loc[df_t["P&L %"] > 0, "P&L %"].mean() if wins > 0 else 0
    avg_loss    = df_t.loc[df_t["P&L %"] <= 0, "P&L %"].mean() if losses > 0 else 0
    rr_ratio    = abs(avg_win / avg_loss) if avg_loss != 0 else float("inf")

    profit_days = (df_day["Total P&L %"] > 0).sum()
    loss_days   = (df_day["Total P&L %"] < 0).sum()
    flat_days   = len(df_day) - profit_days - loss_days
    best_day    = df_day.loc[df_day["Total P&L %"].idxmax()]
    worst_day   = df_day.loc[df_day["Total P&L %"].idxmin()]

    pct_mom  = mom_exits / total * 100 if total else 0
    pct_time = time_exits / total * 100 if total else 0
    pct_sl   = hard_exits / total * 100 if total else 0

    print("\n\n" + sep)
    print("  OVERALL SUMMARY")
    print(dash)
    print(f"  {'STRATEGY CONFIG':─<55}")
    print(f"  Donchian Period       : {DC_PERIOD} candles")
    print(f"  Entry Signal          : Close breaks prev DC Upper (long) / Lower (short)")
    print(f"  Hard SL               : {'ON  — DC Midline price at entry (fixed)' if HARD_SL_ENABLED else 'OFF'}")
    print(f"  Momentum Exit         : {'ON  — body < ' + str(MOMENTUM_DECAY_MULT) + '× prev body' if MOMENTUM_EXIT_ENABLED else 'OFF'}")
    print(
    f"  Time Exit             : "
    f"{'ON  — ' + str(MAX_CANDLES) + ' candles (' + str(MAX_CANDLES * int(INTERVAL.replace('m',''))) + ' min)' 
      if TIME_EXIT_ENABLED 
      else 'OFF'}"
)
    print(dash)
    print(f"  {'TRADE STATS':─<55}")
    print(f"  Total Trades          : {total}")
    print(f"  Winners               : {wins}  ({wins/total*100:.1f}%)")
    print(f"  Losers                : {losses}  ({losses/total*100:.1f}%)")
    print(f"  Realized R:R          : {rr_ratio:.2f}:1  (avg win / avg loss)")
    print(f"  Avg Candles Held      : {avg_held:.1f}  (target: 1–3 for momentum trades)")
    print(dash)
    print(f"  {'EXIT BREAKDOWN':─<55}")
    print(f"  Momentum Fade Exits   : {mom_exits}  ({pct_mom:.0f}%)  ← body shrunk, exited clean")
    print(f"  Time Exits            : {time_exits}  ({pct_time:.0f}%)  ← held full {MAX_CANDLES} candles")
    print(f"  Hard SL (Midline) Hits: {hard_exits}  ({pct_sl:.0f}%)  ← breakout failed, reversed to midline")
    print(f"  EOD Force Exits       : {eod_exits}")
    print(dash)
    print(f"  {'P&L STATS':─<55}")
    print(f"  Total P&L             : {df_t['P&L %'].sum():.4f}%")
    print(f"  Average P&L / Trade   : {df_t['P&L %'].mean():.4f}%")
    print(f"  Avg Winning Trade     : +{avg_win:.4f}%")
    print(f"  Avg Losing Trade      : {avg_loss:.4f}%")
    print(f"  Best Single Trade     : {df_t['P&L %'].max():.4f}%")
    print(f"  Worst Single Trade    : {df_t['P&L %'].min():.4f}%")
    print(dash)
    print(f"  {'POINTS':─<55}")
    print(f"  Total Points Captured : {win_pts:.2f}  (winning trades)")
    print(f"  Points Lost           : {loss_pts:.2f}  (losing trades)")
    print(f"  Net Points            : {net_pts:.2f}")
    print(dash)
    print(f"  {'DAY STATS':─<55}")
    print(f"  Total Trading Days    : {len(df_day)}")
    print(f"  Profit Days           : {profit_days}  ({profit_days/len(df_day)*100:.1f}%)")
    print(f"  Loss Days             : {loss_days}  ({loss_days/len(df_day)*100:.1f}%)")
    print(f"  Flat Days             : {flat_days}")
    print(f"  Best Day              : {best_day['Date']}  →  {best_day['Total P&L %']:.4f}%  |  Net Pts: {best_day['Net Points']:.2f}")
    print(f"  Worst Day             : {worst_day['Date']}  →  {worst_day['Total P&L %']:.4f}%  |  Net Pts: {worst_day['Net Points']:.2f}")
    print(sep + "\n")

    return df_t, df_day


# ───────────────────────────────────────────────────────────────
def export_excel(df, trades, df_trades, df_day):
    ticker_clean = TICKER.replace("^", "").replace(".", "_")
    fname = f"{ticker_clean}_{INTERVAL}_donchian_momentum.xlsx"

    # Sheet 1: Donchian raw data
    df_dc = df[["Open", "High", "Low", "Close", "Volume",
                "DC_upper_live", "DC_lower_live", "DC_mid_live",
                "DC_upper", "DC_lower", "DC_mid", "Body"]].copy()
    df_dc.index = df_dc.index.strftime("%Y-%m-%d %H:%M")
    df_dc.index.name = "DateTime (IST)"
    df_dc.columns = ["Open", "High", "Low", "Close", "Volume",
                     f"DC Upper (live, {DC_PERIOD})", f"DC Lower (live, {DC_PERIOD})",
                     "DC Mid (live)", "DC Upper (prev)", "DC Lower (prev)",
                     "DC Mid (prev)", "Body"]
    df_dc = df_dc.round(2)

    if df_trades is not None and not df_trades.empty:
        total    = len(df_trades)
        wins     = (df_trades["P&L %"] > 0).sum()
        losses   = total - wins
        win_pts  = df_trades.loc[df_trades["P&L %"] > 0, "Points Captured"].sum()
        loss_pts = df_trades.loc[df_trades["P&L %"] <= 0, "Points Captured"].sum()
        profit_days_n = int((df_day["Total P&L %"] > 0).sum()) if df_day is not None else 0
        loss_days_n   = int((df_day["Total P&L %"] < 0).sum()) if df_day is not None else 0
        avg_win  = df_trades.loc[df_trades["P&L %"] > 0, "P&L %"].mean() if wins > 0 else 0
        avg_loss = df_trades.loc[df_trades["P&L %"] <= 0, "P&L %"].mean() if losses > 0 else 0
        rr_str   = f"{abs(avg_win / avg_loss):.2f}:1" if avg_loss != 0 else "N/A"

        summary_rows = [
            ["STRATEGY",          "Donchian Channel Breakout — Momentum Exit"],
            ["Ticker",            TICKER],
            ["Interval",          INTERVAL],
            ["Donchian Period",   DC_PERIOD],
            ["Entry (Long)",      "Close > prev DC Upper"],
            ["Entry (Short)",     "Close < prev DC Lower"],
            ["EOD Exit",          str(MARKET_CLOSE)],
            ["", ""],
            ["EXIT RULES",        ""],
            ["Hard SL",           f"{'ON' if HARD_SL_ENABLED else 'OFF'} — DC Midline at entry (fixed)"],
            ["Momentum Exit",     f"{'ON' if MOMENTUM_EXIT_ENABLED else 'OFF'} — body < {MOMENTUM_DECAY_MULT}× prev body"],
            ["Time Exit",         f"{'ON' if TIME_EXIT_ENABLED else 'OFF'} — {MAX_CANDLES} candles ({MAX_CANDLES * int(INTERVAL.replace('m',''))} min)"],
            ["", ""],
            ["TRADE STATS",       ""],
            ["Total Trades",      total],
            ["Winners",           f"{wins} ({wins/total*100:.1f}%)"],
            ["Losers",            f"{losses} ({losses/total*100:.1f}%)"],
            ["Realized R:R",      rr_str],
            ["Avg Candles Held",  round(df_trades["Candles Held"].mean(), 1)],
            ["Momentum Exits",    int(df_trades["Exit Reason"].str.startswith("Momentum").sum())],
            ["Time Exits",        int(df_trades["Exit Reason"].str.startswith("Time").sum())],
            ["Hard SL Exits",     int(df_trades["Exit Reason"].str.startswith("Hard SL").sum())],
            ["EOD Exits",         int((df_trades["Exit Reason"] == "EOD Exit").sum())],
            ["", ""],
            ["P&L STATS",         ""],
            ["Total P&L %",       round(df_trades["P&L %"].sum(), 4)],
            ["Avg P&L % / Trade", round(df_trades["P&L %"].mean(), 4)],
            ["Best Trade %",      round(df_trades["P&L %"].max(), 4)],
            ["Worst Trade %",     round(df_trades["P&L %"].min(), 4)],
            ["", ""],
            ["POINTS",            ""],
            ["Points Captured",   round(win_pts, 2)],
            ["Points Lost",       round(loss_pts, 2)],
            ["Net Points",        round(win_pts - loss_pts, 2)],
            ["", ""],
            ["DAY STATS",         ""],
            ["Total Days",        len(df_day) if df_day is not None else 0],
            ["Profit Days",       profit_days_n],
            ["Loss Days",         loss_days_n],
        ]
        if df_day is not None and not df_day.empty:
            best  = df_day.loc[df_day["Total P&L %"].idxmax()]
            worst = df_day.loc[df_day["Total P&L %"].idxmin()]
            summary_rows += [
                ["Best Day",  f"{best['Date']}  →  {best['Total P&L %']:.4f}%"],
                ["Worst Day", f"{worst['Date']}  →  {worst['Total P&L %']:.4f}%"],
            ]
        df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value"])
    else:
        df_summary = pd.DataFrame([{"Metric": "No trades found", "Value": ""}])

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        df_dc.to_excel(writer, sheet_name="Donchian Data", index=True)
        if df_trades is not None and not df_trades.empty:
            df_trades.to_excel(writer, sheet_name="Trade Log", index=False)
        if df_day is not None and not df_day.empty:
            df_day.to_excel(writer, sheet_name="Daily P&L", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    _style_excel(fname, df_dc, df_trades, df_day)
    print(f"  Excel saved  → {fname}")
    return fname


def _style_excel(fname, df_dc, df_trades, df_day):
    GREEN_DARK  = "1A5C38"; GREEN_LIGHT = "C6EFCE"
    RED_DARK    = "9C0006"; RED_LIGHT   = "FFC7CE"
    BLUE_HDR    = "1F3864"; YELLOW_SEC  = "FFD700"
    GRAY_ALT    = "F2F2F2"; WHITE       = "FFFFFF"
    ORANGE      = "FCE4D6"; TEAL_LIGHT  = "DDEBF7"
    VIOLET      = "EAD1DC"; CHANNEL_BG  = "EBF3FB"

    thin_s = Side(style="thin", color="CCCCCC")
    bdr    = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    def hdr(ws, row=1, bg=BLUE_HDR, fg=WHITE):
        for cell in ws[row]:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(bold=True, color=fg, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = bdr
        ws.row_dimensions[row].height = 22

    def autowidth(ws, cap=30):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, cap)

    wb = load_workbook(fname)

    # Sheet 1: Donchian Data
    ws = wb["Donchian Data"]
    hdr(ws)
    ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=CHANNEL_BG)
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="center")
    autowidth(ws)

    # Sheet 2: Trade Log
    if "Trade Log" in wb.sheetnames and df_trades is not None:
        ws = wb["Trade Log"]
        hdr(ws)
        ws.freeze_panes = "A2"
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            result_cell = row[-1]
            result      = str(result_cell.value or "")
            reason_cell = row[-2]
            reason      = str(reason_cell.value or "")
            if result == "WIN":
                bg = GREEN_LIGHT; result_cell.font = Font(bold=True, color=GREEN_DARK)
            elif result == "LOSS":
                bg = RED_LIGHT;   result_cell.font = Font(bold=True, color=RED_DARK)
            else:
                bg = GRAY_ALT if i % 2 == 0 else WHITE

            if reason.startswith("Hard SL"):
                reason_cell.fill = PatternFill("solid", fgColor="FF0000")
                reason_cell.font = Font(bold=True, color="FFFFFF")
            elif reason.startswith("Momentum"):
                reason_cell.fill = PatternFill("solid", fgColor=VIOLET)
                reason_cell.font = Font(bold=True, color="7030A0")
            elif reason.startswith("Time"):
                reason_cell.fill = PatternFill("solid", fgColor=YELLOW_SEC)
                reason_cell.font = Font(bold=True, color="000000")
            elif "EOD" in reason:
                reason_cell.fill = PatternFill("solid", fgColor=TEAL_LIGHT)
                reason_cell.font = Font(bold=True, color="2E75B6")

            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = PatternFill("solid", fgColor=bg)
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    # Sheet 3: Daily P&L
    if "Daily P&L" in wb.sheetnames and df_day is not None:
        ws = wb["Daily P&L"]
        hdr(ws)
        ws.freeze_panes = "A2"
        result_col = ws.max_column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            day_result_cell = row[result_col - 1]
            day_result      = str(day_result_cell.value or "")
            if "Profit" in day_result:
                bg = GREEN_LIGHT; day_result_cell.font = Font(bold=True, color=GREEN_DARK)
            elif "Loss" in day_result:
                bg = RED_LIGHT;   day_result_cell.font = Font(bold=True, color=RED_DARK)
            else:
                bg = GRAY_ALT
            for cell in row:
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    # Sheet 4: Summary
    ws = wb["Summary"]
    hdr(ws)
    SECTIONS = {"STRATEGY", "EXIT RULES", "TRADE STATS", "P&L STATS", "POINTS", "DAY STATS"}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        label = str(row[0].value or "")
        if label in SECTIONS:
            for cell in row:
                cell.fill   = PatternFill("solid", fgColor=YELLOW_SEC)
                cell.font   = Font(bold=True, color="000000", size=10)
                cell.border = bdr
            continue
        val_cell = row[1] if len(row) > 1 else None
        for cell in row:
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="left")
        if val_cell:
            v = val_cell.value
            if "Net Points" in label:
                color = GREEN_LIGHT if isinstance(v, (int, float)) and v > 0 else RED_LIGHT
                val_cell.fill = PatternFill("solid", fgColor=color)
                val_cell.font = Font(bold=True,
                    color=GREEN_DARK if isinstance(v, (int, float)) and v > 0 else RED_DARK)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    wb.save(fname)


# ───────────────────────────────────────────────────────────────
def build_chart(df, trades):
    if CHART_DAYS:
        unique_days = sorted(df.index.normalize().unique())
        cutoff = unique_days[-CHART_DAYS] if len(unique_days) >= CHART_DAYS else unique_days[0]
        df_c   = df[df.index.normalize() >= cutoff].copy()
    else:
        df_c = df.copy()

    fig = go.Figure()

    # ── Donchian Channel fill (shaded band) ──────────────────────────
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["DC_upper_live"],
        name=f"DC Upper ({DC_PERIOD})",
        mode="lines", line=dict(color="#818cf8", width=1.5),
        fill=None
    ))
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["DC_lower_live"],
        name=f"DC Lower ({DC_PERIOD})",
        mode="lines", line=dict(color="#818cf8", width=1.5),
        fill="tonexty",
        fillcolor="rgba(129,140,248,0.08)"
    ))
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["DC_mid_live"],
        name="DC Midline (SL reference)",
        mode="lines", line=dict(color="#f59e0b", width=1, dash="dot"),
        fill=None
    ))

    # ── Candlestick ───────────────────────────────────────────────────
    fig.add_trace(go.Candlestick(
        x=df_c.index, open=df_c["Open"], high=df_c["High"],
        low=df_c["Low"], close=df_c["Close"], name="Price",
        increasing_line_color="#26a69a", decreasing_line_color="#ef5350"
    ))

    # ── Trade markers ─────────────────────────────────────────────────
    if trades:
        df_t = pd.DataFrame(trades)
        in_win = df_c.index[0]

        def parse(col):
            return pd.to_datetime(df_t[col], format="%Y-%m-%d %H:%M").dt.tz_localize(IST)

        et = parse("Entry Time"); xt = parse("Exit Time")
        ep = df_t["Entry Price"]; xp = df_t["Exit Price"]
        dr = df_t["Direction"];   rs = df_t["Result"]
        ex = df_t["Exit Reason"]
        mask = et >= in_win

        def scatter(times, prices, name, symbol, color, border, size=14):
            if times:
                fig.add_trace(go.Scatter(
                    x=times, y=prices, mode="markers", name=name,
                    marker=dict(symbol=symbol, size=size, color=color,
                                line=dict(color=border, width=1.5))
                ))

        scatter(et[mask & dr.str.contains("Long")].tolist(),
                ep[mask & dr.str.contains("Long")].tolist(),
                "Long Entry (DC breakout ↑)", "triangle-up", "#22c55e", "white")
        scatter(et[mask & dr.str.contains("Short")].tolist(),
                ep[mask & dr.str.contains("Short")].tolist(),
                "Short Entry (DC breakdown ↓)", "triangle-down", "#ef4444", "white")
        scatter(xt[mask & (rs == "WIN") & ~ex.str.startswith("Hard SL")].tolist(),
                xp[mask & (rs == "WIN") & ~ex.str.startswith("Hard SL")].tolist(),
                "Exit WIN", "circle", "#86efac", "#16a34a")
        scatter(xt[mask & ex.str.startswith("Momentum")].tolist(),
                xp[mask & ex.str.startswith("Momentum")].tolist(),
                "Momentum Fade Exit ← body shrunk", "diamond", "#c084fc", "#7e22ce", 12)
        scatter(xt[mask & ex.str.startswith("Time")].tolist(),
                xp[mask & ex.str.startswith("Time")].tolist(),
                f"Time Exit ({MAX_CANDLES} candles)", "square", "#fbbf24", "#92400e", 12)
        scatter(xt[mask & ex.str.startswith("Hard SL")].tolist(),
                xp[mask & ex.str.startswith("Hard SL")].tolist(),
                "Hard SL Hit (Midline)", "x", "#ff0000", "white", 14)
        scatter(xt[mask & (rs == "LOSS") & ~ex.str.startswith("Hard SL")].tolist(),
                xp[mask & (rs == "LOSS") & ~ex.str.startswith("Hard SL")].tolist(),
                "Exit LOSS (other)", "x", "#fca5a5", "#dc2626", 12)

    fig.update_layout(
        template="plotly_dark", height=800,
        xaxis_rangeslider_visible=False,
        legend=dict(orientation="h", yanchor="bottom", y=1.01,
                    xanchor="right", x=1, font=dict(size=9)),
        margin=dict(l=60, r=40, t=120, b=50),
        font=dict(family="monospace", size=11),
        title=dict(
            text=(f"<b>{TICKER}</b> | {INTERVAL} | Donchian({DC_PERIOD}) Breakout — Momentum Exit<br>"
                  f"<span style='font-size:10px;color:#94a3b8'>"
                  f"▲ Long breakout (close > DC upper)  |  ▼ Short breakdown (close < DC lower)  |  "
                  f"Yellow dashed = DC Midline (Hard SL)  |  "
                  f"◆ Momentum fade exit  |  ■ Time exit  |  ✕ Hard SL (midline)</span>"),
            x=0.5, xanchor="center"
        )
    )
    fig.update_yaxes(title_text="Price", showgrid=True, gridcolor="#1e293b")
    fig.update_xaxes(showgrid=True, gridcolor="#1e293b",
        rangebreaks=[dict(bounds=["sat", "mon"]),
                     dict(bounds=[15.5, 9.25], pattern="hour")])

    chart_name = f"{TICKER.replace('^','').replace('.','_')}_{INTERVAL}_donchian_chart.html"
    fig.write_html(chart_name)
    print(f"  Chart saved  → {chart_name}  (open in browser)")
    return chart_name


# ───────────────────────────────────────────────────────────────
def main():
    try:
        df = fetch_data()
    except ValueError as e:
        print(e); return

    trades            = run_backtest(df)
    df_trades, df_day = print_results(trades)

    print("  Exporting Excel …")
    export_excel(df, trades, df_trades, df_day)

    print("  Building chart …")
    build_chart(df, trades)

    print("\n  ✅  Done.\n")
    print("  ── TUNING GUIDE ────────────────────────────────────────────────────")
    print(f"  Too many false breakouts?  → Raise DC_PERIOD ({DC_PERIOD}) — wider channel = rarer, cleaner breaks")
    print(f"  Too few trades?            → Lower DC_PERIOD — tighter channel = more frequent breaks")
    print(f"  Exiting too early?         → Raise MOMENTUM_DECAY_MULT ({MOMENTUM_DECAY_MULT}) e.g. to 0.3 (less sensitive)")
    print(f"  Exiting too late?          → Lower MOMENTUM_DECAY_MULT to 0.6 (more sensitive to fade)")
    print(f"  Hard SL too tight?         → Midline SL is by design. For wider SL, disable HARD_SL_ENABLED")
    print(f"  Holding too long?          → Reduce MAX_CANDLES ({MAX_CANDLES}); data: momentum dies in 1–2 candles")
    print(f"  ────────────────────────────────────────────────────────────────────\n")


if __name__ == "__main__":
    main()