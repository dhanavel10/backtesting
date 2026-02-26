"""
Supertrend Intraday Backtesting Strategy
─────────────────────────────────────────
LONG ENTRY  : ST is BELOW close price  (ST bull → price > ST line)
LONG EXIT   : Price touches / crosses the ST line from above
              (ST line acts as the dynamic stop loss)
              OR Hard SL fires first (whichever comes first)

SHORT ENTRY : ST is ABOVE close price  (ST bear → price < ST line)
SHORT EXIT  : Price touches / crosses the ST line from below
              (ST line acts as the dynamic stop loss)
              OR Hard SL fires first (whichever comes first)

ADDITIONAL EXITS:
  - Hard Stop Loss : Fixed % loss limit from entry (default 0.2%)
                     Fires when candle LOW (long) or HIGH (short) breaches the hard SL price.
                     Takes priority over ST Stop Loss check.
                     Set HARD_SL_ENABLED = False to revert to pure ST SL behaviour.
  - EOD force exit at 15:25 IST — ALL open positions closed, no exceptions
  - No new entries after 15:15 IST
  - First 5-min candle (9:15) skipped every day
  - ST flat + price converging to ST → entry blocked until price moves away

ADX FILTER:
  - ADX is calculated using period 14.
  - If ADX < 15 at the time of entry signal → skip the trade and wait.
  - Once ADX ≥ 15 and all other conditions are met → entry is allowed.
  - ADX filter applies to FRESH entries only. Exit logic is NOT affected.

OUTPUTS:
  - Console: trade log + per-day P&L + overall P&L summary
  - Excel   : Supertrend data sheet + Trade Log + Daily P&L + Summary
  - HTML    : Interactive candlestick + Supertrend chart

Install: pip install yfinance pandas pandas-ta plotly openpyxl
"""

import yfinance as yf
import pandas as pd
import pandas_ta as ta
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pytz
import numpy as np
from datetime import datetime, timezone, time as dtime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════
TICKER        = "^NSEI"       # ^NSEI | RELIANCE.NS | TCS.NS etc.
INTERVAL      = "5m"         # 1m / 5m / 15m (intraday, last 60d only)
PERIOD        = "60d"         # used when START=None, END=None
START         = None          # e.g. "2026-01-25"
END           = None          # e.g. "2026-02-23"

# Supertrend settings
ST_PERIOD     = 10
ST_MULTIPLIER = 3.0

# ── Hard Stop Loss ──────────────────────────────────────────────
HARD_SL_ENABLED = True          # Set False to disable (reverts to pure ST SL behaviour)
HARD_SL_PCT     = 0.3        # Max loss allowed from entry price (%)
                                 # Exit fires when loss reaches this level,
                                 # BEFORE the ST line is even touched.
                                 # Works independently — does NOT replace ST SL,
                                 # whichever triggers first wins.

# ── ADX Entry Filter ────────────────────────────────────────────
ADX_PERIOD        = 14           # ADX calculation period
ADX_MIN_ENTRY     = 15           # Do NOT enter any trade if ADX < this value.
                                  # Wait until ADX >= 15 before taking a signal.
                                  # Exit logic is completely unaffected by this filter.

# Intraday session (NSE)
MARKET_OPEN        = dtime(9, 15)
MARKET_CLOSE       = dtime(15, 25)  # ALL positions force-exited at 15:25, no exceptions
NO_HOLD_AFTER      = dtime(15, 25)  # alias — same as MARKET_CLOSE, for clarity
NO_NEW_TRADE_AFTER = dtime(15, 15)  # no new entries after 15:15
IST           = pytz.timezone("Asia/Kolkata")

# Chart: how many recent trading days to show (None = all)
CHART_DAYS    = 5
# ═══════════════════════════════════════════════════════════════

MAX_DAYS = {"1m": 7, "2m": 60, "5m": 60, "15m": 60,
            "30m": 60, "60m": 60, "90m": 60, "1h": 60}


def fmt(ts):
    return ts.strftime("%Y-%m-%d %H:%M") if pd.notna(ts) else "—"


def fmtd(ts):
    return ts.strftime("%Y-%m-%d") if pd.notna(ts) else "—"


# ───────────────────────────────────────────────────────────────
def validate():
    if INTERVAL not in MAX_DAYS:
        return
    max_days = MAX_DAYS[INTERVAL]
    if START:
        try:
            start_dt = datetime.strptime(START, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise ValueError(f"START='{START}' must be YYYY-MM-DD format.")
        days_ago = (datetime.now(timezone.utc) - start_dt).days
        if days_ago > max_days:
            raise ValueError(
                f"\n  ❌  '{INTERVAL}' only covers the last {max_days} days.\n"
                f"      START='{START}' is {days_ago} days ago.\n\n"
                f"  ✅  FIX:\n"
                f"      • Use a recent START within last {max_days} days, OR\n"
                f"      • Set START=None and PERIOD='60d', OR\n"
                f"      • Use INTERVAL='1d' for any historical range\n"
            )


# ───────────────────────────────────────────────────────────────
def compute_supertrend(df):
    """
    Wilder ATR-based Supertrend.
    Adds: ST_val (line price), ST_bull (True = bullish = line below price)
    """
    high  = df["High"].values
    low   = df["Low"].values
    close = df["Close"].values
    n, p, m = len(df), ST_PERIOD, ST_MULTIPLIER

    tr = np.maximum(high[1:] - low[1:],
         np.maximum(np.abs(high[1:] - close[:-1]),
                    np.abs(low[1:]  - close[:-1])))

    atr = np.full(n, np.nan)
    atr[p] = tr[:p].mean()
    for j in range(p + 1, n):
        atr[j] = (atr[j-1] * (p - 1) + tr[j-1]) / p

    hl2         = (high + low) / 2.0
    upper_basic = hl2 + m * atr
    lower_basic = hl2 - m * atr

    upper = np.full(n, np.nan)
    lower = np.full(n, np.nan)
    bull  = np.full(n, True, dtype=bool)

    for j in range(p, n):
        if np.isnan(upper[j-1]):
            upper[j] = upper_basic[j]
            lower[j] = lower_basic[j]
            bull[j]  = close[j] >= lower[j]
            continue
        upper[j] = (upper_basic[j]
                    if upper_basic[j] < upper[j-1] or close[j-1] > upper[j-1]
                    else upper[j-1])
        lower[j] = (lower_basic[j]
                    if lower_basic[j] > lower[j-1] or close[j-1] < lower[j-1]
                    else lower[j-1])
        bull[j] = (close[j] >= lower[j]) if bull[j-1] else (close[j] > upper[j])

    df = df.copy()
    df["ST_val"]       = np.where(bull, lower, upper)
    df["ST_bull"]      = bull
    df["ST_direction"] = np.where(bull, "Bullish", "Bearish")
    return df


# ───────────────────────────────────────────────────────────────
def fetch_data():
    validate()
    rng = f"{START} → {END}" if START and END else PERIOD
    print(f"\nFetching {TICKER} | {INTERVAL} | {rng} …")

    kw = dict(interval=INTERVAL, auto_adjust=True, progress=True)
    try:
        df = (yf.download(TICKER, start=START, end=END, **kw)
              if START and END else
              yf.download(TICKER, period=PERIOD, **kw))
    except Exception as e:
        raise ValueError(f"\n  ❌  Download error: {e}\n  Try: pip install --upgrade yfinance\n")

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df = df.loc[:, ~df.columns.duplicated()]

    required = {"Open", "High", "Low", "Close", "Volume"}
    if not required.issubset(df.columns) or df.empty:
        raise ValueError(
            f"\n  ❌  No data for '{TICKER}'.\n"
            f"      Nifty50='^NSEI'  |  Reliance='RELIANCE.NS'\n"
            f"      Check ticker spelling and date range.\n"
        )

    df = df[list(required)].dropna(subset=["Open", "High", "Low", "Close"])
    df.index = pd.to_datetime(df.index)
    if df.index.tz is None:
        df.index = df.index.tz_localize("UTC")
    df.index = df.index.tz_convert(IST)

    df = compute_supertrend(df)

    # ── ADX ──────────────────────────────────────────────────────────────
    adx_result = ta.adx(df["High"], df["Low"], df["Close"], length=ADX_PERIOD)
    if adx_result is not None:
        adx_col = [c for c in adx_result.columns if c.startswith("ADX_")]
        df["ADX"] = adx_result[adx_col[0]].values if adx_col else np.nan
    else:
        df["ADX"] = np.nan

    df.dropna(inplace=True)
    print(f"  {len(df)} candles  |  {fmt(df.index[0])} → {fmt(df.index[-1])}")
    return df


# ───────────────────────────────────────────────────────────────
def run_backtest(df):
    """
    LONG  : ST bull (ST line below price) → enter at close
            ADX FILTER: only enter if ADX >= ADX_MIN_ENTRY (15)
            Stop Loss = current ST_val (dynamic, updates every candle)
            Hard SL   = entry_price * (1 - HARD_SL_PCT/100)  [if HARD_SL_ENABLED]
            Exit when: Hard SL hit  OR  price LOW ≤ ST line  OR  EOD
            (Hard SL is checked first — whichever fires first wins)

    SHORT : ST bear (ST line above price) → enter at close
            ADX FILTER: only enter if ADX >= ADX_MIN_ENTRY (15)
            Stop Loss = current ST_val (dynamic, updates every candle)
            Hard SL   = entry_price * (1 + HARD_SL_PCT/100)  [if HARD_SL_ENABLED]
            Exit when: Hard SL hit  OR  price HIGH ≥ ST line  OR  EOD
            (Hard SL is checked first — whichever fires first wins)

    Re-entry after ST flip: allowed as before (ADX filter applies to re-entries too).
    Re-entry after Hard SL: blocked (hard SL means the move failed, not a clean flip).

    ADDITIONAL RULES:
    1. First candle (9:15) of each day is always skipped — no entry, no analysis.
    2. No new entries after 15:15. Existing positions can be held until EOD at 15:30.
    3. ST flat + price converging filter: if ST_val is unchanged from previous candle
       AND |close - ST| is getting smaller → price is drifting toward ST line.
       Skip entry and wait until close moves AWAY from ST (distance expanding).
    """
    trades    = []
    adx_skips = []   # records every candle where ADX < ADX_MIN_ENTRY blocked an entry

    in_trade      = False
    direction     = None
    entry_price   = None
    entry_time    = None
    peak          = None        # highest (long) / lowest (short)
    hard_sl_price = None        # fixed price level set at entry; None if disabled
    was_in_profit = False       # True once price moved in profit direction from entry

    closes   = df["Close"].to_numpy(dtype=float)
    highs    = df["High"].to_numpy(dtype=float)
    lows     = df["Low"].to_numpy(dtype=float)
    st_vals  = df["ST_val"].to_numpy(dtype=float)
    st_bulls = df["ST_bull"].to_numpy(dtype=bool)
    adx_vals = df["ADX"].to_numpy(dtype=float)
    times    = df.index

    # Track first candle time per day to enforce the skip-first-candle rule
    first_candle_of_day = {}   # date → timestamp of the first candle that day
    for t in times:
        d = t.date()
        if d not in first_candle_of_day:
            first_candle_of_day[d] = t

    def record(dir_, e_time, e_price, x_time, x_price, pk, sl_at_exit, reason, hard_sl):
        if dir_ == "long":
            pnl = round((x_price - e_price) / e_price * 100, 4)
        else:
            pnl = round((e_price - x_price) / e_price * 100, 4)
        pts = round(abs(x_price - e_price), 4)
        trades.append({
            "Date"            : e_time.strftime("%Y-%m-%d"),
            "Direction"       : "Long  ↑" if dir_ == "long" else "Short ↓",
            "Entry Time"      : fmt(e_time),
            "Entry Price"     : round(e_price, 2),
            "Hard SL Price"   : round(hard_sl, 2) if hard_sl is not None else "OFF",
            "ST Stop Loss"    : round(sl_at_exit, 2),   # ST value at exit candle
            "Exit Time"       : fmt(x_time),
            "Exit Price"      : round(x_price, 2),
            "Peak"            : round(pk, 2),
            "Points Captured" : pts,
            "P&L %"           : pnl,
            "Exit Reason"     : reason,
            "Result"          : "WIN" if pnl > 0 else "LOSS",
        })

    current_day = None   # tracks the trading date; force-exit if it changes

    for i in range(1, len(df)):
        c_close   = closes[i]
        c_high    = highs[i]
        c_low     = lows[i]
        c_time    = times[i]
        c_tod     = c_time.time()
        c_st       = st_vals[i]       # current ST line value = dynamic stop loss
        c_bull     = st_bulls[i]
        c_adx      = adx_vals[i]      # current ADX value
        prev_bull  = st_bulls[i-1]
        prev_st    = st_vals[i-1]     # previous ST value
        prev_close = closes[i-1]      # previous candle close

        flipped_bull = (not prev_bull) and c_bull   # bear → bull
        flipped_bear = prev_bull and (not c_bull)   # bull → bear

        # ── RULE 1: Skip first 5-min candle of the day (9:15 candle) ──────
        if c_time == first_candle_of_day.get(c_time.date()):
            continue

        # ── RULE 3 helper: ST flat + price converging check ───────────────
        # ST is "flat" when it has not moved from previous candle
        st_is_flat = (c_st == prev_st)
        # For long setup (ST below price): converging = distance shrinking
        # i.e. (close - ST) now < (prev_close - prev_ST)
        long_converging  = st_is_flat and (c_close - c_st) < (prev_close - prev_st)
        # For short setup (ST above price): converging = distance shrinking
        # i.e. (ST - close) now < (prev_ST - prev_close)
        short_converging = st_is_flat and (c_st - c_close) < (prev_st - prev_close)

        # ── EOD FORCE EXIT — same day close mandatory ────────────────────
        # Case 1: candle is at or after 15:25 → close immediately
        # Case 2: date has changed (next trading day) → trade was never closed,
        #         force-exit at the LAST candle of the previous day (entry_date EOD)
        c_date = c_time.date()

        if in_trade and c_date != entry_time.date():
            # New day arrived while still in trade — exit at entry day's last known close
            # We use c_close of the first candle of the new day as a proxy (or
            # we exit at the previous candle's close to stay on the correct date)
            prev_close_price = closes[i - 1]
            prev_time        = times[i - 1]
            prev_st_val      = st_vals[i - 1]
            pk_final = max(peak, highs[i-1]) if direction == "long" else min(peak, lows[i-1])
            record(direction, entry_time, entry_price,
                   prev_time, prev_close_price, pk_final, prev_st_val,
                   "EOD Exit", hard_sl_price)
            in_trade      = False
            was_in_profit = False
            current_day   = c_date

        if in_trade and c_tod >= MARKET_CLOSE:
            pk_final = max(peak, c_high) if direction == "long" else min(peak, c_low)
            record(direction, entry_time, entry_price,
                   c_time, c_close, pk_final, c_st, "EOD Exit", hard_sl_price)
            in_trade      = False
            was_in_profit = False
            continue

        # Skip candles outside market hours (EOD exit at 15:25 handled above)
        if c_tod < MARKET_OPEN or c_tod >= MARKET_CLOSE:
            continue

        # ── MANAGE OPEN TRADE ──────────────────────────────────────────
        # NOTE: ADX filter does NOT apply here — exit logic is unchanged
        if in_trade:
            if direction == "long":
                if c_high > peak: peak = c_high
                if c_close > entry_price: was_in_profit = True   # trade touched profit territory

                # Hard SL checked FIRST — fires before ST SL if both hit same candle
                hard_sl_hit = (HARD_SL_ENABLED and
                               hard_sl_price is not None and
                               c_low <= hard_sl_price)
                st_sl_hit   = c_low <= c_st or flipped_bear

                if hard_sl_hit or st_sl_hit:
                    if hard_sl_hit:
                        exit_px = round(hard_sl_price, 2)
                        reason  = f"Hard SL ({HARD_SL_PCT}%)"
                    else:
                        exit_px = round(c_st, 2)
                        if c_low <= c_st:
                            reason = "ST Profit Exit" if was_in_profit else "ST Stop Loss"
                        else:
                            reason = "ST Flip Bear (Profit)" if was_in_profit else "ST Flip Bear"
                    record("long", entry_time, entry_price,
                           c_time, exit_px, peak, c_st, reason, hard_sl_price)
                    in_trade = False

                    # Re-entry into short only on a clean ST flip (not a hard SL exit)
                    # All entry filters apply to re-entries too
                    if (not hard_sl_hit and not c_bull
                            and c_adx >= ADX_MIN_ENTRY
                            and c_tod < NO_NEW_TRADE_AFTER
                            and not short_converging):
                        entry_price   = c_close; entry_time = c_time
                        peak          = c_low;   direction  = "short"; in_trade = True
                        hard_sl_price = round(entry_price * (1 + HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
                        was_in_profit = False

            elif direction == "short":
                if c_low < peak: peak = c_low
                if c_close < entry_price: was_in_profit = True   # trade touched profit territory

                # Hard SL checked FIRST
                hard_sl_hit = (HARD_SL_ENABLED and
                               hard_sl_price is not None and
                               c_high >= hard_sl_price)
                st_sl_hit   = c_high >= c_st or flipped_bull

                if hard_sl_hit or st_sl_hit:
                    if hard_sl_hit:
                        exit_px = round(hard_sl_price, 2)
                        reason  = f"Hard SL ({HARD_SL_PCT}%)"
                    else:
                        exit_px = round(c_st, 2)
                        if c_high >= c_st:
                            reason = "ST Profit Exit" if was_in_profit else "ST Stop Loss"
                        else:
                            reason = "ST Flip Bull (Profit)" if was_in_profit else "ST Flip Bull"
                    record("short", entry_time, entry_price,
                           c_time, exit_px, peak, c_st, reason, hard_sl_price)
                    in_trade = False

                    # Re-entry into long only on a clean ST flip (not a hard SL exit)
                    # All entry filters apply to re-entries too
                    if (not hard_sl_hit and c_bull
                            and c_adx >= ADX_MIN_ENTRY
                            and c_tod < NO_NEW_TRADE_AFTER
                            and not long_converging):
                        entry_price   = c_close; entry_time = c_time
                        peak          = c_high;  direction  = "long"; in_trade = True
                        hard_sl_price = round(entry_price * (1 - HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
                        was_in_profit = False
            continue

        # ── NOT IN TRADE — CHECK ENTRY ─────────────────────────────────

        # RULE 2: No new entries after 15:15 (existing positions held till 15:30)
        if c_tod >= NO_NEW_TRADE_AFTER:
            continue

        # ADX FILTER: skip entry if ADX < ADX_MIN_ENTRY (15) — sideways market
        if np.isnan(c_adx) or c_adx < ADX_MIN_ENTRY:
            # Log: what signal was blocked and what was the ADX value
            signal = "Long" if c_bull else "Short"
            adx_skips.append({
                "Date"      : c_time.strftime("%Y-%m-%d"),
                "Time"      : fmt(c_time),
                "Signal"    : signal,
                "ADX"       : round(c_adx, 2) if not np.isnan(c_adx) else "NaN",
                "Close"     : round(c_close, 2),
                "ST Val"    : round(c_st, 2),
            })
            continue

        # RULE 3: ST flat + price converging → wait, do not enter
        # Long  : ST unchanged AND close is getting closer to ST from above → skip
        # Short : ST unchanged AND close is getting closer to ST from below → skip
        if c_bull and long_converging:
            continue   # price drifting toward ST — wait for it to move away
        if not c_bull and short_converging:
            continue   # price drifting toward ST — wait for it to move away

        if c_bull:                          # ST below price → LONG
            entry_price   = c_close; entry_time = c_time
            peak          = c_high;  direction  = "long"; in_trade = True
            hard_sl_price = round(entry_price * (1 - HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
            was_in_profit = False

        elif not c_bull:                    # ST above price → SHORT
            entry_price   = c_close; entry_time = c_time
            peak          = c_low;   direction  = "short"; in_trade = True
            hard_sl_price = round(entry_price * (1 + HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
            was_in_profit = False

    return trades, adx_skips


# ───────────────────────────────────────────────────────────────
def build_daily_summary(trades):
    """Compute per-day P&L breakdown from trade list."""
    if not trades:
        return pd.DataFrame()

    df_t = pd.DataFrame(trades)
    daily = []

    for date, grp in df_t.groupby("Date"):
        total   = len(grp)
        wins    = (grp["P&L %"] > 0).sum()
        losses  = total - wins
        pnl_sum = grp["P&L %"].sum()
        win_pts = grp.loc[grp["P&L %"] > 0, "Points Captured"].sum()
        los_pts = grp.loc[grp["P&L %"] <= 0, "Points Captured"].sum()
        longs   = grp["Direction"].str.contains("Long").sum()
        shorts  = grp["Direction"].str.contains("Short").sum()
        daily.append({
            "Date"           : date,
            "Total Trades"   : total,
            "Longs"          : longs,
            "Shorts"         : shorts,
            "Winners"        : wins,
            "Losers"         : losses,
            "Win Rate %"     : round(wins / total * 100, 1),
            "Total P&L %"    : round(pnl_sum, 4),
            "Best Trade %"   : round(grp["P&L %"].max(), 4),
            "Worst Trade %"  : round(grp["P&L %"].min(), 4),
            "Points Captured": round(win_pts, 2),
            "Points Lost"    : round(los_pts, 2),
            "Net Points"     : round(win_pts - los_pts, 2),
            "Day Result"     : "✅ Profit" if pnl_sum > 0 else "❌ Loss" if pnl_sum < 0 else "⚖ Flat",
        })

    return pd.DataFrame(daily)


# ───────────────────────────────────────────────────────────────
def print_results(trades, adx_skips=None):
    sep  = "═" * 118
    dash = "─" * 118

    print("\n" + sep)
    print(f"  Supertrend({ST_PERIOD},{ST_MULTIPLIER})  |  {TICKER}  |  {INTERVAL}  |  EOD Exit (all trades): {MARKET_CLOSE}  |  No new entry after: {NO_NEW_TRADE_AFTER}")
    print(f"  Long : ST below price  →  Stop Loss = ST line value  →  Exit when Low ≤ ST line")
    print(f"  Short: ST above price  →  Stop Loss = ST line value  →  Exit when High ≥ ST line")
    print(f"  🛑 Hard SL : {'ON  | Max loss = ' + str(HARD_SL_PCT) + '% from entry (fires before ST SL if hit first)' if HARD_SL_ENABLED else 'OFF (pure ST SL mode)'}")
    print(f"  📊 ADX Filter : ADX({ADX_PERIOD}) ≥ {ADX_MIN_ENTRY} to enter  |  Skip 9:15 candle  |  No entry if price converging to flat ST")
    print(sep)

    if not trades:
        print("  ⚠  No trades found. Try a longer PERIOD or different INTERVAL.")
        print(sep)
        return None, None

    df_t   = pd.DataFrame(trades)
    df_day = build_daily_summary(trades)

    # ── TRADE LOG ──────────────────────────────────────────────────────
    print("\n  TRADE LOG")
    print(dash)
    cols = ["Date", "Direction", "Entry Time", "Entry Price", "Hard SL Price",
            "ST Stop Loss", "Exit Time", "Exit Price", "Points Captured",
            "P&L %", "Exit Reason", "Result"]
    print(df_t[cols].to_string(index=False))

    # ── ADX FILTER REPORT ─────────────────────────────────────────────
    if adx_skips:
        df_sk = pd.DataFrame(adx_skips)
        print("\n\n" + sep)
        print(f"  ADX FILTER REPORT  —  {len(adx_skips)} signals blocked (ADX < {ADX_MIN_ENTRY})")
        print(dash)
        # Per-day summary of skips
        skip_by_day = df_sk.groupby("Date").agg(
            Blocked=("Time", "count"),
            Longs_Blocked=("Signal", lambda x: (x=="Long").sum()),
            Shorts_Blocked=("Signal", lambda x: (x=="Short").sum()),
            Avg_ADX=("ADX", lambda x: round(pd.to_numeric(x, errors="coerce").mean(), 2)),
            Min_ADX=("ADX", lambda x: round(pd.to_numeric(x, errors="coerce").min(), 2)),
            Max_ADX=("ADX", lambda x: round(pd.to_numeric(x, errors="coerce").max(), 2)),
        ).reset_index()
        print("  Per-Day Blocked Signals:")
        print(skip_by_day.to_string(index=False))
        print(dash)
        # Overall ADX skip stats
        adx_num = pd.to_numeric(df_sk["ADX"], errors="coerce")
        long_blocked  = (df_sk["Signal"] == "Long").sum()
        short_blocked = (df_sk["Signal"] == "Short").sum()
        print(f"  Total Blocked         : {len(adx_skips)}  (Long: {long_blocked}  Short: {short_blocked})")
        print(f"  ADX Range (blocked)   : {adx_num.min():.2f} – {adx_num.max():.2f}  (avg: {adx_num.mean():.2f})")
        print(f"  Days with ADX blocks  : {df_sk['Date'].nunique()}")
        # Distribution buckets
        b0 = (adx_num < 10).sum()
        b1 = ((adx_num >= 10) & (adx_num < 12)).sum()
        b2 = ((adx_num >= 12) & (adx_num < ADX_MIN_ENTRY)).sum()
        print(f"  ADX < 10              : {b0} signals")
        print(f"  ADX 10–12             : {b1} signals")
        print(f"  ADX 12–{ADX_MIN_ENTRY}            : {b2} signals")

    # ── PER-DAY P&L ────────────────────────────────────────────────────
    print("\n\n" + sep)
    print("  PER-DAY P&L BREAKDOWN")
    print(dash)
    day_cols = ["Date", "Total Trades", "Longs", "Shorts", "Winners", "Losers",
                "Win Rate %", "Total P&L %", "Best Trade %", "Worst Trade %",
                "Points Captured", "Points Lost", "Net Points", "Day Result"]
    print(df_day[day_cols].to_string(index=False))

    # ── OVERALL SUMMARY ────────────────────────────────────────────────
    total      = len(df_t)
    wins       = (df_t["P&L %"] > 0).sum()
    losses     = total - wins
    hard_exits = df_t["Exit Reason"].str.startswith("Hard SL").sum()
    sl_exits       = (df_t["Exit Reason"] == "ST Stop Loss").sum()
    st_profit_exits= (df_t["Exit Reason"] == "ST Profit Exit").sum()
    st_exits       = df_t["Exit Reason"].str.contains("ST Flip").sum()
    eod_exits  = (df_t["Exit Reason"] == "EOD Exit").sum()
    win_pts    = df_t.loc[df_t["P&L %"] > 0, "Points Captured"].sum()
    loss_pts   = df_t.loc[df_t["P&L %"] <= 0, "Points Captured"].sum()
    net_pts    = win_pts - loss_pts

    profit_days = (df_day["Total P&L %"] > 0).sum()
    loss_days   = (df_day["Total P&L %"] < 0).sum()
    flat_days   = len(df_day) - profit_days - loss_days
    best_day    = df_day.loc[df_day["Total P&L %"].idxmax()]
    worst_day   = df_day.loc[df_day["Total P&L %"].idxmin()]

    print("\n\n" + sep)
    print("  OVERALL SUMMARY")
    print(dash)
    print(f"  {'HARD SL CONFIG':─<50}")
    print(f"  Hard Stop Loss        : {'ENABLED  (' + str(HARD_SL_PCT) + '% from entry)' if HARD_SL_ENABLED else 'DISABLED  (pure ST SL mode)'}")
    print(dash)
    print(f"  {'ADX FILTER':─<50}")
    print(f"  ADX Period            : {ADX_PERIOD}")
    print(f"  ADX Min Entry         : {ADX_MIN_ENTRY}  (trades skipped when ADX < {ADX_MIN_ENTRY})")
    print(dash)
    print(f"  {'TRADE STATS':─<50}")
    print(f"  Total Trades          : {total}")
    print(f"  Winners               : {wins}  ({wins/total*100:.1f}%)")
    print(f"  Losers                : {losses}  ({losses/total*100:.1f}%)")
    print(f"  Hard SL Exits         : {hard_exits}  (max {HARD_SL_PCT}% loss cap triggered)")
    print(f"  ST Stop Loss Exits    : {sl_exits}  (price never went into profit)")
    print(f"  ST Profit Exit        : {st_profit_exits}  (hit ST after being in profit — not a loss)")
    print(f"  ST Flip Exits         : {st_exits}  (trend reversal)")
    print(f"  EOD Force Exits       : {eod_exits}")
    print(dash)
    print(f"  {'P&L STATS':─<50}")
    print(f"  Total P&L             : {df_t['P&L %'].sum():.4f}%")
    print(f"  Average P&L / Trade   : {df_t['P&L %'].mean():.4f}%")
    print(f"  Best Single Trade     : {df_t['P&L %'].max():.4f}%")
    print(f"  Worst Single Trade    : {df_t['P&L %'].min():.4f}%")
    print(dash)
    print(f"  {'POINTS':─<50}")
    print(f"  Total Points Captured : {win_pts:.2f}  (winning trades only)")
    print(f"  Points Lost           : {loss_pts:.2f}  (losing trades)")
    print(f"  Net Points            : {net_pts:.2f}")
    print(dash)
    print(f"  {'DAY STATS':─<50}")
    print(f"  Total Trading Days    : {len(df_day)}")
    print(f"  Profit Days           : {profit_days}  ({profit_days/len(df_day)*100:.1f}%)")
    print(f"  Loss Days             : {loss_days}  ({loss_days/len(df_day)*100:.1f}%)")
    print(f"  Flat Days             : {flat_days}")
    print(f"  Best Day              : {best_day['Date']}  →  {best_day['Total P&L %']:.4f}%  |  Net Pts: {best_day['Net Points']:.2f}")
    print(f"  Worst Day             : {worst_day['Date']}  →  {worst_day['Total P&L %']:.4f}%  |  Net Pts: {worst_day['Net Points']:.2f}")
    print(sep + "\n")

    return df_t, df_day


# ───────────────────────────────────────────────────────────────
def export_excel(df, trades, adx_skips, df_trades, df_day):
    ticker_clean = TICKER.replace("^", "").replace(".", "_")
    fname = f"{ticker_clean}_{INTERVAL}_supertrend.xlsx"

    # ── Sheet 1: Supertrend raw data ───────────────────────────────────
    df_st = df[["Open", "High", "Low", "Close", "Volume", "ST_val", "ST_direction"]].copy()
    df_st.index = df_st.index.strftime("%Y-%m-%d %H:%M")
    df_st.index.name = "DateTime (IST)"
    df_st.columns = ["Open", "High", "Low", "Close", "Volume",
                     f"ST Value ({ST_PERIOD},{ST_MULTIPLIER})", "Direction"]
    df_st = df_st.round(2)

    # ── Sheet 4: Summary stats ─────────────────────────────────────────
    if df_trades is not None and not df_trades.empty:
        total     = len(df_trades)
        wins      = (df_trades["P&L %"] > 0).sum()
        losses    = total - wins
        win_pts   = df_trades.loc[df_trades["P&L %"] > 0, "Points Captured"].sum()
        loss_pts  = df_trades.loc[df_trades["P&L %"] <= 0, "Points Captured"].sum()
        profit_days = int((df_day["Total P&L %"] > 0).sum()) if df_day is not None else 0
        loss_days   = int((df_day["Total P&L %"] < 0).sum()) if df_day is not None else 0

        summary_rows = [
            ["STRATEGY INFO",    ""],
            ["Ticker",           TICKER],
            ["Interval",         INTERVAL],
            ["Supertrend",       f"Period={ST_PERIOD}, Multiplier={ST_MULTIPLIER}"],
            ["Stop Loss",        "Dynamic — ST line value"],
            ["EOD Exit",         str(MARKET_CLOSE)],
            ["", ""],
            ["HARD SL CONFIG",   ""],
            ["Hard Stop Loss",   f"{'ENABLED' if HARD_SL_ENABLED else 'DISABLED'} — {HARD_SL_PCT}% max loss from entry"],
            ["", ""],
            ["ADX FILTER",       ""],
            ["ADX Period",       ADX_PERIOD],
            ["ADX Min Entry",    f"{ADX_MIN_ENTRY}  (no entry if ADX < {ADX_MIN_ENTRY})"],
            ["", ""],
            ["TRADE STATS",      ""],
            ["Total Trades",     total],
            ["Winners",          f"{wins} ({wins/total*100:.1f}%)"],
            ["Losers",           f"{losses} ({losses/total*100:.1f}%)"],
            ["Hard SL Exits",    int(df_trades["Exit Reason"].str.startswith("Hard SL").sum())],
            ["ST Stop Loss Exits", int((df_trades["Exit Reason"] == "ST Stop Loss").sum())],
            ["ST Profit Exits",  int((df_trades["Exit Reason"] == "ST Profit Exit").sum())],
            ["ST Flip Exits",    int(df_trades["Exit Reason"].str.contains("ST Flip").sum())],
            ["ADX Blocked",      len(adx_skips) if adx_skips else 0],
            ["EOD Exits",        int((df_trades["Exit Reason"] == "EOD Exit").sum())],
            ["", ""],
            ["P&L STATS",        ""],
            ["Total P&L %",      round(df_trades["P&L %"].sum(), 4)],
            ["Avg P&L % / Trade",round(df_trades["P&L %"].mean(), 4)],
            ["Best Trade %",     round(df_trades["P&L %"].max(), 4)],
            ["Worst Trade %",    round(df_trades["P&L %"].min(), 4)],
            ["", ""],
            ["POINTS",           ""],
            ["Points Captured (wins)", round(win_pts, 2)],
            ["Points Lost",      round(loss_pts, 2)],
            ["Net Points",       round(win_pts - loss_pts, 2)],
            ["", ""],
            ["DAY STATS",        ""],
            ["Total Trading Days", len(df_day) if df_day is not None else 0],
            ["Profit Days",      profit_days],
            ["Loss Days",        loss_days],
        ]
        if df_day is not None and not df_day.empty:
            best  = df_day.loc[df_day["Total P&L %"].idxmax()]
            worst = df_day.loc[df_day["Total P&L %"].idxmin()]
            summary_rows += [
                ["Best Day",     f"{best['Date']}  →  {best['Total P&L %']:.4f}%"],
                ["Worst Day",    f"{worst['Date']}  →  {worst['Total P&L %']:.4f}%"],
            ]
        df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value"])
    else:
        df_summary = pd.DataFrame([{"Metric": "No trades found", "Value": ""}])

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        df_st.to_excel(writer, sheet_name="Supertrend Data", index=True)
        if df_trades is not None and not df_trades.empty:
            df_trades.to_excel(writer, sheet_name="Trade Log", index=False)
        if df_day is not None and not df_day.empty:
            df_day.to_excel(writer, sheet_name="Daily P&L", index=False)
        if adx_skips:
            pd.DataFrame(adx_skips).to_excel(writer, sheet_name="ADX Blocked Signals", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    _style_excel(fname, df_st, df_trades, df_day)
    print(f"  Excel saved  → {fname}")
    return fname


def _style_excel(fname, df_st, df_trades, df_day):
    GREEN_DARK  = "1A5C38";  GREEN_LIGHT = "C6EFCE"
    RED_DARK    = "9C0006";  RED_LIGHT   = "FFC7CE"
    BLUE_HDR    = "1F3864";  YELLOW_SEC  = "FFD700"
    GRAY_ALT    = "F2F2F2";  WHITE       = "FFFFFF"
    ORANGE      = "FCE4D6";  TEAL_LIGHT  = "DDEBF7"
    PURPLE_LIGHT= "E2EFDA"

    thin_s = Side(style="thin", color="CCCCCC")
    bdr    = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    def hdr(ws, row=1, bg=BLUE_HDR, fg=WHITE):
        for cell in ws[row]:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(bold=True, color=fg, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = bdr
        ws.row_dimensions[row].height = 22

    def autowidth(ws, cap=28):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, cap)

    wb = load_workbook(fname)

    # ── Sheet 1: Supertrend Data ─────────────────────────────────────────
    ws = wb["Supertrend Data"]
    hdr(ws)
    ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        dir_cell = row[-1]
        val      = str(dir_cell.value or "")
        if val == "Bullish":
            bg = GREEN_LIGHT;  fg_col = GREEN_DARK
        elif val == "Bearish":
            bg = RED_LIGHT;    fg_col = RED_DARK
        else:
            bg = WHITE;        fg_col = "000000"
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="center")
        dir_cell.font = Font(bold=True, color=fg_col)
    autowidth(ws)

    # ── Sheet 2: Trade Log ───────────────────────────────────────────────
    if "Trade Log" in wb.sheetnames and df_trades is not None:
        ws = wb["Trade Log"]
        hdr(ws)
        ws.freeze_panes = "A2"
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            result_cell = row[-1]
            result      = str(result_cell.value or "")
            reason_cell = row[-2]
            reason      = str(reason_cell.value or "")
            if result == "WIN":
                bg = GREEN_LIGHT;  result_cell.font = Font(bold=True, color=GREEN_DARK)
            elif result == "LOSS":
                bg = RED_LIGHT;    result_cell.font = Font(bold=True, color=RED_DARK)
            else:
                bg = GRAY_ALT if i % 2 == 0 else WHITE
            # Hard SL → solid red (most severe), ST SL → orange, EOD → teal
            if reason.startswith("Hard SL"):
                reason_cell.fill = PatternFill("solid", fgColor="FF0000")
                reason_cell.font = Font(bold=True, color="FFFFFF")
            elif "Stop Loss" in reason:
                reason_cell.fill = PatternFill("solid", fgColor=ORANGE)
                reason_cell.font = Font(bold=True, color="C55A11")
            elif "EOD" in reason:
                reason_cell.fill = PatternFill("solid", fgColor=TEAL_LIGHT)
                reason_cell.font = Font(bold=True, color="2E75B6")
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = PatternFill("solid", fgColor=bg)
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    # ── Sheet 3: Daily P&L ───────────────────────────────────────────────
    if "Daily P&L" in wb.sheetnames and df_day is not None:
        ws = wb["Daily P&L"]
        hdr(ws)
        ws.freeze_panes = "A2"
        result_col = ws.max_column
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            day_result_cell = row[result_col - 1]
            day_result      = str(day_result_cell.value or "")
            if "Profit" in day_result:
                bg = GREEN_LIGHT;  day_result_cell.font = Font(bold=True, color=GREEN_DARK)
            elif "Loss" in day_result:
                bg = RED_LIGHT;    day_result_cell.font = Font(bold=True, color=RED_DARK)
            else:
                bg = GRAY_ALT
            for cell in row:
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    # ── Sheet 4: Summary ─────────────────────────────────────────────────
    ws = wb["Summary"]
    hdr(ws)
    SECTION_LABELS = {"STRATEGY INFO", "HARD SL CONFIG", "ADX FILTER", "TRADE STATS", "P&L STATS", "POINTS", "DAY STATS"}
    HIGHLIGHT_ROWS = {"Hard Stop Loss", "ADX Min Entry"}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        label = str(row[0].value or "")
        if label in SECTION_LABELS:
            for cell in row:
                cell.fill   = PatternFill("solid", fgColor=YELLOW_SEC)
                cell.font   = Font(bold=True, color="000000", size=10)
                cell.border = bdr
            continue
        if label in HIGHLIGHT_ROWS:
            for cell in row:
                cell.fill   = PatternFill("solid", fgColor=PURPLE_LIGHT)
                cell.font   = Font(bold=True, color="375623")
                cell.border = bdr
            continue
        val_cell = row[1] if len(row) > 1 else None
        for cell in row:
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="left")
        if val_cell:
            v = val_cell.value
            if "Net Points" in label:
                color = GREEN_LIGHT if isinstance(v, (int, float)) and v > 0 else RED_LIGHT
                val_cell.fill = PatternFill("solid", fgColor=color)
                val_cell.font = Font(bold=True,
                    color=GREEN_DARK if isinstance(v, (int, float)) and v > 0 else RED_DARK)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    # ── ADX Blocked Signals sheet ────────────────────────────────────────
    if "ADX Blocked Signals" in wb.sheetnames:
        ws = wb["ADX Blocked Signals"]
        hdr(ws)
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            sig = str(row[2].value or "") if len(row) > 2 else ""
            bg  = "FFF2CC" if sig == "Long" else "FCE4D6" if sig == "Short" else GRAY_ALT
            for cell in row:
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    wb.save(fname)


# ───────────────────────────────────────────────────────────────
def build_chart(df, trades):
    if CHART_DAYS:
        unique_days = sorted(df.index.normalize().unique())
        cutoff = unique_days[-CHART_DAYS] if len(unique_days) >= CHART_DAYS else unique_days[0]
        df_c   = df[df.index.normalize() >= cutoff].copy()
    else:
        df_c = df.copy()

    st_green = df_c["ST_val"].where(df_c["ST_bull"])
    st_red   = df_c["ST_val"].where(~df_c["ST_bull"])

    fig = go.Figure()

    # Candlestick
    fig.add_trace(go.Candlestick(
        x=df_c.index, open=df_c["Open"], high=df_c["High"],
        low=df_c["Low"], close=df_c["Close"], name="Price",
        increasing_line_color="#26a69a", decreasing_line_color="#ef5350"
    ))

    # ST line — bullish (green)
    fig.add_trace(go.Scatter(
        x=df_c.index, y=st_green, name="ST Bullish (Stop Loss for Long)",
        mode="lines", line=dict(color="#22c55e", width=2.5), connectgaps=False
    ))

    # ST line — bearish (red)
    fig.add_trace(go.Scatter(
        x=df_c.index, y=st_red, name="ST Bearish (Stop Loss for Short)",
        mode="lines", line=dict(color="#ef4444", width=2.5), connectgaps=False
    ))

    # Trade markers
    if trades:
        df_t = pd.DataFrame(trades)
        in_win = df_c.index[0]

        def parse(col):
            return pd.to_datetime(df_t[col], format="%Y-%m-%d %H:%M").dt.tz_localize(IST)

        et = parse("Entry Time");  xt = parse("Exit Time")
        ep = df_t["Entry Price"];  xp = df_t["Exit Price"]
        dr = df_t["Direction"];    rs = df_t["Result"]
        ex = df_t["Exit Reason"]

        mask = et >= in_win

        lx  = et[mask & dr.str.contains("Long")].tolist()
        ly  = ep[mask & dr.str.contains("Long")].tolist()
        sx  = et[mask & dr.str.contains("Short")].tolist()
        sy  = ep[mask & dr.str.contains("Short")].tolist()
        wx  = xt[mask & (rs == "WIN")].tolist()
        wy  = xp[mask & (rs == "WIN")].tolist()
        lox = xt[mask & (rs == "LOSS")].tolist()
        loy = xp[mask & (rs == "LOSS")].tolist()
        slx = xt[mask & (ex == "ST Stop Loss")].tolist()
        sly = xp[mask & (ex == "ST Stop Loss")].tolist()
        hsx = xt[mask & ex.str.startswith("Hard SL")].tolist()
        hsy = xp[mask & ex.str.startswith("Hard SL")].tolist()

        if lx:
            fig.add_trace(go.Scatter(x=lx, y=ly, mode="markers", name="Long Entry",
                marker=dict(symbol="triangle-up", size=14, color="#22c55e",
                            line=dict(color="white", width=1))))
        if sx:
            fig.add_trace(go.Scatter(x=sx, y=sy, mode="markers", name="Short Entry",
                marker=dict(symbol="triangle-down", size=14, color="#ef4444",
                            line=dict(color="white", width=1))))
        if wx:
            fig.add_trace(go.Scatter(x=wx, y=wy, mode="markers", name="Exit WIN",
                marker=dict(symbol="circle", size=11, color="#86efac",
                            line=dict(color="#16a34a", width=2))))
        if lox:
            fig.add_trace(go.Scatter(x=lox, y=loy, mode="markers", name="Exit LOSS",
                marker=dict(symbol="x", size=12, color="#fca5a5",
                            line=dict(color="#dc2626", width=2))))
        if slx:
            fig.add_trace(go.Scatter(x=slx, y=sly, mode="markers", name="ST Stop Loss Hit",
                marker=dict(symbol="diamond", size=11, color="#fb923c",
                            line=dict(color="#c2410c", width=1.5))))
        if hsx:
            fig.add_trace(go.Scatter(x=hsx, y=hsy, mode="markers",
                name=f"Hard SL Hit ({HARD_SL_PCT}%)",
                marker=dict(symbol="hexagram", size=14, color="#ff0000",
                            line=dict(color="white", width=1.5))))

    hard_sl_label = f"HardSL {HARD_SL_PCT}% + " if HARD_SL_ENABLED else ""
    fig.update_layout(
        template="plotly_dark",
        height=750,
        xaxis_rangeslider_visible=False,
        legend=dict(orientation="h", yanchor="bottom", y=1.01,
                    xanchor="right", x=1, font=dict(size=10)),
        margin=dict(l=60, r=40, t=100, b=50),
        font=dict(family="monospace", size=11),
        title=dict(
            text=(f"<b>{TICKER}</b> | {INTERVAL} | Supertrend({ST_PERIOD},{ST_MULTIPLIER})<br>"
                  f"<span style='font-size:11px;color:#94a3b8'>"
                  f"Long: ST below price  |  Short: ST above price  |  "
                  f"{hard_sl_label}Stop Loss = ST line  |  ADX({ADX_PERIOD}) ≥ {ADX_MIN_ENTRY} to enter</span>"),
            x=0.5, xanchor="center"
        )
    )
    fig.update_yaxes(title_text="Price", showgrid=True, gridcolor="#1e293b")
    fig.update_xaxes(
        showgrid=True, gridcolor="#1e293b",
        rangebreaks=[
            dict(bounds=["sat", "mon"]),
            dict(bounds=[15.5, 9.25], pattern="hour")
        ]
    )

    chart_name = f"{TICKER.replace('^','').replace('.','_')}_{INTERVAL}_chart.html"
    fig.write_html(chart_name)
    print(f"  Chart saved  → {chart_name}  (open in browser)")
    return chart_name


# ───────────────────────────────────────────────────────────────
def main():
    try:
        df = fetch_data()
    except ValueError as e:
        print(e); return

    trades, adx_skips = run_backtest(df)
    df_trades, df_day = print_results(trades, adx_skips)

    print("  Exporting Excel …")
    export_excel(df, trades, adx_skips, df_trades, df_day)

    print("  Building chart …")
    build_chart(df, trades)

    print("\n  ✅  Done.\n")


if __name__ == "__main__":
    main()