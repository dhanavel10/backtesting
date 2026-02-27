"""
Supertrend High-Conviction Momentum Strategy
─────────────────────────────────────────────
PHILOSOPHY:
  Fewer trades. Bigger moves. Minimal losses.
  Every trade must earn its entry through multiple strict conditions.
  The strategy deliberately skips most signals — only the highest-quality
  breakouts from confirmed squeeze zones get a trade.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 ENTRY GATE  (ALL conditions must pass — one fail = no trade)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  1. SUPERTREND DIRECTION
       ST bull → only longs  |  ST bear → only shorts
       Direction must have been stable for ≥ MIN_ST_CANDLES candles
       (prevents trading right after a flip — flips are often false)

  2. SQUEEZE ZONE (price coiling for long enough)
       At least SQUEEZE_MIN_CANDLES of the last SQUEEZE_LOOKBACK candles
       must have body % < SQUEEZE_THRESHOLD
       (need sustained contraction, not just 1-2 quiet candles)

  3. POWER CANDLE (explosive breakout)
       Current candle body % > POWER_MULTIPLIER × rolling avg body %
       Candle colour must match ST direction (green for long, red for short)
       Candle body must be > POWER_MIN_BODY_PCT % (absolute floor, not just relative)

  4. RANGE BREAKOUT
       Long:  current candle High must exceed the high of the last SQUEEZE_LOOKBACK candles
       Short: current candle Low  must breach the low  of the last SQUEEZE_LOOKBACK candles
       (price must actually break out of the squeeze range, not just a big doji)

  5. ATR EXPANSION
       Current candle True Range > ATR_EXPANSION_FACTOR × rolling ATR average
       (overall volatility must be expanding, not just body)

  6. VOLUME SURGE
       Current volume > VOLUME_MIN_FACTOR × rolling volume average
       (real breakouts have fuel behind them)

  7. DAILY TRADE LIMIT
       Max MAX_TRADES_PER_DAY trades per calendar day
       After that, no new entries until next session

  8. POST-LOSS COOLDOWN
       After any losing exit (Hard SL or ST SL loss),
       wait COOLDOWN_CANDLES before looking for next entry
       (gives market time to settle — avoids revenge trading)

  9. NO AUTO RE-ENTRY after ST flip
       ST flip exits close the trade cleanly. The next entry must pass
       ALL conditions above from scratch — no free re-entry on flips.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 EXIT LOGIC  (first trigger wins)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  1. HARD STOP LOSS    : Fixed % loss from entry (capital protection)

  2. TRAILING STOP     : Once trade moves TRAIL_ACTIVATION_PCT in profit,
                         lock in TRAIL_LOCK_PCT of the move (trailing high/low)
                         This protects profits dynamically as the move extends

  3. MOMENTUM FADE     : After MOMENTUM_MIN_CANDLES candles in trade,
                         if MOMENTUM_FADE_CANDLES consecutive candles have
                         body % < MOMENTUM_FADE_THRESHOLD → exit (move dying)

  4. ST STOP LOSS      : Long: Low ≤ ST line | Short: High ≥ ST line

  5. ST FLIP           : Direction reversal — clean exit, no auto re-entry

  6. EOD FORCE EXIT    : 15:15 IST

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Install: pip install yfinance pandas plotly openpyxl pytz numpy
"""

import yfinance as yf
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pytz
import numpy as np
from datetime import datetime, timezone, time as dtime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════
#  ── CONFIGURATION  —  set True / False for every feature ───────
# ═══════════════════════════════════════════════════════════════

TICKER   = "^NSEI"
INTERVAL = "5m"
PERIOD   = "60d"
START    = None          # e.g. "2026-01-20"
END      = None          # e.g. "2026-02-20"

# ── Supertrend ──────────────────────────────────────────────────
ST_PERIOD     = 10
ST_MULTIPLIER = 3.0

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ENTRY GATES  (set True = ON, False = OFF for each gate)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Gate 1 — ST Stability
# ST must stay in same direction for MIN_ST_CANDLES consecutive candles before entry.
# Prevents trading fresh/uncertain flips. True = wait for stable trend.
GATE_ST_STABILITY   = True
MIN_ST_CANDLES      = 3        # only used when GATE_ST_STABILITY = True

# Gate 2 — Squeeze Zone
# At least SQUEEZE_MIN_CANDLES of the last SQUEEZE_LOOKBACK candles must be quiet
# (body % < SQUEEZE_THRESHOLD). Price must be coiling before we enter.
# True = require contraction before breakout.
GATE_SQUEEZE        = True
SQUEEZE_LOOKBACK    = 4
SQUEEZE_MIN_CANDLES = 2        # only used when GATE_SQUEEZE = True
SQUEEZE_THRESHOLD   = 0.18     # body % below this = quiet candle

# Gate 3 — Power Candle (Relative)
# Breakout candle body must be >= POWER_MULTIPLIER × rolling avg body.
# True = demand an explosive candle vs recent average.
GATE_POWER_RELATIVE = True
POWER_MULTIPLIER    = 2.5      # only used when GATE_POWER_RELATIVE = True
POWER_LOOKBACK      = 8

# Gate 3b — Power Candle (Absolute Floor)
# Candle body must be >= POWER_MIN_BODY_PCT % in absolute terms regardless of avg.
# Prevents entering on a "relatively big" candle that is still tiny in flat markets.
# True = enforce a minimum body size floor.
GATE_POWER_FLOOR    = True
POWER_MIN_BODY_PCT  = 0.20     # only used when GATE_POWER_FLOOR = True

# Gate 3c — Candle Colour Direction Match
# Bullish ST → only green (close >= open) breakout candle.
# Bearish ST → only red  (close <  open) breakout candle.
# True = candle colour must confirm the ST direction.
GATE_CANDLE_COLOUR  = True

# Gate 4 — Range Breakout (New High / New Low)
# Long:  breakout candle High must exceed the highest High of the squeeze window.
# Short: breakout candle Low  must breach the lowest  Low  of the squeeze window.
# True = price must actually escape the squeeze box, not just be a big candle inside it.
GATE_RANGE_BREAKOUT = True

# Gate 5 — ATR Expansion
# Current candle True Range must be >= ATR_EXPANSION_FACTOR × rolling ATR average.
# True = overall volatility must be expanding at the moment of entry.
GATE_ATR_EXPANSION  = True
ATR_LOOKBACK        = 14
ATR_EXPANSION_FACTOR = 0.8     # only used when GATE_ATR_EXPANSION = True

# Gate 6 — Volume Surge
# Current volume must be >= VOLUME_MIN_FACTOR × rolling volume average.
# True = real breakouts have fuel; enforce volume confirmation.
GATE_VOLUME_SURGE   = False
VOLUME_LOOKBACK     = 14
VOLUME_MIN_FACTOR   = 1.4      # only used when GATE_VOLUME_SURGE = True

# Gate 7 — Daily Trade Limit
# Hard cap on entries per calendar day. After MAX_TRADES_PER_DAY entries, stop.
# True = enforce the daily cap; False = unlimited entries per day.
GATE_DAILY_LIMIT    = True
MAX_TRADES_PER_DAY  = 3        # only used when GATE_DAILY_LIMIT = True

# Gate 8 — Post-Loss Cooldown
# After any losing exit, skip COOLDOWN_CANDLES candles before looking for entry.
# True = enforce cooldown period after a loss; False = re-hunt immediately.
GATE_COOLDOWN       = True
COOLDOWN_CANDLES    = 6        # only used when GATE_COOLDOWN = True

# Gate 9 — No Auto Re-entry on ST Flip
# When a ST flip closes a trade, the next entry must re-qualify through ALL active gates.
# True = strict re-qualification required; False = immediately re-enter on flip (old behaviour).
GATE_NO_AUTO_REENTRY = True

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  EXIT RULES  (set True = ON, False = OFF for each exit type)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Hard Stop Loss
# Fixed % loss cap from entry price. Fires before all other exits.
# True = capital protection always on; False = rely only on ST line as stop.
HARD_SL_ENABLED     = True
HARD_SL_PCT         = 0.3      # only used when HARD_SL_ENABLED = True

# Trailing Stop
# Once trade is TRAIL_ACTIVATION_PCT in profit, lock in TRAIL_LOCK_PCT of peak gain.
# Dynamically tightens as price moves further — protects profits while letting winner run.
# True = trail stop active; False = no trailing (only ST line and hard SL protect profit).
TRAIL_ENABLED        = True
TRAIL_ACTIVATION_PCT = 0.3     # only used when TRAIL_ENABLED = True
TRAIL_LOCK_PCT       = 0.6     # only used when TRAIL_ENABLED = True

# Momentum Fade Exit
# If MOMENTUM_FADE_CANDLES consecutive candles (after MOMENTUM_MIN_CANDLES settling period)
# have body % < MOMENTUM_FADE_THRESHOLD, the move has stalled — exit at close.
# True = exit when momentum dies inside the trade; False = hold until ST line or hard SL.
MOMENTUM_FADE_ENABLED   = True
MOMENTUM_MIN_CANDLES    = 3    # settling candles before fade check starts
MOMENTUM_FADE_CANDLES   = 4    # consecutive weak candles to trigger exit
MOMENTUM_FADE_THRESHOLD = 0.10 # body % below this = weak/fading candle

# ST Stop Loss  (recommended: always True)
# Long: exit when candle Low  <= ST line value.
# Short: exit when candle High >= ST line value.
# True = ST line acts as dynamic stop loss; False = ignore ST line as stop (not recommended).
ST_SL_ENABLED       = True

# ST Flip Exit
# When ST changes direction, close the current trade immediately at candle close.
# True = exit on trend reversal signal; False = ignore flips (only hard/trail SL exits).
ST_FLIP_EXIT_ENABLED = True

# EOD Force Exit at 15:15 IST
# Close all open positions at end of intraday session regardless of P&L.
# True = intraday only, no overnight; False = hold past close (risky, not recommended).
EOD_EXIT_ENABLED    = True

# ── Intraday session ────────────────────────────────────────────
MARKET_OPEN  = dtime(9, 15)
MARKET_CLOSE = dtime(15, 15)
IST          = pytz.timezone("Asia/Kolkata")

# ── Chart ───────────────────────────────────────────────────────
CHART_DAYS = 5

# ═══════════════════════════════════════════════════════════════
MAX_DAYS = {"1m": 7, "2m": 60, "5m": 60, "15m": 60,
            "30m": 60, "60m": 60, "90m": 60, "1h": 60}


def fmt(ts):
    return ts.strftime("%Y-%m-%d %H:%M") if pd.notna(ts) else "—"


def validate():
    if INTERVAL not in MAX_DAYS:
        return
    max_d = MAX_DAYS[INTERVAL]
    if START:
        start_dt = datetime.strptime(START, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        days_ago = (datetime.now(timezone.utc) - start_dt).days
        if days_ago > max_d:
            raise ValueError(
                f"\n  ❌  '{INTERVAL}' only covers last {max_d} days.\n"
                f"      START='{START}' is {days_ago} days ago.\n"
            )


# ─────────────────────────────────────────────────────────────
def compute_supertrend(df):
    high  = df["High"].values
    low   = df["Low"].values
    close = df["Close"].values
    n, p, m = len(df), ST_PERIOD, ST_MULTIPLIER

    tr = np.maximum(high[1:] - low[1:],
         np.maximum(np.abs(high[1:] - close[:-1]),
                    np.abs(low[1:]  - close[:-1])))
    atr = np.full(n, np.nan)
    atr[p] = tr[:p].mean()
    for j in range(p + 1, n):
        atr[j] = (atr[j-1] * (p - 1) + tr[j-1]) / p

    hl2         = (high + low) / 2.0
    upper_basic = hl2 + m * atr
    lower_basic = hl2 - m * atr
    upper = np.full(n, np.nan)
    lower = np.full(n, np.nan)
    bull  = np.full(n, True, dtype=bool)

    for j in range(p, n):
        if np.isnan(upper[j-1]):
            upper[j] = upper_basic[j]; lower[j] = lower_basic[j]
            bull[j]  = close[j] >= lower[j]; continue
        upper[j] = (upper_basic[j]
                    if upper_basic[j] < upper[j-1] or close[j-1] > upper[j-1]
                    else upper[j-1])
        lower[j] = (lower_basic[j]
                    if lower_basic[j] > lower[j-1] or close[j-1] < lower[j-1]
                    else lower[j-1])
        bull[j] = (close[j] >= lower[j]) if bull[j-1] else (close[j] > upper[j])

    df = df.copy()
    df["ST_val"]  = np.where(bull, lower, upper)
    df["ST_bull"] = bull
    df["ST_dir"]  = np.where(bull, "Bullish", "Bearish")
    return df


# ─────────────────────────────────────────────────────────────
def compute_indicators(df):
    df = df.copy()

    # Candle body as % of open
    df["body_pct"] = abs(df["Close"] - df["Open"]) / df["Open"] * 100

    # Rolling body avg (for power candle check)
    df["body_ma"]  = df["body_pct"].rolling(POWER_LOOKBACK, min_periods=POWER_LOOKBACK).mean()

    # True range
    prev_close = df["Close"].shift(1)
    df["tr"] = np.maximum(
        df["High"] - df["Low"],
        np.maximum(abs(df["High"] - prev_close), abs(df["Low"] - prev_close))
    )
    df["tr_ma"] = df["tr"].rolling(ATR_LOOKBACK, min_periods=ATR_LOOKBACK).mean()

    # Rolling high/low of squeeze window (for range breakout check)
    df["sq_high"] = df["High"].rolling(SQUEEZE_LOOKBACK, min_periods=SQUEEZE_LOOKBACK).max()
    df["sq_low"]  = df["Low"].rolling(SQUEEZE_LOOKBACK,  min_periods=SQUEEZE_LOOKBACK).min()

    # Count of quiet candles in squeeze window
    df["squeeze_count"] = (
        df["body_pct"]
        .lt(SQUEEZE_THRESHOLD)
        .rolling(SQUEEZE_LOOKBACK, min_periods=SQUEEZE_LOOKBACK)
        .sum()
    )

    # ST stability counter (how many consecutive candles in same direction)
    st_bull_arr = df["ST_bull"].values
    st_stable   = np.zeros(len(df), dtype=int)
    for j in range(1, len(df)):
        if st_bull_arr[j] == st_bull_arr[j-1]:
            st_stable[j] = st_stable[j-1] + 1
        else:
            st_stable[j] = 0
    df["st_stable"] = st_stable

    # Volume rolling avg
    df["vol_ma"] = df["Volume"].rolling(VOLUME_LOOKBACK, min_periods=VOLUME_LOOKBACK).mean()

    return df


# ─────────────────────────────────────────────────────────────
def fetch_data():
    validate()
    rng = f"{START} → {END}" if START and END else PERIOD
    print(f"\nFetching {TICKER} | {INTERVAL} | {rng} …")
    kw = dict(interval=INTERVAL, auto_adjust=True, progress=True)
    df = (yf.download(TICKER, start=START, end=END, **kw)
          if START and END else
          yf.download(TICKER, period=PERIOD, **kw))
    if df.empty:
        raise ValueError(f"\n  ❌  No data for '{TICKER}'.\n")
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df.index = pd.to_datetime(df.index)
    if df.index.tz is None:
        df.index = df.index.tz_localize("UTC")
    df.index = df.index.tz_convert(IST)
    df = compute_supertrend(df)
    df = compute_indicators(df)
    df.dropna(inplace=True)
    print(f"  {len(df)} candles  |  {fmt(df.index[0])} → {fmt(df.index[-1])}")
    return df


# ─────────────────────────────────────────────────────────────
def entry_allowed(i, opens, closes, highs, lows, st_bulls, body_pcts,
                  body_mas, trs, tr_mas, sq_highs, sq_lows,
                  squeeze_counts, st_stables, volumes, vol_mas,
                  trades_today, cooldown_left):
    """
    Returns (True, direction) or (False, None).
    Each gate is checked only if its True/False flag is enabled.
    """
    c_bull      = st_bulls[i]
    c_body      = body_pcts[i]
    c_body_ma   = body_mas[i]
    c_tr        = trs[i]
    c_tr_ma     = tr_mas[i]
    c_high      = highs[i]
    c_low       = lows[i]
    c_open      = opens[i]
    c_close     = closes[i]
    c_vol       = volumes[i]
    c_vol_ma    = vol_mas[i]
    sq_hi       = sq_highs[i-1]
    sq_lo       = sq_lows[i-1]
    sq_count    = squeeze_counts[i-1]
    st_stable_n = st_stables[i]
    candle_bull = c_close >= c_open

    # Gate 7 — Daily trade limit
    if GATE_DAILY_LIMIT and trades_today >= MAX_TRADES_PER_DAY:
        return False, None

    # Gate 8 — Post-loss cooldown
    if GATE_COOLDOWN and cooldown_left > 0:
        return False, None

    # Gate 1 — ST stability
    if GATE_ST_STABILITY and st_stable_n < MIN_ST_CANDLES:
        return False, None

    # Gate 2 — Squeeze zone
    if GATE_SQUEEZE:
        if np.isnan(sq_count) or sq_count < SQUEEZE_MIN_CANDLES:
            return False, None

    # Gate 3 — Power candle (relative size)
    if GATE_POWER_RELATIVE:
        if np.isnan(c_body_ma) or c_body_ma == 0:
            return False, None
        if c_body < (POWER_MULTIPLIER * c_body_ma):
            return False, None

    # Gate 3b — Power candle (absolute floor)
    if GATE_POWER_FLOOR:
        if c_body < POWER_MIN_BODY_PCT:
            return False, None

    # Gate 3c — Candle colour must match ST direction
    if GATE_CANDLE_COLOUR:
        if candle_bull != c_bull:
            return False, None

    # Gate 4 — Range breakout
    if GATE_RANGE_BREAKOUT:
        if np.isnan(sq_hi) or np.isnan(sq_lo):
            return False, None
        if c_bull     and c_high <= sq_hi:
            return False, None
        if not c_bull and c_low  >= sq_lo:
            return False, None

    # Gate 5 — ATR expansion
    if GATE_ATR_EXPANSION:
        if np.isnan(c_tr_ma) or c_tr_ma == 0:
            return False, None
        if c_tr < (ATR_EXPANSION_FACTOR * c_tr_ma):
            return False, None

    # Gate 6 — Volume surge
    if GATE_VOLUME_SURGE:
        if np.isnan(c_vol_ma) or c_vol_ma == 0:
            return False, None
        if c_vol < (VOLUME_MIN_FACTOR * c_vol_ma):
            return False, None

    direction = "long" if c_bull else "short"
    return True, direction


# ─────────────────────────────────────────────────────────────
def run_backtest(df):
    trades = []

    in_trade      = False
    direction     = None
    entry_price   = None
    entry_time    = None
    peak          = None
    hard_sl_price = None
    trail_active  = False
    trail_sl      = None
    fade_count    = 0
    candles_in    = 0       # candles since entry
    cooldown_left = 0       # candles remaining in post-loss cooldown
    trades_today  = 0
    current_day   = None

    # Pre-extract arrays for speed
    closes         = df["Close"].to_numpy(dtype=float)
    opens          = df["Open"].to_numpy(dtype=float)
    highs          = df["High"].to_numpy(dtype=float)
    lows           = df["Low"].to_numpy(dtype=float)
    st_vals        = df["ST_val"].to_numpy(dtype=float)
    st_bulls       = df["ST_bull"].to_numpy(dtype=bool)
    body_pcts      = df["body_pct"].to_numpy(dtype=float)
    body_mas       = df["body_ma"].to_numpy(dtype=float)
    trs            = df["tr"].to_numpy(dtype=float)
    tr_mas         = df["tr_ma"].to_numpy(dtype=float)
    sq_highs       = df["sq_high"].to_numpy(dtype=float)
    sq_lows        = df["sq_low"].to_numpy(dtype=float)
    squeeze_counts = df["squeeze_count"].to_numpy(dtype=float)
    st_stables     = df["st_stable"].to_numpy(dtype=int)
    volumes        = df["Volume"].to_numpy(dtype=float)
    vol_mas        = df["vol_ma"].to_numpy(dtype=float)
    times          = df.index

    def record(dir_, e_time, e_price, x_time, x_price, pk, st_sl, reason, h_sl, fc, ci):
        pnl = round(((x_price - e_price) if dir_ == "long" else (e_price - x_price))
                    / e_price * 100, 4)
        trades.append({
            "Date"            : e_time.strftime("%Y-%m-%d"),
            "Direction"       : "Long  ↑" if dir_ == "long" else "Short ↓",
            "Entry Time"      : fmt(e_time),
            "Entry Price"     : round(e_price, 2),
            "Hard SL"         : round(h_sl, 2) if h_sl else "OFF",
            "Trail SL"        : round(trail_sl, 2) if trail_sl else "—",
            "ST SL"           : round(st_sl, 2),
            "Exit Time"       : fmt(x_time),
            "Exit Price"      : round(x_price, 2),
            "Peak"            : round(pk, 2),
            "Points"          : round(abs(x_price - e_price), 2),
            "P&L %"           : pnl,
            "Candles Held"    : ci,
            "Fade Count"      : fc,
            "Exit Reason"     : reason,
            "Result"          : "WIN" if pnl > 0 else "LOSS",
        })

    def reset_trade(was_loss):
        nonlocal in_trade, direction, entry_price, entry_time, peak
        nonlocal hard_sl_price, trail_active, trail_sl, fade_count
        nonlocal candles_in, cooldown_left
        in_trade = False; direction = None; entry_price = None
        entry_time = None; peak = None; hard_sl_price = None
        trail_active = False; trail_sl = None; fade_count = 0; candles_in = 0
        if was_loss:
            cooldown_left = COOLDOWN_CANDLES

    for i in range(1, len(df)):
        c_close = closes[i]; c_open  = opens[i]
        c_high  = highs[i];  c_low   = lows[i]
        c_time  = times[i];  c_tod   = c_time.time()
        c_st    = st_vals[i]; c_bull  = st_bulls[i]
        c_body  = body_pcts[i]
        prev_bull = st_bulls[i-1]
        flipped_bull = (not prev_bull) and c_bull
        flipped_bear = prev_bull and (not c_bull)

        # ── Reset daily counter ─────────────────────────────────────────
        day_str = c_time.strftime("%Y-%m-%d")
        if day_str != current_day:
            current_day  = day_str
            trades_today = 0

        # ── Cooldown tick ───────────────────────────────────────────────
        if not in_trade and cooldown_left > 0:
            cooldown_left -= 1

        # ── EOD FORCE EXIT ──────────────────────────────────────────────
        if EOD_EXIT_ENABLED and in_trade and c_tod >= MARKET_CLOSE:
            pk_final = max(peak, c_high) if direction == "long" else min(peak, c_low)
            record(direction, entry_time, entry_price, c_time, c_close,
                   pk_final, c_st, "EOD Exit", hard_sl_price, fade_count, candles_in)
            reset_trade(was_loss=(c_close < entry_price if direction == "long"
                                  else c_close > entry_price))
            continue

        if c_tod < MARKET_OPEN or c_tod >= MARKET_CLOSE:
            continue

        # ── MANAGE OPEN TRADE ───────────────────────────────────────────
        if in_trade:
            candles_in += 1

            # Update peak
            if direction == "long":
                if c_high > peak: peak = c_high
            else:
                if c_low < peak: peak = c_low

            # ── Update trailing stop ───────────────────────────────────
            if TRAIL_ENABLED:
                if direction == "long":
                    profit_pct = (peak - entry_price) / entry_price * 100
                    if profit_pct >= TRAIL_ACTIVATION_PCT:
                        trail_active = True
                        locked_profit = profit_pct * TRAIL_LOCK_PCT
                        new_trail = entry_price * (1 + locked_profit / 100)
                        if trail_sl is None or new_trail > trail_sl:
                            trail_sl = new_trail
                else:
                    profit_pct = (entry_price - peak) / entry_price * 100
                    if profit_pct >= TRAIL_ACTIVATION_PCT:
                        trail_active = True
                        locked_profit = profit_pct * TRAIL_LOCK_PCT
                        new_trail = entry_price * (1 - locked_profit / 100)
                        if trail_sl is None or new_trail < trail_sl:
                            trail_sl = new_trail

            # ── Momentum fade tracker ──────────────────────────────────
            if MOMENTUM_FADE_ENABLED:
                if c_body < MOMENTUM_FADE_THRESHOLD:
                    fade_count += 1
                else:
                    fade_count = 0  # strong candle resets the counter

            exit_px = None; reason = None; is_loss = False

            if direction == "long":
                # Priority 1: Hard SL
                if HARD_SL_ENABLED and c_low <= hard_sl_price:
                    exit_px = hard_sl_price; reason = f"Hard SL ({HARD_SL_PCT}%)"
                    is_loss = True
                # Priority 2: Trailing SL (only fires in profit zone)
                elif trail_active and trail_sl and c_low <= trail_sl:
                    exit_px = trail_sl; reason = "Trail Stop"
                    is_loss = (trail_sl < entry_price)
                # Priority 3: Momentum fade (only after settling in)
                elif (MOMENTUM_FADE_ENABLED and candles_in >= MOMENTUM_MIN_CANDLES
                      and fade_count >= MOMENTUM_FADE_CANDLES):
                    exit_px = c_close; reason = "Momentum Fade"
                    is_loss = (c_close < entry_price)
                # Priority 4: ST stop loss
                elif ST_SL_ENABLED and c_low <= c_st:
                    exit_px = c_st; reason = "ST Stop Loss"
                    is_loss = (c_st < entry_price)
                # Priority 5: ST flip
                elif ST_FLIP_EXIT_ENABLED and flipped_bear:
                    exit_px = c_close; reason = "ST Flip Bear"
                    is_loss = (c_close < entry_price)

            else:  # short
                if HARD_SL_ENABLED and c_high >= hard_sl_price:
                    exit_px = hard_sl_price; reason = f"Hard SL ({HARD_SL_PCT}%)"
                    is_loss = True
                elif trail_active and trail_sl and c_high >= trail_sl:
                    exit_px = trail_sl; reason = "Trail Stop"
                    is_loss = (trail_sl > entry_price)
                elif (MOMENTUM_FADE_ENABLED and candles_in >= MOMENTUM_MIN_CANDLES
                      and fade_count >= MOMENTUM_FADE_CANDLES):
                    exit_px = c_close; reason = "Momentum Fade"
                    is_loss = (c_close > entry_price)
                elif ST_SL_ENABLED and c_high >= c_st:
                    exit_px = c_st; reason = "ST Stop Loss"
                    is_loss = (c_st > entry_price)
                elif ST_FLIP_EXIT_ENABLED and flipped_bull:
                    exit_px = c_close; reason = "ST Flip Bull"
                    is_loss = (c_close > entry_price)

            if exit_px is not None:
                pk_final = (max(peak, c_high) if direction == "long"
                            else min(peak, c_low))
                record(direction, entry_time, entry_price, c_time,
                       round(exit_px, 2), pk_final, c_st, reason,
                       hard_sl_price, fade_count, candles_in)
                # Auto re-entry on ST flip only if GATE_NO_AUTO_REENTRY is False
                auto_reenter = False
                if not GATE_NO_AUTO_REENTRY and not is_loss:
                    if reason == "ST Flip Bear" and not c_bull:
                        auto_reenter = True
                        new_dir = "short"
                    elif reason == "ST Flip Bull" and c_bull:
                        auto_reenter = True
                        new_dir = "long"
                reset_trade(was_loss=is_loss)
                if auto_reenter:
                    entry_price   = closes[i]
                    entry_time    = c_time
                    direction     = new_dir
                    in_trade      = True
                    candles_in    = 0
                    fade_count    = 0
                    trail_active  = False
                    trail_sl      = None
                    trades_today += 1
                    peak = c_high if new_dir == "long" else c_low
                    hard_sl_price = (
                        round(entry_price * (1 - HARD_SL_PCT / 100), 4)
                        if new_dir == "long"
                        else round(entry_price * (1 + HARD_SL_PCT / 100), 4)
                    ) if HARD_SL_ENABLED else None
            continue

        # ── NOT IN TRADE — CHECK ENTRY ──────────────────────────────────
        ok, dir_ = entry_allowed(
            i, opens, closes, highs, lows, st_bulls, body_pcts,
            body_mas, trs, tr_mas, sq_highs, sq_lows,
            squeeze_counts, st_stables, volumes, vol_mas,
            trades_today, cooldown_left
        )

        if ok:
            entry_price   = closes[i]
            entry_time    = c_time
            direction     = dir_
            in_trade      = True
            candles_in    = 0
            fade_count    = 0
            trail_active  = False
            trail_sl      = None
            trades_today += 1
            peak = c_high if dir_ == "long" else c_low
            hard_sl_price = (
                round(entry_price * (1 - HARD_SL_PCT / 100), 4) if dir_ == "long"
                else round(entry_price * (1 + HARD_SL_PCT / 100), 4)
            ) if HARD_SL_ENABLED else None

    return trades


# ─────────────────────────────────────────────────────────────
def build_daily_summary(trades):
    if not trades:
        return pd.DataFrame()
    df_t = pd.DataFrame(trades)
    rows = []
    for date, g in df_t.groupby("Date"):
        total = len(g); wins = (g["P&L %"] > 0).sum()
        wp = g.loc[g["P&L %"] > 0, "Points"].sum()
        lp = g.loc[g["P&L %"] <= 0, "Points"].sum()
        rows.append({
            "Date"       : date,
            "Trades"     : total,
            "Longs"      : g["Direction"].str.contains("Long").sum(),
            "Shorts"     : g["Direction"].str.contains("Short").sum(),
            "Wins"       : wins,
            "Losses"     : total - wins,
            "Win Rate %"      : round(wins / total * 100, 1),
            "Total P&L %"     : round(g["P&L %"].sum(), 4),
            "Best Trade %"    : round(g["P&L %"].max(), 4),
            "Worst Trade %"   : round(g["P&L %"].min(), 4),
            "Points Won"      : round(wp, 2),
            "Points Lost"     : round(lp, 2),
            "Net Points"      : round(wp - lp, 2),
            "Avg Candles Held": round(g["Candles Held"].mean(), 1),
            "Day Result"      : ("✅ Profit" if g["P&L %"].sum() > 0
                                 else "❌ Loss" if g["P&L %"].sum() < 0 else "⚖ Flat"),
        })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────
def print_results(trades):
    SEP  = "═" * 128
    DASH = "─" * 128
    print("\n" + SEP)
    print(f"  HIGH-CONVICTION Supertrend({ST_PERIOD},{ST_MULTIPLIER}) | {TICKER} | {INTERVAL}")
    print(DASH)
    print("  ENTRY GATES")
    print(f"    Gate 1  ST Stability       : {'✅ True ' if GATE_ST_STABILITY    else '❌ False'} — ST stable for {MIN_ST_CANDLES} candles required")
    print(f"    Gate 2  Squeeze Zone       : {'✅ True ' if GATE_SQUEEZE         else '❌ False'} — {SQUEEZE_MIN_CANDLES}/{SQUEEZE_LOOKBACK} candles < {SQUEEZE_THRESHOLD}% body")
    print(f"    Gate 3  Power Candle (Rel) : {'✅ True ' if GATE_POWER_RELATIVE  else '❌ False'} — body >= {POWER_MULTIPLIER}x rolling avg")
    print(f"    Gate 3b Power Candle (Abs) : {'✅ True ' if GATE_POWER_FLOOR     else '❌ False'} — body >= {POWER_MIN_BODY_PCT}% absolute floor")
    print(f"    Gate 3c Candle Colour Match: {'✅ True ' if GATE_CANDLE_COLOUR   else '❌ False'} — colour must match ST direction")
    print(f"    Gate 4  Range Breakout     : {'✅ True ' if GATE_RANGE_BREAKOUT  else '❌ False'} — must break {SQUEEZE_LOOKBACK}-candle high/low")
    print(f"    Gate 5  ATR Expansion      : {'✅ True ' if GATE_ATR_EXPANSION   else '❌ False'} — TR >= {ATR_EXPANSION_FACTOR}x ATR avg")
    print(f"    Gate 6  Volume Surge       : {'✅ True ' if GATE_VOLUME_SURGE    else '❌ False'} — volume >= {VOLUME_MIN_FACTOR}x avg")
    print(f"    Gate 7  Daily Limit        : {'✅ True ' if GATE_DAILY_LIMIT     else '❌ False'} — max {MAX_TRADES_PER_DAY} trades/day")
    print(f"    Gate 8  Post-Loss Cooldown : {'✅ True ' if GATE_COOLDOWN        else '❌ False'} — skip {COOLDOWN_CANDLES} candles after loss")
    print(f"    Gate 9  No Auto Re-entry   : {'✅ True ' if GATE_NO_AUTO_REENTRY else '❌ False'} — must re-qualify after ST flip")
    print(DASH)
    print("  EXIT RULES")
    print(f"    Hard Stop Loss             : {'✅ True ' if HARD_SL_ENABLED         else '❌ False'} — {HARD_SL_PCT}% max loss from entry")
    print(f"    Trailing Stop              : {'✅ True ' if TRAIL_ENABLED            else '❌ False'} — activates at {TRAIL_ACTIVATION_PCT}% profit, locks {int(TRAIL_LOCK_PCT*100)}%")
    print(f"    Momentum Fade Exit         : {'✅ True ' if MOMENTUM_FADE_ENABLED    else '❌ False'} — {MOMENTUM_FADE_CANDLES} weak candles < {MOMENTUM_FADE_THRESHOLD}%")
    print(f"    ST Stop Loss               : {'✅ True ' if ST_SL_ENABLED            else '❌ False'} — dynamic ST line as stop")
    print(f"    ST Flip Exit               : {'✅ True ' if ST_FLIP_EXIT_ENABLED     else '❌ False'} — exit on trend direction change")
    print(f"    EOD Force Exit             : {'✅ True ' if EOD_EXIT_ENABLED         else '❌ False'} — close all at 15:15 IST")
    print(SEP)

    if not trades:
        print("\n  ⚠  No trades found. Conditions may be too strict.")
        print("  Try: ↓ SQUEEZE_MIN_CANDLES, ↓ POWER_MULTIPLIER, ↓ VOLUME_MIN_FACTOR")
        print(SEP + "\n")
        return None, None

    df_t   = pd.DataFrame(trades)
    df_day = build_daily_summary(trades)

    print("\n  TRADE LOG")
    print(DASH)
    cols = ["Date","Direction","Entry Time","Entry Price","Hard SL","Trail SL",
            "Exit Time","Exit Price","Points","P&L %","Candles Held","Exit Reason","Result"]
    print(df_t[cols].to_string(index=False))

    print("\n\n" + SEP)
    print("  DAILY P&L")
    print(DASH)
    print(df_day.to_string(index=False))

    total = len(df_t); wins = (df_t["P&L %"] > 0).sum(); losses = total - wins
    wp = df_t.loc[df_t["P&L %"] > 0, "Points"].sum()
    lp = df_t.loc[df_t["P&L %"] <= 0, "Points"].sum()
    profit_days = (df_day["Total P&L %"] > 0).sum()
    loss_days   = (df_day["Total P&L %"] < 0).sum()

    exit_counts = df_t["Exit Reason"].value_counts()

    print("\n\n" + SEP)
    print("  SUMMARY")
    print(DASH)
    print(f"  Total Trades          : {total}  ({total/len(df_day):.1f} avg / day)")
    print(f"  Winners               : {wins}  ({wins/total*100:.1f}%)")
    print(f"  Losers                : {losses}  ({losses/total*100:.1f}%)")
    print(f"  Avg Candles Held      : {df_t['Candles Held'].mean():.1f}")
    print(DASH)
    print("  EXIT BREAKDOWN")
    for reason, count in exit_counts.items():
        print(f"    {reason:<26}: {count}")
    print(DASH)
    print(f"  Total P&L %           : {df_t['P&L %'].sum():.4f}%")
    print(f"  Avg P&L / Trade       : {df_t['P&L %'].mean():.4f}%")
    print(f"  Best Trade            : {df_t['P&L %'].max():.4f}%")
    print(f"  Worst Trade           : {df_t['P&L %'].min():.4f}%")
    print(f"  Net Points            : {wp - lp:.2f}  (won: {wp:.2f}  lost: {lp:.2f})")
    print(DASH)
    best  = df_day.loc[df_day["Total P&L %"].idxmax()]
    worst = df_day.loc[df_day["Total P&L %"].idxmin()]
    print(f"  Trading Days          : {len(df_day)}")
    print(f"  Profit Days           : {profit_days}  ({profit_days/len(df_day)*100:.1f}%)")
    print(f"  Loss Days             : {loss_days}  ({loss_days/len(df_day)*100:.1f}%)")
    print(f"  Best Day              : {best['Date']}  →  {best['Total P&L %']:.4f}%")
    print(f"  Worst Day             : {worst['Date']}  →  {worst['Total P&L %']:.4f}%")
    print(SEP + "\n")

    return df_t, df_day


# ─────────────────────────────────────────────────────────────
def export_excel(df, trades, df_trades, df_day):
    fname = f"{TICKER.replace('^','').replace('.','_')}_{INTERVAL}_hc_momentum.xlsx"

    df_raw = df[["Open","High","Low","Close","Volume",
                 "ST_val","ST_dir","body_pct","body_ma",
                 "tr","tr_ma","squeeze_count","st_stable"]].copy()
    df_raw.index = df_raw.index.strftime("%Y-%m-%d %H:%M")
    df_raw.index.name = "DateTime (IST)"
    df_raw.columns = ["Open","High","Low","Close","Volume",
                      f"ST({ST_PERIOD},{ST_MULTIPLIER})","ST Dir",
                      "Body %",f"Body MA({POWER_LOOKBACK})",
                      "True Range",f"ATR MA({ATR_LOOKBACK})",
                      f"Squeeze Cnt({SQUEEZE_LOOKBACK})",f"ST Stable Cnt"]
    df_raw = df_raw.round(4)

    if df_trades is not None and not df_trades.empty:
        total = len(df_trades); wins = (df_trades["P&L %"] > 0).sum()
        wp = df_trades.loc[df_trades["P&L %"] > 0, "Points"].sum()
        lp = df_trades.loc[df_trades["P&L %"] <= 0, "Points"].sum()
        ec = df_trades["Exit Reason"].value_counts()
        summary_rows = [
            ["STRATEGY",           "High-Conviction Supertrend Momentum"],
            ["Ticker",             TICKER], ["Interval", INTERVAL],
            ["Supertrend",         f"({ST_PERIOD},{ST_MULTIPLIER})"],
            ["",""],
            ["ENTRY GATES",        ""],
            ["ST Stability",       f"≥ {MIN_ST_CANDLES} candles same direction"],
            ["Squeeze",            f"≥ {SQUEEZE_MIN_CANDLES}/{SQUEEZE_LOOKBACK} candles < {SQUEEZE_THRESHOLD}%"],
            ["Power Candle",       f"≥ {POWER_MULTIPLIER}× avg body + ≥ {POWER_MIN_BODY_PCT}% absolute"],
            ["Range Breakout",     f"Must break {SQUEEZE_LOOKBACK}-candle high/low"],
            ["ATR Expansion",      f"TR ≥ {ATR_EXPANSION_FACTOR}× ATR avg (lookback {ATR_LOOKBACK})"],
            ["Volume Surge",       f"Vol ≥ {VOLUME_MIN_FACTOR}× avg (lookback {VOLUME_LOOKBACK})"],
            ["Max Trades/Day",     MAX_TRADES_PER_DAY],
            ["Post-Loss Cooldown", f"{COOLDOWN_CANDLES} candles"],
            ["",""],
            ["EXIT CONFIG",        ""],
            ["Hard SL",            f"{'ON ' + str(HARD_SL_PCT) + '%' if HARD_SL_ENABLED else 'OFF'}"],
            ["Trailing Stop",      f"{'ON — activate at ' + str(TRAIL_ACTIVATION_PCT) + '%, lock ' + str(TRAIL_LOCK_PCT*100) + '%' if TRAIL_ENABLED else 'OFF'}"],
            ["Momentum Fade",      f"{'ON — ' + str(MOMENTUM_FADE_CANDLES) + ' candles < ' + str(MOMENTUM_FADE_THRESHOLD) + '%' if MOMENTUM_FADE_ENABLED else 'OFF'}"],
            ["",""],
            ["RESULTS",            ""],
            ["Total Trades",       total],
            ["Winners",            f"{wins} ({wins/total*100:.1f}%)"],
            ["Losers",             f"{total-wins} ({(total-wins)/total*100:.1f}%)"],
            ["Avg Candles Held",   round(df_trades["Candles Held"].mean(), 1)],
        ]
        for r, c in ec.items():
            summary_rows.append([f"  Exit: {r}", c])
        summary_rows += [
            ["",""],
            ["Total P&L %",        round(df_trades["P&L %"].sum(), 4)],
            ["Avg P&L / Trade",    round(df_trades["P&L %"].mean(), 4)],
            ["Best Trade %",       round(df_trades["P&L %"].max(), 4)],
            ["Worst Trade %",      round(df_trades["P&L %"].min(), 4)],
            ["Net Points",         round(wp - lp, 2)],
        ]
        if df_day is not None and not df_day.empty:
            bd = df_day.loc[df_day["Total P&L %"].idxmax()]
            wd = df_day.loc[df_day["Total P&L %"].idxmin()]
            summary_rows += [
                ["",""],
                ["Trading Days",   len(df_day)],
                ["Profit Days",    int((df_day["Total P&L %"] > 0).sum())],
                ["Loss Days",      int((df_day["Total P&L %"] < 0).sum())],
                ["Best Day",       f"{bd['Date']} → {bd['Total P&L %']:.4f}%"],
                ["Worst Day",      f"{wd['Date']} → {wd['Total P&L %']:.4f}%"],
            ]
        df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value"])
    else:
        df_summary = pd.DataFrame([{"Metric": "No trades", "Value": ""}])

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        df_raw.to_excel(writer,     sheet_name="Data",      index=True)
        if df_trades is not None and not df_trades.empty:
            df_trades.to_excel(writer, sheet_name="Trades", index=False)
        if df_day is not None and not df_day.empty:
            df_day.to_excel(writer,    sheet_name="Daily",  index=False)
        df_summary.to_excel(writer, sheet_name="Summary",   index=False)

    _style_excel(fname, df_trades, df_day)
    print(f"  Excel → {fname}")
    return fname


def _style_excel(fname, df_trades, df_day):
    GL = "C6EFCE"; GD = "1A5C38"; RL = "FFC7CE"; RD = "9C0006"
    BH = "1F3864"; YS = "FFD700"; GR = "F2F2F2"; WH = "FFFFFF"
    OR = "FCE4D6"; TL = "DDEBF7"; CY = "D6F4F4"
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws):
        for c in ws[1]:
            c.fill = PatternFill("solid", fgColor=BH)
            c.font = Font(bold=True, color=WH, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr
        ws.row_dimensions[1].height = 22

    def aw(ws, cap=28):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+3, cap)

    wb = load_workbook(fname)

    ws = wb["Data"]; hdr(ws); ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        d = str(row[6].value or "")
        bg = GL if d == "Bullish" else (RL if d == "Bearish" else WH)
        for c in row:
            c.fill = PatternFill("solid", fgColor=bg); c.border = bdr
            c.alignment = Alignment(horizontal="center")
    aw(ws)

    if "Trades" in wb.sheetnames and df_trades is not None:
        ws = wb["Trades"]; hdr(ws); ws.freeze_panes = "A2"
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            res = str(row[-1].value or ""); rea = str(row[-3].value or "")
            bg = GL if res == "WIN" else (RL if res == "LOSS" else (GR if i%2==0 else WH))
            if res == "WIN":   row[-1].font = Font(bold=True, color=GD)
            elif res == "LOSS": row[-1].font = Font(bold=True, color=RD)
            if rea.startswith("Hard SL"):
                row[-3].fill = PatternFill("solid", fgColor="FF0000")
                row[-3].font = Font(bold=True, color=WH)
            elif rea == "Trail Stop":
                row[-3].fill = PatternFill("solid", fgColor="92D050")
                row[-3].font = Font(bold=True, color="375623")
            elif rea == "Momentum Fade":
                row[-3].fill = PatternFill("solid", fgColor=CY)
                row[-3].font = Font(bold=True, color="1a7a7a")
            elif "Stop Loss" in rea:
                row[-3].fill = PatternFill("solid", fgColor=OR)
                row[-3].font = Font(bold=True, color="C55A11")
            elif "EOD" in rea:
                row[-3].fill = PatternFill("solid", fgColor=TL)
                row[-3].font = Font(bold=True, color="2E75B6")
            for c in row:
                if not c.fill or c.fill.fgColor.rgb in ("00000000","FFFFFFFF"):
                    c.fill = PatternFill("solid", fgColor=bg)
                c.border = bdr; c.alignment = Alignment(horizontal="center")
        aw(ws)

    if "Daily" in wb.sheetnames and df_day is not None:
        ws = wb["Daily"]; hdr(ws); ws.freeze_panes = "A2"
        rc = ws.max_column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            dr = str(row[rc-1].value or "")
            bg = GL if "Profit" in dr else (RL if "Loss" in dr else GR)
            if "Profit" in dr: row[rc-1].font = Font(bold=True, color=GD)
            elif "Loss" in dr: row[rc-1].font = Font(bold=True, color=RD)
            for c in row:
                c.fill = PatternFill("solid", fgColor=bg); c.border = bdr
                c.alignment = Alignment(horizontal="center")
        aw(ws)

    ws = wb["Summary"]; hdr(ws)
    SECS = {"ENTRY GATES","EXIT CONFIG","RESULTS","STRATEGY"}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        lbl = str(row[0].value or "")
        if lbl in SECS:
            for c in row:
                c.fill = PatternFill("solid", fgColor=YS)
                c.font = Font(bold=True, color="000000"); c.border = bdr
            continue
        for c in row:
            c.border = bdr; c.alignment = Alignment(horizontal="left")
        if len(row) > 1 and "Net Points" in lbl:
            v = row[1].value
            bg = GL if isinstance(v,(int,float)) and v > 0 else RL
            fg = GD if isinstance(v,(int,float)) and v > 0 else RD
            row[1].fill = PatternFill("solid", fgColor=bg)
            row[1].font = Font(bold=True, color=fg)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 36
    wb.save(fname)


# ─────────────────────────────────────────────────────────────
def build_chart(df, trades):
    if CHART_DAYS:
        ud = sorted(df.index.normalize().unique())
        co = ud[-CHART_DAYS] if len(ud) >= CHART_DAYS else ud[0]
        df_c = df[df.index.normalize() >= co].copy()
    else:
        df_c = df.copy()

    st_green = df_c["ST_val"].where(df_c["ST_bull"])
    st_red   = df_c["ST_val"].where(~df_c["ST_bull"])
    squeeze_mask = df_c["squeeze_count"] >= SQUEEZE_MIN_CANDLES

    fig = make_subplots(
        rows=2, cols=1, shared_xaxes=True,
        row_heights=[0.75, 0.25], vertical_spacing=0.02
    )

    # Candlestick
    fig.add_trace(go.Candlestick(
        x=df_c.index, open=df_c["Open"], high=df_c["High"],
        low=df_c["Low"],  close=df_c["Close"], name="Price",
        increasing_line_color="#26a69a", decreasing_line_color="#ef5350"
    ), row=1, col=1)

    # Squeeze zone markers (yellow triangles at bottom)
    sq_y = df_c["Low"].where(squeeze_mask) * 0.9995
    fig.add_trace(go.Scatter(
        x=df_c.index[squeeze_mask], y=sq_y[squeeze_mask],
        mode="markers", name="Squeeze Zone",
        marker=dict(symbol="triangle-down", size=5, color="rgba(255,215,0,0.5)")
    ), row=1, col=1)

    fig.add_trace(go.Scatter(
        x=df_c.index, y=st_green, name="ST Bullish",
        mode="lines", line=dict(color="#22c55e", width=2.5), connectgaps=False
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=df_c.index, y=st_red, name="ST Bearish",
        mode="lines", line=dict(color="#ef4444", width=2.5), connectgaps=False
    ), row=1, col=1)

    IST_tz = pytz.timezone("Asia/Kolkata")
    if trades:
        df_t = pd.DataFrame(trades)
        t0   = df_c.index[0]

        def parse(col):
            return pd.to_datetime(df_t[col], format="%Y-%m-%d %H:%M").dt.tz_localize(IST_tz)

        et = parse("Entry Time"); xt = parse("Exit Time")
        ep = df_t["Entry Price"]; xp = df_t["Exit Price"]
        dr = df_t["Direction"];   rs = df_t["Result"]
        ex = df_t["Exit Reason"]
        m  = et >= t0

        for dir_, sym, col in [("Long", "triangle-up", "#22c55e"),
                                ("Short","triangle-down","#ef4444")]:
            mm = m & dr.str.contains(dir_)
            if mm.any():
                fig.add_trace(go.Scatter(
                    x=et[mm].tolist(), y=ep[mm].tolist(),
                    mode="markers", name=f"{dir_} Entry",
                    marker=dict(symbol=sym, size=15, color=col,
                                line=dict(color="white", width=1.5))
                ), row=1, col=1)

        for cond, sym, col, brd, lbl in [
            (rs=="WIN",                   "circle",  "#86efac","#16a34a","WIN Exit"),
            (ex=="Trail Stop",            "star",    "#00ff88","#007a40","Trail Stop"),
            (ex=="Momentum Fade",         "star",    "#00e5ff","#007a8a","Momentum Fade"),
            (ex=="ST Stop Loss",          "diamond", "#fb923c","#c2410c","ST Stop Loss"),
            (ex.str.startswith("Hard SL"),"hexagram","#ff0000","white",  f"Hard SL"),
            (ex.str.contains("Flip"),     "square",  "#a78bfa","#5b21b6","ST Flip"),
            (ex=="EOD Exit",              "circle",  "#94a3b8","#475569","EOD Exit"),
        ]:
            mm = m & cond & (rs == "LOSS") if lbl in ("ST Stop Loss","Hard SL") else m & cond
            if mm.any():
                fig.add_trace(go.Scatter(
                    x=xt[mm].tolist(), y=xp[mm].tolist(),
                    mode="markers", name=lbl,
                    marker=dict(symbol=sym, size=11, color=col,
                                line=dict(color=brd, width=1.5))
                ), row=1, col=1)

    # Body % panel
    fig.add_trace(go.Bar(
        x=df_c.index,
        y=df_c["body_pct"],
        name="Body %",
        marker_color=np.where(df_c["body_pct"] >= POWER_MIN_BODY_PCT,
                               "#22c55e", "#facc15")
    ), row=2, col=1)
    fig.add_trace(go.Scatter(
        x=df_c.index, y=df_c["body_ma"],
        name=f"Body MA({POWER_LOOKBACK})",
        mode="lines", line=dict(color="#60a5fa", width=1.5)
    ), row=2, col=1)
    fig.add_hline(y=SQUEEZE_THRESHOLD, line_dash="dash", line_color="#fbbf24",
                  annotation_text=f"Squeeze ({SQUEEZE_THRESHOLD}%)",
                  annotation_position="bottom right", row=2, col=1)
    fig.add_hline(y=POWER_MIN_BODY_PCT, line_dash="dot", line_color="#22c55e",
                  annotation_text=f"Power floor ({POWER_MIN_BODY_PCT}%)",
                  annotation_position="top right", row=2, col=1)

    fig.update_layout(
        template="plotly_dark", height=900,
        xaxis_rangeslider_visible=False,
        legend=dict(orientation="h", yanchor="bottom", y=1.01,
                    xanchor="right", x=1, font=dict(size=9)),
        margin=dict(l=60, r=40, t=120, b=50),
        font=dict(family="monospace", size=11),
        title=dict(
            text=(f"<b>{TICKER}</b> | {INTERVAL} | High-Conviction Supertrend({ST_PERIOD},{ST_MULTIPLIER})<br>"
                  f"<span style='font-size:10px;color:#94a3b8'>"
                  f"🟡 Squeeze zone  →  💥 Power breakout  →  Entry | "
                  f"Max {MAX_TRADES_PER_DAY}/day | "
                  f"Trail SL protects profit | Momentum Fade exits chopping moves</span>"),
            x=0.5, xanchor="center"
        )
    )
    fig.update_yaxes(title_text="Price",  showgrid=True, gridcolor="#1e293b", row=1, col=1)
    fig.update_yaxes(title_text="Body %", showgrid=True, gridcolor="#1e293b", row=2, col=1)
    fig.update_xaxes(showgrid=True, gridcolor="#1e293b",
                     rangebreaks=[dict(bounds=["sat","mon"]),
                                  dict(bounds=[15.5, 9.25], pattern="hour")])
    fname = f"{TICKER.replace('^','').replace('.','_')}_{INTERVAL}_hc_chart.html"
    fig.write_html(fname)
    print(f"  Chart → {fname}")


# ─────────────────────────────────────────────────────────────
def print_tuning_guide():
    print("""
  ┌──────────────────────────────────────────────────────────────────────┐
  │  QUICK TUNING GUIDE                                                  │
  ├──────────────────────────────────────────────────────────────────────┤
  │  Zero trades? (too strict)                                           │
  │    ↓ SQUEEZE_MIN_CANDLES   (e.g. 4 instead of 5)                    │
  │    ↓ POWER_MULTIPLIER      (e.g. 2.0 instead of 2.5)               │
  │    ↓ VOLUME_MIN_FACTOR     (e.g. 1.2 instead of 1.4)               │
  │    ↓ ATR_EXPANSION_FACTOR  (e.g. 1.1 instead of 1.2)               │
  │    ↑ MAX_TRADES_PER_DAY    (allow more per day)                     │
  │                                                                      │
  │  Too many losers? (setups not clean enough)                         │
  │    ↑ POWER_MULTIPLIER      (demand stronger breakout)               │
  │    ↑ SQUEEZE_MIN_CANDLES   (need more coiling)                      │
  │    ↑ VOLUME_MIN_FACTOR     (demand more volume conviction)          │
  │    ↓ MAX_TRADES_PER_DAY    (1 = ultra-selective, best setup only)   │
  │    ↑ COOLDOWN_CANDLES      (more rest after loss)                   │
  │                                                                      │
  │  Exiting profitable trades too early?                               │
  │    ↑ MOMENTUM_FADE_CANDLES  (more patient before fade exit)         │
  │    ↓ TRAIL_LOCK_PCT         (looser trail, lets winner breathe)     │
  │    ↑ TRAIL_ACTIVATION_PCT   (trail kicks in later)                  │
  │                                                                      │
  │  Giving back profits at end of move?                                │
  │    ↓ MOMENTUM_FADE_CANDLES  (exit sooner on fade, e.g. 3)          │
  │    ↑ TRAIL_LOCK_PCT         (lock more profit, e.g. 0.6–0.7)       │
  │                                                                      │
  │  Recommended defaults for NSE Nifty 5m:                            │
  │    SQUEEZE_THRESHOLD   0.15–0.20  |  POWER_MULTIPLIER   2.0–3.0   │
  │    VOLUME_MIN_FACTOR   1.3–1.5    |  MAX_TRADES_PER_DAY  1–2      │
  └──────────────────────────────────────────────────────────────────────┘
""")


# ─────────────────────────────────────────────────────────────
def main():
    try:
        df = fetch_data()
    except ValueError as e:
        print(e); return

    trades            = run_backtest(df)
    df_trades, df_day = print_results(trades)

    print_tuning_guide()

    print("  Exporting Excel …")
    export_excel(df, trades, df_trades, df_day)

    print("  Building chart …")
    build_chart(df, trades)

    print("\n  ✅  Done.\n")


if __name__ == "__main__":
    main()