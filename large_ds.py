"""
Supertrend Intraday Backtesting Strategy — v2.0
─────────────────────────────────────────────────────────────────────────────
DATA SOURCE (choose one):
  DATA_SOURCE = "yfinance"  → download from Yahoo Finance (live)
  DATA_SOURCE = "csv"       → read from a local CSV file (CSV_PATH)

CSV FORMAT (any of these column naming styles are auto-detected):
  Required : datetime, open, high, low, close  (case-insensitive)
  Optional : volume
  DateTime : any standard format (auto-parsed), with or without timezone

FILTERS ADDED IN v2.0:
  1. Higher-Timeframe Filter  — only trade in direction of 5m ST (yfinance mode)
                                or HTF ST computed on resampled CSV data
  2. Trade Window Filter      — only enter trades in profitable time slots
  3. Confirmation Candle      — wait N candles before entering after an ST flip
  4. Min Gap Filter           — skip entries when price is too close to ST line
  5. Volume Filter            — require volume > multiplier × 20-bar average
  6. Max Trades Per Day Cap   — hard cap on daily trade count
  7. Time-Based P&L Analysis  — shows P&L breakdown per 30-min slot

LONG ENTRY  : ST is BELOW close price  (ST bull → price > ST line)
LONG EXIT   : Low ≤ ST line  OR  Hard SL fires first  OR  EOD

SHORT ENTRY : ST is ABOVE close price  (ST bear → price < ST line)
SHORT EXIT  : High ≥ ST line  OR  Hard SL fires first  OR  EOD

Install: pip install yfinance pandas pandas-ta plotly openpyxl pytz numpy
"""

import yfinance as yf
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import pytz
from datetime import datetime, timezone, time as dtime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

# ── Data Source ────────────────────────────────────────────────────────────────
DATA_SOURCE   = "csv"          # "yfinance" | "csv"
CSV_PATH      = "bt.csv"     # Path to your CSV file (used only when DATA_SOURCE="csv")
#
# CSV column names are auto-detected (case-insensitive). Accepted names:
#   datetime : "datetime", "date", "time", "timestamp", "Date", "Datetime" etc.
#   open     : "open", "Open", "OPEN"
#   high     : "high", "High", "HIGH"
#   low      : "low", "Low", "LOW"
#   close    : "close", "Close", "CLOSE", "ltp", "LTP"
#   volume   : "volume", "Volume", "VOLUME", "vol"  (optional)

# ── yfinance settings (ignored when DATA_SOURCE="csv") ────────────────────────
TICKER        = "^NSEI"             # ^NSEI | RELIANCE.NS | TCS.NS etc.
INTERVAL      = "1m"                # 1m / 5m / 15m (intraday, last 60d only)
PERIOD        = "7d"                # used when START=None, END=None
START         = None                # e.g. "2026-01-25"
END           = None                # e.g. "2026-02-23"

# ── Supertrend settings ────────────────────────────────────────────────────────
ST_PERIOD     = 14                  # Increased from 10 → fewer flips
ST_MULTIPLIER = 3.5                 # Increased from 3.0 → smoother signals

# ── Hard Stop Loss ─────────────────────────────────────────────────────────────
HARD_SL_ENABLED = True
HARD_SL_PCT     = 0.25              # Slightly wider (0.30%) for 1m noise

# ── Filter 1: Higher-Timeframe Confirmation ────────────────────────────────────
HTF_FILTER_ENABLED = False           # Only trade in direction of HTF Supertrend
HTF_RESAMPLE       = "5min"         # Resample base data to this TF for HTF ST
                                     # Use "5min","15min","30min" etc.
HTF_ST_PERIOD      = 10
HTF_ST_MULTIPLIER  = 3.0

# ── Filter 2: Trade Window (time-of-day filter) ────────────────────────────────
TRADE_WINDOW_ENABLED = False
TRADE_WINDOWS = [                    # Only enter trades within these windows
    (dtime(9, 30), dtime(11, 30)),   # Morning momentum (best window)
    (dtime(14, 0), dtime(15, 10)),   # Pre-close momentum
]
# Set TRADE_WINDOW_ENABLED = False to trade all day

# ── Filter 3: Confirmation Candle ──────────────────────────────────────────────
CONFIRM_CANDLES = 10                  # Wait N candles after ST flip before entering
                                     # Set to 1 to disable (original behaviour)

# ── Filter 4: Min Gap Filter (anti-chop) ──────────────────────────────────────
MIN_GAP_ENABLED = True
MIN_GAP_PCT     = 0.15               # Skip entry if price is within 0.15% of ST line

# ── Filter 5: Volume Filter ────────────────────────────────────────────────────
VOLUME_FILTER_ENABLED = False        # Disable if CSV has no volume column
VOLUME_MULTIPLIER     = 1.5          # Volume must be > 1.5× 20-bar average
VOLUME_LOOKBACK       = 20

# ── Filter 6: Max Trades Per Day ───────────────────────────────────────────────
MAX_TRADES_PER_DAY_ENABLED = True
MAX_TRADES_PER_DAY         = 5

# ── Intraday session (NSE) ─────────────────────────────────────────────────────
MARKET_OPEN   = dtime(9, 15)
MARKET_CLOSE  = dtime(15, 15)
IST           = pytz.timezone("Asia/Kolkata")

# ── Chart ──────────────────────────────────────────────────────────────────────
CHART_DAYS = 5                       # How many recent days to show (None = all)

# ═══════════════════════════════════════════════════════════════════════════════

MAX_DAYS = {"1m": 7, "2m": 60, "5m": 60, "15m": 60,
            "30m": 60, "60m": 60, "90m": 60, "1h": 60}


def fmt(ts):
    return ts.strftime("%Y-%m-%d %H:%M") if pd.notna(ts) else "—"


# ───────────────────────────────────────────────────────────────────────────────
#  SUPERTREND ENGINE
# ───────────────────────────────────────────────────────────────────────────────
def compute_supertrend(df, period=None, multiplier=None):
    """Wilder ATR-based Supertrend. Returns df with ST_val, ST_bull, ST_direction."""
    p = period or ST_PERIOD
    m = multiplier or ST_MULTIPLIER

    high  = df["High"].values.astype(float)
    low   = df["Low"].values.astype(float)
    close = df["Close"].values.astype(float)
    n     = len(df)

    tr = np.maximum(high[1:] - low[1:],
         np.maximum(np.abs(high[1:] - close[:-1]),
                    np.abs(low[1:]  - close[:-1])))

    atr = np.full(n, np.nan)
    if n > p:
        atr[p] = tr[:p].mean()
        for j in range(p + 1, n):
            atr[j] = (atr[j-1] * (p - 1) + tr[j-1]) / p

    hl2         = (high + low) / 2.0
    upper_basic = hl2 + m * atr
    lower_basic = hl2 - m * atr

    upper = np.full(n, np.nan)
    lower = np.full(n, np.nan)
    bull  = np.full(n, True, dtype=bool)

    for j in range(p, n):
        if np.isnan(upper[j-1]):
            upper[j] = upper_basic[j]
            lower[j] = lower_basic[j]
            bull[j]  = close[j] >= lower[j]
            continue
        upper[j] = (upper_basic[j]
                    if upper_basic[j] < upper[j-1] or close[j-1] > upper[j-1]
                    else upper[j-1])
        lower[j] = (lower_basic[j]
                    if lower_basic[j] > lower[j-1] or close[j-1] < lower[j-1]
                    else lower[j-1])
        bull[j] = (close[j] >= lower[j]) if bull[j-1] else (close[j] > upper[j])

    df = df.copy()
    df["ST_val"]       = np.where(bull, lower, upper)
    df["ST_bull"]      = bull
    df["ST_direction"] = np.where(bull, "Bullish", "Bearish")
    return df


# ───────────────────────────────────────────────────────────────────────────────
#  CSV LOADER
# ───────────────────────────────────────────────────────────────────────────────
def _detect_column(columns, candidates):
    """Find the first matching column name (case-insensitive)."""
    col_map = {c.lower(): c for c in columns}
    for cand in candidates:
        if cand.lower() in col_map:
            return col_map[cand.lower()]
    return None


def load_csv(path):
    """
    Load OHLC data from a CSV file.
    Auto-detects column names, parses datetime, localizes to IST.
    """
    print(f"\nReading CSV: {path} …")
    df = pd.read_csv(path)
    print(f"  Raw shape: {df.shape}  |  Columns: {list(df.columns)}")

    # ── Detect columns ─────────────────────────────────────────────────────────
    dt_col  = _detect_column(df.columns, ["datetime","date","time","timestamp","Date","Datetime","TIME","DATETIME","DATE"])
    o_col   = _detect_column(df.columns, ["open","Open","OPEN","o"])
    h_col   = _detect_column(df.columns, ["high","High","HIGH","h"])
    l_col   = _detect_column(df.columns, ["low","Low","LOW","l"])
    c_col   = _detect_column(df.columns, ["close","Close","CLOSE","ltp","LTP","c"])
    v_col   = _detect_column(df.columns, ["volume","Volume","VOLUME","vol","Vol","VOL"])

    missing = [name for name, col in [("datetime",dt_col),("open",o_col),
                                       ("high",h_col),("low",l_col),("close",c_col)]
               if col is None]
    if missing:
        raise ValueError(
            f"\n  ❌  Could not detect columns: {missing}\n"
            f"      Found columns: {list(df.columns)}\n"
            f"      Rename your columns to: datetime, open, high, low, close, volume\n"
        )

    print(f"  Detected → datetime='{dt_col}' | open='{o_col}' | high='{h_col}' | low='{l_col}' | close='{c_col}'"
          + (f" | volume='{v_col}'" if v_col else " | volume=NOT FOUND"))

    # ── Build clean DataFrame ──────────────────────────────────────────────────
    rename = {dt_col: "__dt", o_col: "Open", h_col: "High", l_col: "Low", c_col: "Close"}
    if v_col:
        rename[v_col] = "Volume"

    df = df[list(rename.keys())].rename(columns=rename)

    # ── Parse datetime ─────────────────────────────────────────────────────────
    df["__dt"] = pd.to_datetime(
    df["__dt"],
    format="%d-%m-%Y %H:%M",
    errors="coerce"
)
    df = df.set_index("__dt")
    df.index.name = "Datetime"

    if df.index.tz is None:
        df.index = df.index.tz_localize(IST, ambiguous="infer", nonexistent="shift_forward")
    else:
        df.index = df.index.tz_convert(IST)

    df.sort_index(inplace=True)

    # ── Numeric conversion ─────────────────────────────────────────────────────
    for col in ["Open", "High", "Low", "Close"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    if "Volume" in df.columns:
        df["Volume"] = pd.to_numeric(df["Volume"], errors="coerce").fillna(0)
    else:
        df["Volume"] = 0

    df.dropna(subset=["Open","High","Low","Close"], inplace=True)

    # ── Market hours filter ────────────────────────────────────────────────────
    tod = df.index.time
    df  = df[(tod >= MARKET_OPEN) & (tod <= MARKET_CLOSE)]

    df  = compute_supertrend(df)
    df.dropna(inplace=True)

    print(f"  {len(df)} candles  |  {fmt(df.index[0])} → {fmt(df.index[-1])}")
    return df


# ───────────────────────────────────────────────────────────────────────────────
#  YFINANCE LOADER
# ───────────────────────────────────────────────────────────────────────────────
def validate_yf():
    if INTERVAL not in MAX_DAYS:
        return
    max_days = MAX_DAYS[INTERVAL]
    if START:
        try:
            start_dt = datetime.strptime(START, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise ValueError(f"START='{START}' must be YYYY-MM-DD format.")
        days_ago = (datetime.now(timezone.utc) - start_dt).days
        if days_ago > max_days:
            raise ValueError(
                f"\n  ❌  '{INTERVAL}' only covers the last {max_days} days.\n"
                f"      START='{START}' is {days_ago} days ago.\n"
                f"  ✅  Use DATA_SOURCE='csv' for historical data.\n"
            )


def load_yfinance():
    validate_yf()
    rng = f"{START} → {END}" if START and END else PERIOD
    print(f"\nFetching {TICKER} | {INTERVAL} | {rng} …")
    kw = dict(interval=INTERVAL, auto_adjust=True, progress=True)
    df = (yf.download(TICKER, start=START, end=END, **kw)
          if START and END else
          yf.download(TICKER, period=PERIOD, **kw))

    if df.empty:
        raise ValueError(f"\n  ❌  No data for '{TICKER}'.\n")

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    df.index = pd.to_datetime(df.index)
    if df.index.tz is None:
        df.index = df.index.tz_localize("UTC")
    df.index = df.index.tz_convert(IST)

    df = compute_supertrend(df)
    df.dropna(inplace=True)
    print(f"  {len(df)} candles  |  {fmt(df.index[0])} → {fmt(df.index[-1])}")
    return df


# ───────────────────────────────────────────────────────────────────────────────
#  HTF SUPERTREND (Higher Timeframe)
# ───────────────────────────────────────────────────────────────────────────────
def compute_htf_st(df):
    """
    Resample base data to HTF_RESAMPLE, compute ST, then forward-fill
    back to original frequency. Returns a boolean Series aligned to df.index.
    """
    df_htf = df[["Open","High","Low","Close","Volume"]].resample(HTF_RESAMPLE).agg({
        "Open"  : "first",
        "High"  : "max",
        "Low"   : "min",
        "Close" : "last",
        "Volume": "sum",
    }).dropna()

    df_htf = compute_supertrend(df_htf, period=HTF_ST_PERIOD, multiplier=HTF_ST_MULTIPLIER)
    df_htf.dropna(inplace=True)

    # Forward-fill HTF ST_bull back to original index
    htf_bull = df_htf["ST_bull"].reindex(df.index, method="ffill")
    return htf_bull


# ───────────────────────────────────────────────────────────────────────────────
#  MAIN FETCH / LOAD
# ───────────────────────────────────────────────────────────────────────────────
def fetch_data():
    if DATA_SOURCE.lower() == "csv":
        df = load_csv(CSV_PATH)
        ticker_label = CSV_PATH.split("/")[-1].replace(".csv","")
    else:
        df = load_yfinance()
        ticker_label = TICKER

    # Compute HTF ST if filter enabled
    htf_bull = None
    if HTF_FILTER_ENABLED:
        print(f"  Computing {HTF_RESAMPLE} HTF Supertrend({HTF_ST_PERIOD},{HTF_ST_MULTIPLIER}) …")
        htf_bull = compute_htf_st(df)
        df["HTF_bull"] = htf_bull
        df["HTF_bull"].ffill(inplace=True)
        df.dropna(subset=["HTF_bull"], inplace=True)
        print(f"  HTF filter applied. Remaining candles: {len(df)}")

    return df, ticker_label


# ───────────────────────────────────────────────────────────────────────────────
#  BACKTEST ENGINE
# ───────────────────────────────────────────────────────────────────────────────
def run_backtest(df):
    trades = []

    in_trade      = False
    direction     = None
    entry_price   = None
    entry_time    = None
    peak          = None
    hard_sl_price = None
    flip_candle   = -999   # index of last ST flip candle

    closes   = df["Close"].to_numpy(dtype=float)
    highs    = df["High"].to_numpy(dtype=float)
    lows     = df["Low"].to_numpy(dtype=float)
    st_vals  = df["ST_val"].to_numpy(dtype=float)
    st_bulls = df["ST_bull"].to_numpy(dtype=bool)
    volumes  = df["Volume"].to_numpy(dtype=float) if "Volume" in df.columns else np.ones(len(df))
    htf_bulls = df["HTF_bull"].to_numpy(dtype=bool) if "HTF_bull" in df.columns else np.ones(len(df), dtype=bool)
    times    = df.index

    # Rolling volume average
    vol_series = pd.Series(volumes)
    vol_avg    = vol_series.rolling(VOLUME_LOOKBACK).mean().to_numpy()

    # Daily trade counter
    daily_trades = {}

    def in_trade_window(t):
        if not TRADE_WINDOW_ENABLED:
            return True
        return any(s <= t <= e for s, e in TRADE_WINDOWS)

    def record(dir_, e_time, e_price, x_time, x_price, pk, sl_at_exit, reason, hard_sl):
        pnl = round(((x_price - e_price) / e_price * 100)
                    if dir_ == "long"
                    else ((e_price - x_price) / e_price * 100), 4)
        pts = round(abs(x_price - e_price), 4)
        trades.append({
            "Date"            : e_time.strftime("%Y-%m-%d"),
            "Direction"       : "Long  ↑" if dir_ == "long" else "Short ↓",
            "Entry Time"      : fmt(e_time),
            "Entry Price"     : round(e_price, 2),
            "Hard SL Price"   : round(hard_sl, 2) if hard_sl is not None else "OFF",
            "ST Stop Loss"    : round(sl_at_exit, 2),
            "Exit Time"       : fmt(x_time),
            "Exit Price"      : round(x_price, 2),
            "Peak"            : round(pk, 2),
            "Points Captured" : pts,
            "P&L %"           : pnl,
            "Exit Reason"     : reason,
            "Result"          : "WIN" if pnl > 0 else "LOSS",
        })

    for i in range(1, len(df)):
        c_close    = closes[i]
        c_high     = highs[i]
        c_low      = lows[i]
        c_time     = times[i]
        c_tod      = c_time.time()
        c_st       = st_vals[i]
        c_bull     = st_bulls[i]
        prev_bull  = st_bulls[i-1]
        c_htf_bull = htf_bulls[i]
        date_key   = c_time.date()

        flipped_bull = (not prev_bull) and c_bull
        flipped_bear = prev_bull and (not c_bull)
        if flipped_bull or flipped_bear:
            flip_candle = i

        # ── EOD FORCE EXIT ─────────────────────────────────────────────────────
        if in_trade and c_tod >= MARKET_CLOSE:
            pk_final = max(peak, c_high) if direction == "long" else min(peak, c_low)
            record(direction, entry_time, entry_price,
                   c_time, c_close, pk_final, c_st, "EOD Exit", hard_sl_price)
            in_trade = False
            continue

        # Skip outside market hours
        if c_tod < MARKET_OPEN or c_tod >= MARKET_CLOSE:
            continue

        # ── MANAGE OPEN TRADE ──────────────────────────────────────────────────
        if in_trade:
            if direction == "long":
                if c_high > peak: peak = c_high
                hard_sl_hit = (HARD_SL_ENABLED and hard_sl_price is not None and c_low <= hard_sl_price)
                st_sl_hit   = c_low <= c_st or flipped_bear
                if hard_sl_hit or st_sl_hit:
                    exit_px = round(hard_sl_price, 2) if hard_sl_hit else round(c_st, 2)
                    reason  = (f"Hard SL ({HARD_SL_PCT}%)" if hard_sl_hit
                               else ("ST Stop Loss" if c_low <= c_st else "ST Flip Bear"))
                    record("long", entry_time, entry_price, c_time, exit_px, peak, c_st, reason, hard_sl_price)
                    in_trade = False
                    if not hard_sl_hit and not c_bull:
                        # Re-entry short only if HTF agrees and window OK
                        if (not HTF_FILTER_ENABLED or not c_htf_bull) and in_trade_window(c_tod):
                            cnt = daily_trades.get(date_key, 0)
                            if not MAX_TRADES_PER_DAY_ENABLED or cnt < MAX_TRADES_PER_DAY:
                                entry_price   = c_close; entry_time = c_time
                                peak          = c_low;   direction  = "short"; in_trade = True
                                hard_sl_price = round(entry_price * (1 + HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
                                daily_trades[date_key] = cnt + 1
            elif direction == "short":
                if c_low < peak: peak = c_low
                hard_sl_hit = (HARD_SL_ENABLED and hard_sl_price is not None and c_high >= hard_sl_price)
                st_sl_hit   = c_high >= c_st or flipped_bull
                if hard_sl_hit or st_sl_hit:
                    exit_px = round(hard_sl_price, 2) if hard_sl_hit else round(c_st, 2)
                    reason  = (f"Hard SL ({HARD_SL_PCT}%)" if hard_sl_hit
                               else ("ST Stop Loss" if c_high >= c_st else "ST Flip Bull"))
                    record("short", entry_time, entry_price, c_time, exit_px, peak, c_st, reason, hard_sl_price)
                    in_trade = False
                    if not hard_sl_hit and c_bull:
                        if (not HTF_FILTER_ENABLED or c_htf_bull) and in_trade_window(c_tod):
                            cnt = daily_trades.get(date_key, 0)
                            if not MAX_TRADES_PER_DAY_ENABLED or cnt < MAX_TRADES_PER_DAY:
                                entry_price   = c_close; entry_time = c_time
                                peak          = c_high;  direction  = "long"; in_trade = True
                                hard_sl_price = round(entry_price * (1 - HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
                                daily_trades[date_key] = cnt + 1
            continue

        # ── NOT IN TRADE — CHECK ENTRY ─────────────────────────────────────────
        # Filter 2: Trade window
        if not in_trade_window(c_tod):
            continue

        # Filter 3: Confirmation candle — must wait CONFIRM_CANDLES after flip
        if (i - flip_candle) < CONFIRM_CANDLES:
            continue

        # Filter 4: Min gap between price and ST line
        if MIN_GAP_ENABLED:
            gap_pct = abs(c_close - c_st) / c_close * 100
            if gap_pct < MIN_GAP_PCT:
                continue

        # Filter 5: Volume
        if VOLUME_FILTER_ENABLED and not np.isnan(vol_avg[i]) and vol_avg[i] > 0:
            if volumes[i] < VOLUME_MULTIPLIER * vol_avg[i]:
                continue

        # Filter 6: Max trades per day
        if MAX_TRADES_PER_DAY_ENABLED:
            cnt = daily_trades.get(date_key, 0)
            if cnt >= MAX_TRADES_PER_DAY:
                continue

        # Filter 1: HTF confirmation
        if c_bull and (not HTF_FILTER_ENABLED or c_htf_bull):
            entry_price   = c_close; entry_time = c_time
            peak          = c_high;  direction  = "long"; in_trade = True
            hard_sl_price = round(entry_price * (1 - HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
            daily_trades[date_key] = daily_trades.get(date_key, 0) + 1

        elif not c_bull and (not HTF_FILTER_ENABLED or not c_htf_bull):
            entry_price   = c_close; entry_time = c_time
            peak          = c_low;   direction  = "short"; in_trade = True
            hard_sl_price = round(entry_price * (1 + HARD_SL_PCT / 100), 4) if HARD_SL_ENABLED else None
            daily_trades[date_key] = daily_trades.get(date_key, 0) + 1

    return trades


# ───────────────────────────────────────────────────────────────────────────────
#  SUMMARIES
# ───────────────────────────────────────────────────────────────────────────────
def build_daily_summary(trades):
    if not trades:
        return pd.DataFrame()
    df_t = pd.DataFrame(trades)
    daily = []
    for date, grp in df_t.groupby("Date"):
        total = len(grp); wins = (grp["P&L %"] > 0).sum(); losses = total - wins
        pnl_sum = grp["P&L %"].sum()
        win_pts = grp.loc[grp["P&L %"] > 0, "Points Captured"].sum()
        los_pts = grp.loc[grp["P&L %"] <= 0, "Points Captured"].sum()
        daily.append({
            "Date"           : date,
            "Total Trades"   : total,
            "Longs"          : grp["Direction"].str.contains("Long").sum(),
            "Shorts"         : grp["Direction"].str.contains("Short").sum(),
            "Winners"        : wins,
            "Losers"         : losses,
            "Win Rate %"     : round(wins / total * 100, 1),
            "Total P&L %"    : round(pnl_sum, 4),
            "Best Trade %"   : round(grp["P&L %"].max(), 4),
            "Worst Trade %"  : round(grp["P&L %"].min(), 4),
            "Points Captured": round(win_pts, 2),
            "Points Lost"    : round(los_pts, 2),
            "Net Points"     : round(win_pts - los_pts, 2),
            "Day Result"     : "✅ Profit" if pnl_sum > 0 else "❌ Loss" if pnl_sum < 0 else "⚖ Flat",
        })
    return pd.DataFrame(daily)


def build_time_slot_summary(trades):
    """Break down P&L by intraday 30-min time slots."""
    if not trades:
        return pd.DataFrame()

    SLOTS = [
        ("09:15", "09:30"), ("09:30", "10:00"), ("10:00", "10:30"),
        ("10:30", "11:00"), ("11:00", "11:30"), ("11:30", "12:00"),
        ("12:00", "12:30"), ("12:30", "13:00"), ("13:00", "13:30"),
        ("13:30", "14:00"), ("14:00", "14:30"), ("14:30", "15:00"),
        ("15:00", "15:15"),
    ]

    df_t = pd.DataFrame(trades)
    df_t["Entry_dt"]  = pd.to_datetime(df_t["Entry Time"], format="%Y-%m-%d %H:%M")
    df_t["Entry_tod"] = df_t["Entry_dt"].dt.strftime("%H:%M")

    rows = []
    for s, e in SLOTS:
        mask = (df_t["Entry_tod"] >= s) & (df_t["Entry_tod"] < e)
        grp  = df_t[mask]
        if grp.empty:
            rows.append({"Time Slot": f"{s}–{e}", "Trades": 0, "Winners": 0,
                         "Losers": 0, "Win Rate %": "—", "Total P&L %": 0.0,
                         "Avg P&L %": 0.0, "Best %": 0.0, "Worst %": 0.0,
                         "Verdict": "—"})
            continue
        total = len(grp); wins = (grp["P&L %"] > 0).sum()
        pnl   = grp["P&L %"].sum()
        wr    = wins / total * 100
        rows.append({
            "Time Slot"  : f"{s}–{e}",
            "Trades"     : total,
            "Winners"    : wins,
            "Losers"     : total - wins,
            "Win Rate %" : round(wr, 1),
            "Total P&L %": round(pnl, 4),
            "Avg P&L %"  : round(grp["P&L %"].mean(), 4),
            "Best %"     : round(grp["P&L %"].max(), 4),
            "Worst %"    : round(grp["P&L %"].min(), 4),
            "Verdict"    : "🟢 Trade" if pnl > 0 and wr >= 50 else "🔴 Avoid",
        })
    return pd.DataFrame(rows)


# ───────────────────────────────────────────────────────────────────────────────
#  PRINT RESULTS
# ───────────────────────────────────────────────────────────────────────────────
def print_results(trades, ticker_label):
    sep  = "═" * 120
    dash = "─" * 120

    print("\n" + sep)
    print(f"  Supertrend({ST_PERIOD},{ST_MULTIPLIER})  |  {ticker_label}  |  EOD Exit: {MARKET_CLOSE}")
    print(f"  🔎 Filters: HTF={HTF_FILTER_ENABLED}({HTF_RESAMPLE})  "
          f"| Window={TRADE_WINDOW_ENABLED}  "
          f"| ConfirmCandles={CONFIRM_CANDLES}  "
          f"| MinGap={MIN_GAP_ENABLED}({MIN_GAP_PCT}%)  "
          f"| Volume={VOLUME_FILTER_ENABLED}  "
          f"| MaxTrades/Day={MAX_TRADES_PER_DAY if MAX_TRADES_PER_DAY_ENABLED else 'OFF'}")
    print(f"  🛑 Hard SL: {'ON — ' + str(HARD_SL_PCT) + '% from entry' if HARD_SL_ENABLED else 'OFF'}")
    print(sep)

    if not trades:
        print("  ⚠  No trades found. Adjust filters or time range.")
        print(sep)
        return None, None, None

    df_t   = pd.DataFrame(trades)
    df_day = build_daily_summary(trades)
    df_ts  = build_time_slot_summary(trades)

    # ── Trade Log ──────────────────────────────────────────────────────────────
    print("\n  TRADE LOG")
    print(dash)
    cols = ["Date","Direction","Entry Time","Entry Price","Hard SL Price",
            "ST Stop Loss","Exit Time","Exit Price","Points Captured",
            "P&L %","Exit Reason","Result"]
    print(df_t[cols].to_string(index=False))

    # ── Per-Day P&L ────────────────────────────────────────────────────────────
    print("\n\n" + sep)
    print("  PER-DAY P&L BREAKDOWN")
    print(dash)
    print(df_day.to_string(index=False))

    # ── Time Slot Analysis ─────────────────────────────────────────────────────
    print("\n\n" + sep)
    print("  TIME-SLOT P&L ANALYSIS  (Entry time grouped by 30-min windows)")
    print(dash)
    print(df_ts.to_string(index=False))

    # ── Overall Summary ────────────────────────────────────────────────────────
    total     = len(df_t)
    wins      = (df_t["P&L %"] > 0).sum()
    losses    = total - wins
    hard_exits = df_t["Exit Reason"].str.startswith("Hard SL").sum()
    sl_exits  = (df_t["Exit Reason"] == "ST Stop Loss").sum()
    st_exits  = df_t["Exit Reason"].str.startswith("ST Flip").sum()
    eod_exits = (df_t["Exit Reason"] == "EOD Exit").sum()
    win_pts   = df_t.loc[df_t["P&L %"] > 0, "Points Captured"].sum()
    loss_pts  = df_t.loc[df_t["P&L %"] <= 0, "Points Captured"].sum()
    profit_days = (df_day["Total P&L %"] > 0).sum()
    loss_days   = (df_day["Total P&L %"] < 0).sum()
    flat_days   = len(df_day) - profit_days - loss_days
    best_day    = df_day.loc[df_day["Total P&L %"].idxmax()]
    worst_day   = df_day.loc[df_day["Total P&L %"].idxmin()]

    # Best time slot
    best_slot = df_ts.loc[df_ts["Total P&L %"].idxmax()] if not df_ts.empty else None
    worst_slot= df_ts.loc[df_ts["Total P&L %"].idxmin()] if not df_ts.empty else None

    print("\n\n" + sep)
    print("  OVERALL SUMMARY")
    print(dash)
    print(f"  {'FILTERS ACTIVE':─<55}")
    print(f"  HTF Filter          : {'ON  (' + HTF_RESAMPLE + ' ST ' + str(HTF_ST_PERIOD) + ',' + str(HTF_ST_MULTIPLIER) + ')' if HTF_FILTER_ENABLED else 'OFF'}")
    print(f"  Trade Window        : {'ON  (' + str(len(TRADE_WINDOWS)) + ' windows)' if TRADE_WINDOW_ENABLED else 'OFF'}")
    print(f"  Confirm Candles     : {CONFIRM_CANDLES}")
    print(f"  Min Gap Filter      : {'ON  (' + str(MIN_GAP_PCT) + '%)' if MIN_GAP_ENABLED else 'OFF'}")
    print(f"  Volume Filter       : {'ON  (' + str(VOLUME_MULTIPLIER) + 'x)' if VOLUME_FILTER_ENABLED else 'OFF'}")
    print(f"  Max Trades/Day      : {MAX_TRADES_PER_DAY if MAX_TRADES_PER_DAY_ENABLED else 'OFF'}")
    print(f"  Hard Stop Loss      : {'ON  (' + str(HARD_SL_PCT) + '%)' if HARD_SL_ENABLED else 'OFF'}")
    print(dash)
    print(f"  {'TRADE STATS':─<55}")
    print(f"  Total Trades        : {total}")
    print(f"  Winners             : {wins}  ({wins/total*100:.1f}%)")
    print(f"  Losers              : {losses}  ({losses/total*100:.1f}%)")
    print(f"  Hard SL Exits       : {hard_exits}")
    print(f"  ST Stop Loss Exits  : {sl_exits}")
    print(f"  ST Flip Exits       : {st_exits}")
    print(f"  EOD Force Exits     : {eod_exits}")
    print(dash)
    print(f"  {'P&L STATS':─<55}")
    print(f"  Total P&L           : {df_t['P&L %'].sum():.4f}%")
    print(f"  Avg P&L / Trade     : {df_t['P&L %'].mean():.4f}%")
    print(f"  Best Single Trade   : {df_t['P&L %'].max():.4f}%")
    print(f"  Worst Single Trade  : {df_t['P&L %'].min():.4f}%")
    print(dash)
    print(f"  {'POINTS':─<55}")
    print(f"  Points Captured     : {win_pts:.2f}  (winning trades)")
    print(f"  Points Lost         : {loss_pts:.2f}  (losing trades)")
    print(f"  Net Points          : {win_pts - loss_pts:.2f}")
    print(dash)
    print(f"  {'DAY STATS':─<55}")
    print(f"  Total Days          : {len(df_day)}")
    print(f"  Profit Days         : {profit_days}  ({profit_days/len(df_day)*100:.1f}%)")
    print(f"  Loss Days           : {loss_days}  ({loss_days/len(df_day)*100:.1f}%)")
    print(f"  Flat Days           : {flat_days}")
    print(f"  Best Day            : {best_day['Date']}  →  {best_day['Total P&L %']:.4f}%  |  Net Pts: {best_day['Net Points']:.2f}")
    print(f"  Worst Day           : {worst_day['Date']}  →  {worst_day['Total P&L %']:.4f}%  |  Net Pts: {worst_day['Net Points']:.2f}")
    if best_slot is not None:
        print(dash)
        print(f"  {'TIME SLOT INSIGHTS':─<55}")
        print(f"  Best Time Slot      : {best_slot['Time Slot']}  →  {best_slot['Total P&L %']:.4f}%  |  WR: {best_slot['Win Rate %']}%")
        print(f"  Worst Time Slot     : {worst_slot['Time Slot']}  →  {worst_slot['Total P&L %']:.4f}%  |  WR: {worst_slot['Win Rate %']}%")
    print(sep + "\n")

    return df_t, df_day, df_ts


# ───────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ───────────────────────────────────────────────────────────────────────────────
def export_excel(df, trades, df_trades, df_day, df_ts, ticker_label):
    ticker_clean = ticker_label.replace("^","").replace(".","_").replace("/","_")
    fname = f"{ticker_clean}_supertrend_v2.xlsx"

    df_st = df[["Open","High","Low","Close","Volume","ST_val","ST_direction"]].copy()
    df_st.index = df_st.index.strftime("%Y-%m-%d %H:%M")
    df_st.index.name = "DateTime (IST)"
    df_st.columns = ["Open","High","Low","Close","Volume",
                     f"ST Value ({ST_PERIOD},{ST_MULTIPLIER})","Direction"]
    df_st = df_st.round(2)

    # Build summary
    if df_trades is not None and not df_trades.empty:
        total    = len(df_trades); wins = (df_trades["P&L %"] > 0).sum(); losses = total - wins
        win_pts  = df_trades.loc[df_trades["P&L %"] > 0, "Points Captured"].sum()
        loss_pts = df_trades.loc[df_trades["P&L %"] <= 0, "Points Captured"].sum()
        pd_ = int((df_day["Total P&L %"] > 0).sum()) if df_day is not None else 0
        ld_ = int((df_day["Total P&L %"] < 0).sum()) if df_day is not None else 0
        summary_rows = [
            ["STRATEGY INFO",    ""],
            ["Ticker/Source",    ticker_label],
            ["Data Source",      DATA_SOURCE.upper()],
            ["Supertrend",       f"Period={ST_PERIOD}, Multiplier={ST_MULTIPLIER}"],
            ["EOD Exit",         str(MARKET_CLOSE)],
            ["", ""],
            ["FILTERS",          ""],
            ["HTF Filter",       f"{'ON (' + HTF_RESAMPLE + ')' if HTF_FILTER_ENABLED else 'OFF'}"],
            ["Trade Window",     f"{'ON' if TRADE_WINDOW_ENABLED else 'OFF'}"],
            ["Confirm Candles",  CONFIRM_CANDLES],
            ["Min Gap Filter",   f"{'ON (' + str(MIN_GAP_PCT) + '%)' if MIN_GAP_ENABLED else 'OFF'}"],
            ["Volume Filter",    f"{'ON (' + str(VOLUME_MULTIPLIER) + 'x)' if VOLUME_FILTER_ENABLED else 'OFF'}"],
            ["Max Trades/Day",   MAX_TRADES_PER_DAY if MAX_TRADES_PER_DAY_ENABLED else "OFF"],
            ["Hard Stop Loss",   f"{'ON (' + str(HARD_SL_PCT) + '%)' if HARD_SL_ENABLED else 'OFF'}"],
            ["", ""],
            ["TRADE STATS",      ""],
            ["Total Trades",     total],
            ["Winners",          f"{wins} ({wins/total*100:.1f}%)"],
            ["Losers",           f"{losses} ({losses/total*100:.1f}%)"],
            ["Hard SL Exits",    int(df_trades["Exit Reason"].str.startswith("Hard SL").sum())],
            ["ST SL Exits",      int((df_trades["Exit Reason"] == "ST Stop Loss").sum())],
            ["ST Flip Exits",    int(df_trades["Exit Reason"].str.startswith("ST Flip").sum())],
            ["EOD Exits",        int((df_trades["Exit Reason"] == "EOD Exit").sum())],
            ["", ""],
            ["P&L STATS",        ""],
            ["Total P&L %",      round(df_trades["P&L %"].sum(), 4)],
            ["Avg P&L % / Trade",round(df_trades["P&L %"].mean(), 4)],
            ["Best Trade %",     round(df_trades["P&L %"].max(), 4)],
            ["Worst Trade %",    round(df_trades["P&L %"].min(), 4)],
            ["", ""],
            ["POINTS",           ""],
            ["Points Captured",  round(win_pts, 2)],
            ["Points Lost",      round(loss_pts, 2)],
            ["Net Points",       round(win_pts - loss_pts, 2)],
            ["", ""],
            ["DAY STATS",        ""],
            ["Total Days",       len(df_day) if df_day is not None else 0],
            ["Profit Days",      pd_],
            ["Loss Days",        ld_],
        ]
        if df_day is not None and not df_day.empty:
            best  = df_day.loc[df_day["Total P&L %"].idxmax()]
            worst = df_day.loc[df_day["Total P&L %"].idxmin()]
            summary_rows += [
                ["Best Day",  f"{best['Date']}  →  {best['Total P&L %']:.4f}%"],
                ["Worst Day", f"{worst['Date']}  →  {worst['Total P&L %']:.4f}%"],
            ]
        df_summary = pd.DataFrame(summary_rows, columns=["Metric","Value"])
    else:
        df_summary = pd.DataFrame([{"Metric":"No trades found","Value":""}])

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        df_st.to_excel(writer, sheet_name="Supertrend Data", index=True)
        if df_trades is not None and not df_trades.empty:
            df_trades.to_excel(writer, sheet_name="Trade Log", index=False)
        if df_day is not None and not df_day.empty:
            df_day.to_excel(writer, sheet_name="Daily P&L", index=False)
        if df_ts is not None and not df_ts.empty:
            df_ts.to_excel(writer, sheet_name="Time Slot P&L", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    _style_excel(fname, df_trades, df_day, df_ts)
    print(f"  Excel saved  → {fname}")
    return fname


def _style_excel(fname, df_trades, df_day, df_ts):
    GREEN_DARK   = "1A5C38"; GREEN_LIGHT = "C6EFCE"
    RED_DARK     = "9C0006"; RED_LIGHT   = "FFC7CE"
    BLUE_HDR     = "1F3864"; YELLOW_SEC  = "FFD700"
    GRAY_ALT     = "F2F2F2"; WHITE       = "FFFFFF"
    ORANGE       = "FCE4D6"; TEAL_LIGHT  = "DDEBF7"
    PURPLE_LIGHT = "E2EFDA"

    thin_s = Side(style="thin", color="CCCCCC")
    bdr    = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    def hdr(ws, row=1, bg=BLUE_HDR, fg=WHITE):
        for cell in ws[row]:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(bold=True, color=fg, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = bdr
        ws.row_dimensions[row].height = 22

    def autowidth(ws, cap=28):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, cap)

    wb = load_workbook(fname)

    # Sheet 1: Supertrend Data
    ws = wb["Supertrend Data"]
    hdr(ws); ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        val = str(row[-1].value or "")
        bg  = GREEN_LIGHT if val == "Bullish" else RED_LIGHT if val == "Bearish" else WHITE
        fg_ = GREEN_DARK  if val == "Bullish" else RED_DARK  if val == "Bearish" else "000000"
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=bg); cell.border = bdr
            cell.alignment = Alignment(horizontal="center")
        row[-1].font = Font(bold=True, color=fg_)
    autowidth(ws)

    # Sheet 2: Trade Log
    if "Trade Log" in wb.sheetnames and df_trades is not None:
        ws = wb["Trade Log"]; hdr(ws); ws.freeze_panes = "A2"
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            result = str(row[-1].value or ""); reason = str(row[-2].value or "")
            bg = GREEN_LIGHT if result == "WIN" else RED_LIGHT if result == "LOSS" else (GRAY_ALT if i%2==0 else WHITE)
            if result == "WIN":   row[-1].font = Font(bold=True, color=GREEN_DARK)
            elif result == "LOSS":row[-1].font = Font(bold=True, color=RED_DARK)
            if reason.startswith("Hard SL"):
                row[-2].fill = PatternFill("solid", fgColor="FF0000"); row[-2].font = Font(bold=True, color="FFFFFF")
            elif "Stop Loss" in reason:
                row[-2].fill = PatternFill("solid", fgColor=ORANGE); row[-2].font = Font(bold=True, color="C55A11")
            elif "EOD" in reason:
                row[-2].fill = PatternFill("solid", fgColor=TEAL_LIGHT); row[-2].font = Font(bold=True, color="2E75B6")
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000","FFFFFFFF"):
                    cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = bdr; cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    # Sheet 3: Daily P&L
    if "Daily P&L" in wb.sheetnames and df_day is not None:
        ws = wb["Daily P&L"]; hdr(ws); ws.freeze_panes = "A2"
        rc = ws.max_column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            dr = str(row[rc-1].value or "")
            bg = GREEN_LIGHT if "Profit" in dr else RED_LIGHT if "Loss" in dr else GRAY_ALT
            if "Profit" in dr: row[rc-1].font = Font(bold=True, color=GREEN_DARK)
            elif "Loss"  in dr:row[rc-1].font = Font(bold=True, color=RED_DARK)
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=bg); cell.border = bdr
                cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    # Sheet 4: Time Slot P&L
    if "Time Slot P&L" in wb.sheetnames and df_ts is not None:
        ws = wb["Time Slot P&L"]; hdr(ws, bg="2E4057"); ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            verdict = str(row[-1].value or "")
            pnl_cell = row[5]  # "Total P&L %" column
            try:
                pnl_val = float(pnl_cell.value) if pnl_cell.value not in (None, "—") else 0
            except (ValueError, TypeError):
                pnl_val = 0
            bg = GREEN_LIGHT if pnl_val > 0 else RED_LIGHT if pnl_val < 0 else GRAY_ALT
            if "🟢" in verdict: row[-1].font = Font(bold=True, color=GREEN_DARK)
            elif "🔴" in verdict: row[-1].font = Font(bold=True, color=RED_DARK)
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=bg); cell.border = bdr
                cell.alignment = Alignment(horizontal="center")
        autowidth(ws)

    # Sheet 5: Summary
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]; hdr(ws)
        SECTIONS = {"STRATEGY INFO","FILTERS","TRADE STATS","P&L STATS","POINTS","DAY STATS"}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            label = str(row[0].value or "")
            if label in SECTIONS:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=YELLOW_SEC)
                    cell.font = Font(bold=True, color="000000", size=10); cell.border = bdr
                continue
            val_cell = row[1] if len(row) > 1 else None
            for cell in row:
                cell.border = bdr; cell.alignment = Alignment(horizontal="left")
            if val_cell and "Net Points" in label:
                try: v = float(val_cell.value)
                except: v = 0
                val_cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT if v > 0 else RED_LIGHT)
                val_cell.font = Font(bold=True, color=GREEN_DARK if v > 0 else RED_DARK)
        ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 28
    wb.save(fname)


# ───────────────────────────────────────────────────────────────────────────────
#  CHART
# ───────────────────────────────────────────────────────────────────────────────
def build_chart(df, trades, ticker_label):
    if CHART_DAYS:
        unique_days = sorted(df.index.normalize().unique())
        cutoff = unique_days[-CHART_DAYS] if len(unique_days) >= CHART_DAYS else unique_days[0]
        df_c   = df[df.index.normalize() >= cutoff].copy()
    else:
        df_c = df.copy()

    st_green = df_c["ST_val"].where(df_c["ST_bull"])
    st_red   = df_c["ST_val"].where(~df_c["ST_bull"])

    fig = go.Figure()
    fig.add_trace(go.Candlestick(
        x=df_c.index, open=df_c["Open"], high=df_c["High"],
        low=df_c["Low"], close=df_c["Close"], name="Price",
        increasing_line_color="#26a69a", decreasing_line_color="#ef5350"))
    fig.add_trace(go.Scatter(x=df_c.index, y=st_green, name="ST Bullish",
        mode="lines", line=dict(color="#22c55e", width=2.5), connectgaps=False))
    fig.add_trace(go.Scatter(x=df_c.index, y=st_red, name="ST Bearish",
        mode="lines", line=dict(color="#ef4444", width=2.5), connectgaps=False))

    # HTF ST band (if available)
    if "HTF_bull" in df_c.columns:
        htf_mid = df_c["ST_val"].rolling(3).mean()
        htf_bull_line = htf_mid.where(df_c["HTF_bull"])
        htf_bear_line = htf_mid.where(~df_c["HTF_bull"])
        fig.add_trace(go.Scatter(x=df_c.index, y=htf_bull_line,
            name=f"HTF ({HTF_RESAMPLE}) Bullish", mode="lines",
            line=dict(color="#86efac", width=1, dash="dot"), connectgaps=False))
        fig.add_trace(go.Scatter(x=df_c.index, y=htf_bear_line,
            name=f"HTF ({HTF_RESAMPLE}) Bearish", mode="lines",
            line=dict(color="#fca5a5", width=1, dash="dot"), connectgaps=False))

    if trades:
        df_t  = pd.DataFrame(trades)
        in_win = df_c.index[0]
        def parse(col): return pd.to_datetime(df_t[col], format="%Y-%m-%d %H:%M").dt.tz_localize(IST)
        et = parse("Entry Time"); xt = parse("Exit Time")
        ep = df_t["Entry Price"]; xp = df_t["Exit Price"]
        dr = df_t["Direction"];   rs = df_t["Result"]; ex = df_t["Exit Reason"]
        mask = et >= in_win
        for x_val, y_val, name, sym, col, edgecol in [
            (et[mask & dr.str.contains("Long")].tolist(),  ep[mask & dr.str.contains("Long")].tolist(),  "Long Entry",  "triangle-up",   "#22c55e", "white"),
            (et[mask & dr.str.contains("Short")].tolist(), ep[mask & dr.str.contains("Short")].tolist(), "Short Entry", "triangle-down", "#ef4444", "white"),
            (xt[mask & (rs=="WIN")].tolist(), xp[mask & (rs=="WIN")].tolist(), "Exit WIN",  "circle",   "#86efac", "#16a34a"),
            (xt[mask & (rs=="LOSS")].tolist(),xp[mask & (rs=="LOSS")].tolist(),"Exit LOSS", "x",        "#fca5a5", "#dc2626"),
            (xt[mask & ex.str.startswith("Hard SL")].tolist(), xp[mask & ex.str.startswith("Hard SL")].tolist(), f"Hard SL", "hexagram", "#ff0000","white"),
        ]:
            if x_val:
                fig.add_trace(go.Scatter(x=x_val, y=y_val, mode="markers", name=name,
                    marker=dict(symbol=sym, size=13, color=col, line=dict(color=edgecol, width=1.5))))

    filters_txt = (f"HTF={HTF_RESAMPLE if HTF_FILTER_ENABLED else 'OFF'} | "
                   f"Window={'ON' if TRADE_WINDOW_ENABLED else 'OFF'} | "
                   f"ConfirmN={CONFIRM_CANDLES} | HardSL={HARD_SL_PCT}%")
    fig.update_layout(
        template="plotly_dark", height=800,
        xaxis_rangeslider_visible=False,
        legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="right", x=1, font=dict(size=9)),
        margin=dict(l=60, r=40, t=110, b=50), font=dict(family="monospace", size=11),
        title=dict(
            text=(f"<b>{ticker_label}</b> | Supertrend({ST_PERIOD},{ST_MULTIPLIER})<br>"
                  f"<span style='font-size:10px;color:#94a3b8'>{filters_txt}</span>"),
            x=0.5, xanchor="center"))
    fig.update_yaxes(title_text="Price", showgrid=True, gridcolor="#1e293b")
    fig.update_xaxes(showgrid=True, gridcolor="#1e293b",
        rangebreaks=[dict(bounds=["sat","mon"]), dict(bounds=[15.5,9.25], pattern="hour")])

    chart_name = f"{ticker_label.replace('^','').replace('.','_').replace('/','_')}_chart_v2.html"
    fig.write_html(chart_name)
    print(f"  Chart saved  → {chart_name}  (open in browser)")
    return chart_name


# ───────────────────────────────────────────────────────────────────────────────
#  MAIN
# ───────────────────────────────────────────────────────────────────────────────
def main():
    try:
        df, ticker_label = fetch_data()
    except (ValueError, FileNotFoundError) as e:
        print(e); return

    trades                  = run_backtest(df)
    df_trades, df_day, df_ts = print_results(trades, ticker_label)

    print("  Exporting Excel …")
    export_excel(df, trades, df_trades, df_day, df_ts, ticker_label)

    print("  Building chart …")
    build_chart(df, trades, ticker_label)

    print("\n  ✅  Done.\n")


if __name__ == "__main__":
    main()